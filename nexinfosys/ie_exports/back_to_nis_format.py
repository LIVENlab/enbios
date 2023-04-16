"""
From memory model, elaborate a spreadsheet in NIS format that generates the same memory model ("idempotent spreadsheet")
"""
import pandas as pd
import toposort
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook

from nexinfosys.command_generators.parser_ast_evaluators import ast_evaluator
from nexinfosys.command_generators.parser_field_parsers import string_to_ast, expression_with_parameters
from nexinfosys.common.helper import PartialRetrievalDictionary, create_dictionary, ifnull
from nexinfosys.model_services import State, get_case_study_registry_objects
from nexinfosys.models.musiasem_concepts import Parameter, List, Hierarchy, HierarchyNode, FactorType, Processor, \
    ProcessorsRelationPartOfObservation, Factor, FactorsRelationScaleObservation, \
    FactorsRelationDirectedFlowObservation, FactorTypesRelationUnidirectionalLinearTransformObservation, \
    ProblemStatement, Taxon


def nis_format_spreadsheet(s: State):
    """

    :param s:
    :return:
    """
    lst = []
    # Obtain principal structures
    glb_idx, p_sets, hierarchies, datasets, mappings = get_case_study_registry_objects(s)
    # (ProcessorScalings, DatasetQry, ListOfCommands, ImportCommands are lost)
    # Metadata
    lst.append(("Metadata", get_metadata(s)))
    # CodeHierarchies
    lst.append(("CodeHierarchies", get_code_hierarchies(glb_idx)))
    # CodeHierarchiesMapping
    # Parameters
    lst.append(("Parameters", get_parameters(glb_idx)))
    # DatasetDef (only imports)
    # DatasetData

    # RefProvenance
    # RefGeographic
    # RefBibliographic
    # InterfaceTypes
    lst.append(("InterfaceTypes", get_interface_types(glb_idx)))
    # ScaleChangeMap
    lst.append(("ScaleChangeMap", get_scale_change_map(glb_idx)))
    # Processors
    lst.append(("BareProcessors", get_processors(glb_idx)))
    # Interfaces
    lst.append(("Interfaces", get_interfaces(glb_idx)))
    # Relationships (scales)
    lst.append(("Relationships Scales", get_scale_relationships(glb_idx)))
    # Relationships (flows)
    lst.append(("Relationships Exchanges", get_exchange_relationships(glb_idx)))
    # ProblemStatement
    lst.append(("ProblemStatement", get_problem_statement(glb_idx)))
    # ScalarBenchmarks
    # ScalarIndicators
    # MatrixIndicators

    # Convert list of pd.DataFrames to Excel workbook
    wb = Workbook(write_only=True)
    for name, df in lst:
        if df.shape[0] < 2:
            continue

        ws = wb.create_sheet(name)
        widths = [0]*(df.shape[1]+1)  # A maximum of 100 columns
        max_columns = 0
        for r in dataframe_to_rows(df, index=False, header=True):
            if len(r) > max_columns:
                max_columns = len(r)
            for i in range(len(r)):
                width = int(len(str(r[i])) * 1.1)
                if width > widths[i]:
                    widths[i] = width

        for i, column_width in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    return save_virtual_workbook(wb)


def list_to_dataframe(lst: List) -> pd.DataFrame:
    return pd.DataFrame(data=lst[1:], columns=lst[0])


def get_metadata(s: State) -> pd.DataFrame:
    metadata_dictionary = s.get("_metadata")
    lst = list()
    lst.append(("Case study code", metadata_dictionary.get("case_study_code", "") if metadata_dictionary else ""))
    lst.append(("Case study name", metadata_dictionary.get("case_study_name", "") if metadata_dictionary else ""))
    lst.append(("Title", metadata_dictionary.get("title", "") if metadata_dictionary else ""))
    lst.append(("Subject, topic and/or keywords", metadata_dictionary.get("subject_topic_keywords", "") if metadata_dictionary else ""))
    lst.append(("Description", metadata_dictionary.get("description", "") if metadata_dictionary else ""))
    lst.append(("Geographical level", metadata_dictionary.get("geographical_level", "") if metadata_dictionary else ""))  # A list
    lst.append(("Dimensions", metadata_dictionary.get("dimensions", "") if metadata_dictionary else "")) # A list
    lst.append(("Reference documentation", metadata_dictionary.get("reference_documentation", "") if metadata_dictionary else ""))
    lst.append(("Authors", metadata_dictionary.get("authors", "") if metadata_dictionary else ""))
    lst.append(("Date of elaboration", metadata_dictionary.get("date_of_elaboration", "") if metadata_dictionary else ""))
    lst.append(("Temporal situation", metadata_dictionary.get("temporal_situation", "") if metadata_dictionary else ""))
    lst.append(("Geographical location", metadata_dictionary.get("geographical_situation", "") if metadata_dictionary else ""))
    if metadata_dictionary:
        lst.append(("DOI", metadata_dictionary.get("doi", "")[0] if metadata_dictionary.get("doi") else ""))
    else:
        lst.append(("DOI", ""))
    lst.append(("Language", metadata_dictionary.get("language", "") if metadata_dictionary else ""))
    lst.append(("Restriction level", metadata_dictionary.get("restriction_level", "") if metadata_dictionary else ""))  # A list
    if metadata_dictionary:
        lst.append(("Version", metadata_dictionary.get("version", "")[0] if metadata_dictionary.get("version") else ""))
    else:
        lst.append(("Version", ""))

    # Expand lists
    for i, t in enumerate(lst):
        if isinstance(t[1], list):
            tmp = [t[0]]
            tmp.extend(t[1])
            lst[i] = tuple(tmp)

    return list_to_dataframe(lst)


def get_parameters(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    lst = [["Parameter", "Type", "Domain", "Value", "Group", "Description", "Attributes"]]
    for p in glb_idx.get(Parameter.partial_key()):
        lst.append([p.name, p.type, p._range if p._range else "", ifnull(p.default_value, ""), ifnull(p.group, ""), ifnull(p._description, ""), ""])

    return list_to_dataframe(lst)


def get_code_hierarchies(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    def hierarchy_to_list(nodes: List[HierarchyNode]) -> List[HierarchyNode]:
        nodes_list = []
        if nodes:
            for node in sorted(nodes, key=lambda n: n.name):
                nodes_list.append(node)
                nodes_list.extend(hierarchy_to_list(list(node.get_children())))
        return nodes_list

    lst = [["Source", "HierarchyGroup", "Hierarchy", "Level", "ReferredHierarchy", "Code", "ParentCode", "Label", "Description", "Expression", "GeolocationRef", "GeolocationCode", "Attributes"]]
    for hh in glb_idx.get(Hierarchy.partial_key()):
        if hh.hierarchy_type == Taxon:
            for c in hierarchy_to_list(hh.roots):
                lst.append(["", "", c.hierarchy.name, ifnull(c.level, ""), c.referred_node.hierarchy.name if c.referred_node else "", c.name, c.parent.name if c.parent else "", c.label, c.description, "", "", "", ""])

    return list_to_dataframe(lst)


def get_interface_types(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    lst = [["InterfaceTypeHierarchy", "InterfaceType", "Sphere", "RoegenType", "ParentInterfaceType", "Formula",
            "Description", "Unit", "OppositeSubsystemType", "Attributes"]]
    for itype in glb_idx.get(FactorType.partial_key()):
        hierarchy = itype.hierarchy.name if itype.hierarchy else ""
        roegen_type = itype.roegen_type.name if itype.roegen_type else ""
        parent = itype.parent.name if itype.parent else ""
        lst.append([hierarchy, itype.name, itype.sphere, roegen_type, parent, "",
                    itype.attributes.get("description", ""), itype.attributes.get("unit", ""),
                    itype._opposite_processor_type, ""])
    return list_to_dataframe(lst)


def get_scale_change_map(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    lst = [["OriginHierarchy", "OriginInterfaceType", "DestinationHierarchy", "DestinationInterfaceType", "OriginContext", "DestinationContext", "Scale", "OriginUnit", "DestinationUnit"]]
    for sc in glb_idx.get(FactorTypesRelationUnidirectionalLinearTransformObservation.partial_key()):
        lst.append([sc.origin.hierarchy.name if sc.origin.hierarchy else "", sc.origin.name,
                    sc.destination.hierarchy.name if sc.destination.hierarchy else "", sc.destination.name,
                    sc._origin_context.name if sc._origin_context else "",
                    sc._destination_context.name if sc._destination_context else "",
                    sc._weight, sc._origin_unit, sc._destination_unit])

    return list_to_dataframe(lst)


def get_processors(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    procs = set(glb_idx.get(Processor.partial_key()))  # Unique processors
    d = {}
    for p in procs:
        parent_relations = glb_idx.get(ProcessorsRelationPartOfObservation.partial_key(child=p))
        d[p.ident] = set([p.parent_processor.ident for p in parent_relations])

    lst = [["ProcessorGroup", "Processor", "ParentProcessor", "SubsystemType", "System", "FunctionalOrStructural", "Accounted", "Stock", "Description", "GeolocationRef", "GeolocationCode", "GeolocationLatLong", "Attributes"]]
    # Elaborate a DAG, then iterate over it
    for ident in list(toposort.toposort_flatten(d)):
        p = glb_idx.get(Processor.partial_key(ident=ident))[0]
        gref = p.geolocation.reference if p.geolocation else ""
        gcode = p.geolocation.code if p.geolocation else ""
        pgroup = p.attributes.get("processor_group")
        fname = p.full_hierarchy_names(glb_idx)[0]
        t = [pgroup if pgroup else "", p.name, "", p.attributes.get("subsystem_type", ""), p.attributes.get("processor_system", ""), p.attributes.get("functional_or_structural", ""), p.attributes.get("instance_or_archetype", ""), p.attributes.get("stock"), "", gref, gcode, "", ""]
        parent_relations = glb_idx.get(ProcessorsRelationPartOfObservation.partial_key(child=p))
        if len(parent_relations) == 0:
            lst.append(t)
        else:
            first = True
            for rel in parent_relations:
                t[2] = rel.parent_processor.full_hierarchy_names(glb_idx)[0]
                lst.append(t.copy())
                if first:
                    first = False
                    proper = None
                    for n in p.full_hierarchy_names(glb_idx):
                        if t[2] in n:
                            proper = n
                            break
                    t[1] = proper if proper else t[1]

    return list_to_dataframe(lst)


def get_interfaces(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    # Used to examine "value" as expression, and find variables that are interface names vs parameter names
    params = create_dictionary(data={p.name: None for p in glb_idx.get(Parameter.partial_key())})
    s = State()
    procs = glb_idx.get(Processor.partial_key())
    d = {}
    for p in procs:
        parent_relations = glb_idx.get(ProcessorsRelationPartOfObservation.partial_key(child=p))
        d[p.ident] = set([p.parent_processor.ident for p in parent_relations])

    lst = [["Processor", "InterfaceType", "Interface", "Sphere", "RoegenType", "Orientation", "OppositeSubsystemType", "GeolocationRef", "GeolocationCode", "InterfaceAttributes", "Value", "Unit", "RelativeTo", "Uncertainty", "Assessment", "PedigreeMatrix", "Pedigree", "Time", "Source", "NumberAttributes", "Comments"]]
    # Elaborate a DAG, then iterate over it
    for ident in list(toposort.toposort_flatten(d)):
        p = glb_idx.get(Processor.partial_key(ident=ident))[0]
        ifaces = glb_idx.get((Factor.partial_key(processor=p)))
        iface_names = create_dictionary(data={iface.name: iface for iface in ifaces})
        # Elaborate DAG of Interfaces because of Observations
        d = {}
        for iface in ifaces:
            if iface.ident not in d:
                d[iface.ident] = set()
            for obs in iface.quantitative_observations:
                if obs.relative_factor:
                    d[iface.ident].add(obs.relative_factor.ident)
                # Consider obs.value and non linear dependencies
                if isinstance(obs.value, str):
                    try:
                        value = float(obs.value)
                        unresolved_vars = []
                    except ValueError:
                        ast = string_to_ast(expression_with_parameters, obs.value)
                        evaluation_issues = []
                        value, unresolved_vars = ast_evaluator(exp=ast, state=s, obj=None, issue_lst=evaluation_issues)
                        for unresolved in unresolved_vars:
                            if unresolved not in params:
                                d[iface.ident].add(iface_names[unresolved].ident)

        for ident2 in list(toposort.toposort_flatten(d)):
            iface = glb_idx.get(Factor.partial_key(ident=ident2))[0]
            lst1 = [iface.processor.name, iface.taxon.name, iface.name, iface.sphere, iface.roegen_type.name,
                    iface.orientation, iface.opposite_processor_type, "", "", ""]
            observations = iface.quantitative_observations
            if len(observations) > 0:
                for obs in observations:
                    lst2 = [obs.value, obs.attributes.get("unit", ""), obs.relative_factor.name if obs.relative_factor else "", obs.attributes.get("spread", ""), obs.attributes.get("assessment", ""), obs.attributes.get("pedigree_template", ""), obs.attributes.get("pedigree", ""), obs.attributes.get("time", ""), obs.observer.name if obs.observer else "", "", obs.attributes.get("comments", "")]
                    lst.append(lst1+lst2)
            else:
                lst.append(lst1+["", "", "", "", "", "", "", "", "", ""])

    return list_to_dataframe(lst)


def get_scale_relationships(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    lst = [["OriginProcessors", "OriginInterface", "DestinationProcessors", "DestinationInterface", "BackInterface", "RelationType", "Weight", "ChangeOfTypeScale", "OriginCardinality", "DestinationCardinality", "Attributes"]]
    for rel in glb_idx.get(FactorsRelationScaleObservation.partial_key()):
        lst.append([rel.origin.processor.name, rel.origin.name, rel.destination.processor.name, rel.destination.name, "", "Scale", rel.quantity, "", "", "", ""])

    return list_to_dataframe(lst)


def get_exchange_relationships(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    lst = [["OriginProcessors", "OriginInterface", "DestinationProcessors", "DestinationInterface", "BackInterface", "RelationType", "Weight", "ChangeOfTypeScale", "OriginCardinality", "DestinationCardinality", "Attributes"]]
    for rel in glb_idx.get(FactorsRelationDirectedFlowObservation.partial_key()):
        lst.append([rel.source_factor.processor.name, rel.source_factor.name, rel.target_factor.processor.name, rel.target_factor.name, rel.back_factor.name if rel.back_factor else "", ">", rel.weight, "", "", "", ""])

    return list_to_dataframe(lst)


def get_problem_statement(glb_idx: PartialRetrievalDictionary) -> pd.DataFrame:
    ps = glb_idx.get(ProblemStatement.partial_key())
    if len(ps) == 0:
        ps = [ProblemStatement()]

    lst = [["Scenario", "Parameter", "Value", "Description"]]
    for solving_parameter in ps[0].solving_parameters.items():
        lst.append(["", solving_parameter[0], solving_parameter[1], ""])

    for scenario, params in ps[0].scenarios.items():
        for k, v in params.items():
            lst.append([scenario, k, v, ""])

    return list_to_dataframe(lst)


