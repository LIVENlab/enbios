import logging

import numpy as np
from copy import copy
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from openpyxl.worksheet.worksheet import Worksheet
#from openpyxl.worksheet import Worksheet

from openpyxl.worksheet.copier import WorksheetCopy

global_fill = PatternFill("none")

# #################################### #
#  Worksheet related helper functions  #
# #################################### #


def worksheet_to_numpy_array(sh_in):
    """
    Obtain a replica of the worksheet into a Numpy NDArray, with combined cells (combined cells are repeated)

    :param sh_in:
    :return: The numpy array with the values of the worksheet
    """
    max_col = sh_in.max_column
    max_row = sh_in.max_row
    m = np.zeros((max_row, max_col)).astype(object)
    for r in range(max_row):
        for c in range(max_col):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                m[r, c] = v
            else:
                m[r, c] = 0.0

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    max_col = sh_in.max_column
    max_row = sh_in.max_row
    m = np.zeros((max_row, max_col), dtype=bool)
    for r in range(max_row):
        for c in range(max_col):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(str(ra))  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def obtain_rectangular_submatrices(mask, region=None, only_remove_empty_bottom=False):
    """
    Obtain rectangular submatrices of mask
    IMPORTANT: currently it only obtains ONE region

    :param mask: The original matrix, numpy.NDArray, containing only 0/1 (1 is "some content")
    :param region: A tuple (top, bottom, left, right) with indices to search. bottom and right are not included
    :return: The list of rectangular regions as tuples (top, bottom, left, right)
    """

    def nonzero_sequences(a):
        # Create an array that is 1 where a is non-zero, and pad each end with an extra 0.
        isnonzero = np.concatenate(([0], a != 0, [0]))
        absdiff = np.abs(np.diff(isnonzero))
        # Runs start and end where absdiff is 1.
        ranges = np.where(absdiff == 1)[0].reshape(-1, 2)
        return ranges

    lst = []
    if not region:
        region = (0, mask.shape[0], 0, mask.shape[1])  # All the mask
    submask = mask[region[0]:region[1], region[2]:region[3]]
    offset_col, offset_row = (region[2], region[0])
    # Accumulation of elements by row (resulting in a column vector)
    row_sum = np.sum(submask, axis=1)
    # Accumulation of elements by column (resulting in a row vector)
    col_sum = np.sum(submask, axis=0)

    # Ranges
    rs = nonzero_sequences(row_sum.flatten())
    cs = nonzero_sequences(col_sum.flatten())
    if only_remove_empty_bottom:
        if len(rs) > 0:
            lst.append((rs[0][0], rs[-1][1], cs[0][0], cs[-1][1]))
    else:
        # Take the first rectangle
        if len(rs) > 0:
            lst.append((rs[0][0], rs[0][1], cs[0][0], cs[0][1]))

    return lst


def reset_cell_format(sh_writable, r, c):
    """
    When writing, reset cell's format

    :param sh_writable: Output worksheet
    :param r: Row number
    :param c: Col number
    :return:
    """
    cell = sh_writable.cell(row=r, column=c)
    cell.fill = global_fill
    cell.comment = None


def reset_cells_format(sh_writable):
    """
    When writing, reset all worksheet cells format

    :param sh_writable: Output worksheet
    :return:
    """
    max_col = sh_writable.max_column
    max_row = sh_writable.max_row
    for r in range(max_row):
        for c in range(max_col):
            reset_cell_format(sh_writable, r + 1, c + 1)


def cell_content_to_str(v):
    """
    Convert the value of a cell to string

    :param v: Value of a cell
    :return:
    """
    if v:
        if isinstance(v, float) or isinstance(v, int):
            return str(int(v))
        else:
            return str(v).strip()
    else:
        return None


def show_message(sh, r, c, message, type="error", accumulate=True):
    """
    It serves to show a cell in a worksheet
    It shows some type of error (warning or error) with a message in a comment
    The name of the sheet is changed with a prefix indicating there is at least an issue to be solved

    :param sh:
    :param r:
    :param c:
    :param message:
    :param type:
    :return:
    """
    cell = sh.cell(row=r, column=c)
    fill = cell.fill
    if type == "error":
        fill = PatternFill("solid", fgColor="CC0000")
    elif type == "warning":
        fill = PatternFill("solid", fgColor="FFFF33")
    elif type == "info":
        fill = PatternFill("solid", fgColor="87CEEB")
    cell.fill = fill
    if accumulate:
        comment = cell.comment
        if comment:
            comment.text += "\n" + message
        else:
            comment = Comment(message, "NIS")
    else:
        comment = Comment(message, "NIS")
    cell.comment = comment
    # if type == "error":
    #     sh.title = "!" + sh.title


class WorksheetCopy2(object):
    """
    Copy the values, styles, dimensions and merged cells from one worksheet
    to another within the same workbook.

    Adapted from "WorksheetCopy" of OpenPyXL. To cope with issue regarding a badly specified input ColumnDimension
    """

    def __init__(self, source_worksheet, target_worksheet, copy_style=True):
        self.source = source_worksheet
        self.target = target_worksheet
        self._copy_style = copy_style
        self._verify_resources()

    def _verify_resources(self):
        if not isinstance(self.source, Worksheet) and not isinstance(self.target, Worksheet):
            raise TypeError("Can only copy worksheets")

        if self.source is self.target:
            raise ValueError("Cannot copy a worksheet to itself")

        if self.source.parent != self.target.parent:
            raise ValueError('Cannot copy between worksheets from different workbooks')

    def copy_worksheet(self):
        self._copy_cells()
        self._copy_dimensions()

        if self._copy_style:
            self.target.sheet_format = copy(self.source.sheet_format)
        if hasattr(self.source, "_merged_cells"):
            self.target._merged_cells = copy(self.source._merged_cells)
        elif hasattr(self.source, "merged_cells"):
            self.target.merged_cells = copy(self.source.merged_cells)
        # else:
        #     self.target._merged_cells = None
        self.target.sheet_properties = copy(self.source.sheet_properties)

    def _copy_cells(self):
        for (row, col), source_cell in self.source._cells.items():
            target_cell = self.target.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style and self._copy_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    def _copy_dimensions(self):
        for attr in ('row_dimensions', 'column_dimensions'):
            src = getattr(self.source, attr)
            target = getattr(self.target, attr)
            for key, dim in src.items():
                # TODO COPY only if the dimension is correct. From Kendo UI ticket
                if attr == 'column_dimensions' and dim.min < 1021:
                    target[key] = copy(dim)
                    target[key].worksheet = self.target


def rewrite_xlsx_file(xl, copy_style=True):
    """
    Regenerates the worksheets of the input file. The aim is to calculate correctly the "dimension@ref" attribute of
    each of the Worksheets, in order to have it correctly processed by the Kendo UI Spreadsheet

    :param xl: A Workbook object, constructed with OpenPyXL
    :return: Nothing, the "xl" object is modified *inplace*
    """
    def copy_worksheet(xl, xf):
        new_title = u"{0} Copy".format(xf.title)
        to_worksheet = xl.create_sheet(title=new_title)
        cp = WorksheetCopy2(source_worksheet=xf, target_worksheet=to_worksheet, copy_style=copy_style)
        cp.copy_worksheet()
        return to_worksheet

    sn = xl.sheetnames
    for c, sh_name in enumerate(sn):
        source = xl.get_sheet_by_name(sh_name)
        copy_sheet = source.sheet_state != "hidden"
        if copy_sheet:
            tmp = copy_worksheet(xl, source)
        else:
            logging.debug(f"Skipping copy of '{sh_name}' worksheet")
        xl.remove_sheet(source)
        if copy_sheet:
            tmp.title = sh_name
