from typing import List, Tuple, Optional, Dict

from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import CommandField, IssuesLabelContentTripleType, AreaTupleType
from nexinfosys.command_generators import Issue, parser_field_parsers, IssueLocation, IType
from nexinfosys.command_generators.parser_field_parsers import simple_h_name, arith_boolean_expression, unquoted_string


def check_columns(sh, name: str, area: Tuple, cols: List[CommandField], command_name: str, ignore_not_found=True):
    """
    When parsing of a command starts, check columns
    Try to match each column with declared column fields. If a column is not declared, raise an error (or ignore it)
    If mandatory columns are not found, raise an error

    :param sh: The worksheet being analyzed
    :param name: The name of the worksheet
    :param area: Area inside the worksheet that will be scanned
    :param cols: List of CommandField
    :param command_name: A string with the name of the command
    :param ignore_not_found: True if a column not matching declared ones has to be ignored, False if an error has to be raised in this case
    :return: The map column name to column index (or indices for multiply declared columns); The issues found
    """

    issues: List[Issue] = []

    # Set of mandatory columns
    mandatory_not_found = set([c.name for c in cols if c.mandatory])

    # Check columns
    col_map = {}  # From CommandField to a list of tuples (column, index)
    for c in range(area[2], area[3]):  # For each column of row 0 (Header Row)
        ##val = sh.get((area[0], c), None)
        val = sh.cell(row=area[0], column=c).value
        if not val:
            continue
        col_name = val.strip()
        for col in cols:  # Find matching CommandField from the attribute "regex_allowed_names"
            if col.regex_allowed_names.match(col_name):
                # Found matching CommandField "col". Process
                if "@" in col_name:  # In case of use of "@", remove prefix
                    col_name = col_name[col_name.index("@")+1:]
                # Column Name to Column Index
                if not col.many_appearances:  # Column appears once
                    if col in col_map:
                        issues.append(Issue(itype=IType.ERROR,
                                            description="The column '"+col.name+"' should not appear more than one time",
                                            location=IssueLocation(sheet_name=name, row=1, column=c)))
                    col_map[col] = [(col_name, c)]
                else:  # Column appears one or more times
                    if col not in col_map:
                        col_map[col] = []
                    col_map[col].append((col_name, c))
                # Mandatory found (good)
                if col.name in mandatory_not_found:
                    mandatory_not_found.discard(col.name)
                break
        else:  # No match for the column "col_name"
            if not col_name.startswith("@"):
                issues.append(Issue(itype=IType.WARNING if ignore_not_found else IType.ERROR,
                                    description=f"In Header row, the column name '{col_name}' does not match any of the "
                                                f"allowed column names (internal command '{command_name}')",
                                    location=IssueLocation(sheet_name=name, row=1, column=c)))

    if len(mandatory_not_found) > 0:
        issues.append(Issue(itype=IType.ERROR,
                            description="In Header row, mandatory columns: " + ", ".join(
                                mandatory_not_found) + " have not been specified",
                            location=IssueLocation(sheet_name=name, row=1, column=None)))

    return col_map, issues


def read_worksheet(sh: Worksheet) -> Dict:
    rows = sh.rows
    data = {}
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            if cell.data_type == 's':
                data[(r+1, c+1)] = cell.value.strip()
            else:
                data[(r+1, c+1)] = cell.value
    return data


def parse_command_in_worksheet(sh: Worksheet, area: AreaTupleType, name: Optional[str], cmd_name: str) -> IssuesLabelContentTripleType:
    """
    Parse command in general
    Generate a JSON
    Generate a list of issues

    :param sh: Worksheet to read
    :param area: Area of the worksheet
    :param name: Name of the worksheet
    :param cmd_name: Name of the command. Key to access "command_fields" variable. Also, shown in issue descriptions
    :return: issues List, None, content (JSON)
    """

    def check_expandable(v, location):
        """
        Check if curly braces match, that what is inside is syntactically correct, (and that the value exists)

        :param v:
        :return:
        """
        import re
        reg = re.compile(r"{.*?}")
        matches = reg.findall(v)
        output = set()
        if len(matches) == 0:
            issues.append(
                Issue(itype=IType.ERROR,
                      description=f"Incorrect syntax, no macro expansion found",
                      location=location)
            )
        else:
            for m in matches:
                h_name = m[1:-1]
                try:
                    parser_field_parsers.string_to_ast(arith_boolean_expression, h_name)  # simple_h_name
                    output.add(h_name)
                except:
                    issues.append(
                        Issue(itype=IType.ERROR,
                          description=f"The value {m[1:-1]} is not a valid hierarchical name",
                          location=location)
                    )
        return output

    def commented_row(rn):
        commented = False
        v = sh.cell(row=r, column=1).value
        if v is not None:
            if str(v).startswith("#"):
                commented = True
        return commented

    issues: List[Issue] = []

    from nexinfosys.command_field_definitions import command_fields

    cols = command_fields[cmd_name]  # List of CommandField that will guide the parsing
    col_map, local_issues = check_columns(sh, name, area, cols, cmd_name)

    if any([i.itype == IType.ERROR for i in local_issues]):
        return local_issues, None, None

    issues.extend(local_issues)

    # The "mandatoriness" of a field may depend on values in other fields (like in RefBibliographic command fields)
    # Elaborate a list of fields having this "complex" mandatory property
    complex_mandatory_cols = [c for c in cols if isinstance(c.mandatory, str)]

    content = []  # The output JSON
    # Parse each Row
    for r in range(area[0] + 1, area[1]):
        line = {}
        expandable = set()  # A set of variables to be expanded. If empty, it is a literal line (not expandable)
        complex = False  # The line contains at least one field with a complex rule (which cannot be evaluated with a simple cast)

        # A row is commented if the value in the first column starts with "#" (a first empty column could be inserted
        # to ease this, just to signal commented rows)
        if commented_row(r):
            continue

        # Constant mandatory values
        mandatory_not_found = set([c.name for c in cols if c.mandatory and isinstance(c.mandatory, bool)])

        # Each "field"
        for field_def in col_map.keys():
            field_name = field_def.name
            field_defined = False
            # Appearances of field (normally just once, there are attributes allowing more than one appearance)
            for col_name, col_idx in col_map[field_def]:
                # Read and prepare "value"
                value = sh.cell(row=r, column=col_idx).value
                if value is not None:
                    if isinstance(value, float):
                        if value == int(value):
                            value = str(int(value))
                        else:
                            value = str(value)
                    elif not isinstance(value, str):
                        value = str(value)
                    value = value.strip()
                    field_defined = True
                else:
                    continue

                # Check if value contains "{", expansion
                if "{" in value:
                    # Expandable. Do not parse now. Check: curly pairs, and that what is between is a
                    #  simple_h_name and that it exists: as dataset
                    expandable.update(
                        check_expandable(value, IssueLocation(sheet_name=name, row=r, column=col_idx))
                    )
                    # With many appearances, just a "Key-Value list" syntax is permitted
                    if field_def.many_appearances:
                        if field_name in line:
                            line[field_name] += ", " + col_name + "='" + value + "'"
                        else:
                            line[field_name] = col_name + "='" + value + "'"
                    else:
                        if field_name in line:
                            line[field_name] += ", " + value
                        else:
                            line[field_name] = value  # Store the value
                else:
                    if field_def.allowed_values:  # If the CommandField checks for a list of allowed values
                        allowed_values_dict: Dict[str, str] = {v.lower(): v for v in field_def.allowed_values}
                        if value.lower() not in allowed_values_dict:  # TODO Case insensitive CI
                            issues.append(
                                Issue(itype=IType.ERROR,
                                      description=f"Field '{col_name}' of command '{cmd_name}' has invalid category "
                                      f"'{value}'. Allowed values are: {', '.join(field_def.allowed_values)}.",
                                      location=IssueLocation(sheet_name=name, row=r, column=col_idx)))
                        else:
                            # Use case from allowed values
                            line[field_name] = allowed_values_dict[value.lower()]
                    else:  # Instead of a list of values, check if a syntactic rule is met by the value
                        if field_def.parser:  # Parse, just check syntax (do not store the AST)
                            try:
                                standalone_attribute_value = "@" in field_def.allowed_names[0]
                                if not standalone_attribute_value:
                                    ast = parser_field_parsers.string_to_ast(field_def.parser, value)
                                else:
                                    try:
                                        ast = parser_field_parsers.string_to_ast(field_def.parser, value)
                                    except:
                                        ast = parser_field_parsers.string_to_ast(unquoted_string, value)

                                # Rules are in charge of informing if the result is expandable and if it complex
                                if "expandable" in ast and ast["expandable"]:
                                    issues.append(Issue(itype=IType.ERROR,
                                                        description=f"The value in field '{col_header}' of command "
                                                        f"'{cmd_name}' should not be expandable. Entered: {value}",
                                                        location=IssueLocation(sheet_name=name, row=r, column=col_idx)))
                                if "complex" in ast and ast["complex"]:
                                    complex = True

                                # With many appearances, just a "Key-Value list" syntax is permitted
                                if field_def.many_appearances:
                                    if field_name in line:
                                        line[field_name] += ", " + col_name + "='" + value + "'"
                                    else:
                                        line[field_name] = col_name + "='" + value + "'"
                                else:
                                    if field_name in line:
                                        line[field_name] += ", " + value
                                    else:
                                        line[field_name] = value  # Store the value
                            except:
                                import traceback
                                traceback.print_exc()
                                col_header = sh.cell(row=1, column=col_idx).value
                                issues.append(Issue(itype=IType.ERROR,
                                                    description=f"The value in field '{col_header}' of command "
                                                    f"'{cmd_name}' is not syntactically correct. Entered: {value}",
                                                    location=IssueLocation(sheet_name=name, row=r, column=col_idx)))
                        else:
                            line[field_name] = value  # No parser, just store blindly the value

            if field_defined and field_def.name in mandatory_not_found:
                mandatory_not_found.discard(field_def.name)

        if len(line) == 0:
            continue  # Empty line (allowed)

        # Flags to accelerate the second evaluation, during execution
        line["_row"] = r
        line["_expandable"] = list(expandable)
        line["_complex"] = complex

        # Append if all mandatory fields have been filled
        may_append = True
        if len(mandatory_not_found) > 0:
            issues.append(Issue(itype=IType.ERROR,
                                description="Mandatory columns: " + ", ".join(
                                    mandatory_not_found) + " have not been specified",
                                location=IssueLocation(sheet_name=name, row=r, column=None)))
            may_append = False

        # Check varying mandatory fields (fields depending on the value of other fields)
        for c in complex_mandatory_cols:
            field_def = c.name  # next(c2 for c2 in col_map if strcmp(c.name, c2.name))
            if isinstance(c.mandatory, str):
                # Evaluate
                mandatory = eval(c.mandatory, None, line)
                may_append = (mandatory and field_def in line) or (not mandatory)
                if mandatory and field_def not in line:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="Mandatory column: " + field_def + " has not been specified",
                                        location=IssueLocation(sheet_name=name, row=r, column=None)))

        if may_append:
            content.append(line)

    return issues, None, {"items": content, "command_name": name}
