from collections import OrderedDict

from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import AreaTupleType, IssuesLabelContentTripleType
from nexinfosys.command_generators import parser_field_parsers, Issue, IssueLocation, IType
from nexinfosys.command_generators.parser_field_parsers import simple_ident
from nexinfosys.common.helper import obtain_dataset_metadata, create_dictionary, strcmp
from nexinfosys.model_services import get_case_study_registry_objects


# TODO Currently is just a copy of "parse_etl_external_dataset_command" function
# TODO It has two new parameters: "InputDataset" and "AvailableAtDateTime"
# TODO Time dimension can be specified as "Time" or as "StartTime" "EndTime"
# TODO Result parameter column also change a bit
# TODO For a reference of fields, see "DatasetQry" command in "MuSIASEM case study commands" Google Spreadsheet

def parse_dataset_qry_command(sh: Worksheet, area: AreaTupleType, name, state) -> IssuesLabelContentTripleType:
    """
    Check that the syntax of the input spreadsheet is correct
    Return the analysis in JSON compatible format, for execution

    :param sh:   Input worksheet
    :param area: Area of the input worksheet to be analysed
    :return:     The command in a dict-list object (JSON ready)
    """
    def obtain_column(cn, r1, r2):
        """
        Obtain a list with the values of a column, in the range of rows [r1, r2)

        :param cn: Column number
        :param r1: Starting row
        :param r2: End+1 row
        :return: list with the cell values
        """
        lst = []
        for row in range(r1, r2):
            value = sh.cell(row=row, column=cn).value
            if value is None:
                continue
            if isinstance(value, str):
                lst.append(value.strip())
            else:
                lst.append(value)
        return lst

    issues = []
    # Global variables (at parse time they may not be defined, so process carefully...)
    glb_idx, p_sets, hh, datasets, mappings = get_case_study_registry_objects(state)

    # Look for the name of the input Dataset
    dataset_name = None
    available_at_datetime = None
    for c in range(area[2], area[3]):
        col_name = sh.cell(row=1, column=c).value
        if not col_name:
            continue
        if col_name.lower().strip() in ["inputdataset"]:
            lst = obtain_column(c, area[0]+1, area[1])
            for v in lst:
                if v:
                    dataset_name = v
                    break  # Stop on first definition
        elif col_name.lower().strip() in ["availableatdatetime"]:
            lst = obtain_column(c, area[0]+1, area[1])
            for v in lst:
                if v:
                    available_at_datetime = v
                    break  # Stop on first definition

    if dataset_name is None:
        issues.append(Issue(itype=IType.ERROR,
                            description=f"The name of the input dataset must be specified under column 'InputDataset'. Skipping {name} command",
                            location=IssueLocation(sheet_name=name, row=None, column=None)))
        return issues, None, None

    # Obtain the source
    from nexinfosys.ie_imports.data_source_manager import DataSourceManager
    source = DataSourceManager.obtain_dataset_source(dataset_name, datasets)
    # Obtain metadata
    dims, attrs, meas = obtain_dataset_metadata(dataset_name, source, datasets)
    # Load all code lists in a temporary dictionary of sets
    # Also check if there is a TIME dimension in the dataset
    cl = create_dictionary()
    we_have_time = False
    for d in dims:
        if dims[d].code_list:
            cl[d] = create_dictionary(data={k: None for k in dims[d].code_list.keys()})  # Attach the code list
        else:
            cl[d] = None  # No code list (TIME_PERIOD for instance)
        if dims[d].istime:
            we_have_time = True

    # Add matching mappings as more dimensions
    for m in mappings:
        if strcmp(mappings[m].source, source) and \
                strcmp(mappings[m].dataset, dataset_name) and \
                mappings[m].origin in dims:
            # Add a dictionary entry for the new dimension, add also the codes present in the map
            # tmp = [to["d"] for o in mappings[m].map for to in o["to"] if to["d"]]
            tmp = create_dictionary(data={to["d"]: None for o in mappings[m].map for to in o["to"] if to["d"]})
            cl[mappings[m].destination] = tmp  # [t[1] for t in mappings[m].map]

    # Scan columns for Dimensions, Measures and Aggregation.
    # Pivot Table is a Visualization, so now it is not in the command, there will be a command aside.

    # TODO The result COULD be an automatic BI cube (with a separate field)
    # TODO - Write into a set of tables in Mondrian
    # TODO - Generate Schema for Mondrian
    # TODO - Write the Schema for Mondrian

    out_dims = []

    out_measures = OrderedDict()
    for r in range(area[0]+1, area[1]+1):
        out_measures[r] = dict(measure=None, agg_func=None, measure_as=None)

    filter_ = {}  # Cannot use "create_dictionary()" because CaseInsensitiveDict is NOT serializable (which is a requirement)
    result_name = None  # By default, no name for the result. It will be dynamically obtained
    measure_names_column = None
    aggregations_column = None
    for c in range(area[2], area[3]):  # Each column
        col_name = sh.cell(row=1, column=c).value
        if not col_name:
            continue
        if col_name.lower().strip() in ["resultdimensions", "dimensions"]:  # "GROUP BY"
            lst = obtain_column(c, area[0] + 1, area[1])
            for r, d in enumerate(lst):
                if not d:
                    continue
                if d not in cl:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="The dimension specified for output, '"+d+"' is neither a dataset dimension nor a mapped dimension. ["+', '.join([d2 for d2 in cl])+"]",
                                        location=IssueLocation(sheet_name=name, row=r + 1, column=c + 1)))
                else:
                    out_dims.append(d)
        elif col_name.lower().strip() in ["resultmeasures", "measures"]:  # "SELECT"
            measure_names_column = c
            lst = obtain_column(c, area[0] + 1, area[1])
            # Check for measures
            # TODO (and attributes?)
            for r, m in enumerate(lst):
                if not m:
                    continue
                if m not in meas:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="The specified measure, '"+m+"' is not a measure available in the dataset. ["+', '.join([m2["measure"] for m2 in out_measures.values])+"]",
                                        location=IssueLocation(sheet_name=name, row=r + 1, column=c + 1)))
                else:
                    out_measures[r+area[0]+1]["measure"] = m
        elif col_name.lower().strip() in ["resultmeasuresaggregation", "resultmeasuresaggregator", "aggregation"]:  # "SELECT AGGREGATORS"
            aggregations_column = c
            lst = obtain_column(c, area[0] + 1, area[1])
            for r, f in enumerate(lst):
                if not f:
                    continue

                if f.lower() not in ["sum", "avg", "count", "sumna", "countav", "avgna", "pctna"]:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="The specified aggregation function, '"+f+"' is not one of the supported ones: 'sum', 'avg', 'count', 'sumna', 'avgna', 'countav', 'pctna'",
                                        location=IssueLocation(sheet_name=name, row=r + 1, column=c + 1)))
                else:
                    out_measures[r+area[0]+1]["agg_func"] = f
        elif col_name.lower().strip() in ["resultmeasurename", "resultmeasuresnames", "resultmeasuresas", "measuresas"]:  # "AS <name>"
            lst = obtain_column(c, area[0] + 1, area[1])
            for r, m in enumerate(lst):
                out_measures[r+area[0]+1]["measure_as"] = m
        elif col_name in cl:  # A dimension -> "WHERE"
            # Check codes, and add them to the "filter"
            lst = obtain_column(c, area[0] + 1, area[1])
            for r, cd in enumerate(lst):
                if not cd:
                    continue
                if str(cd) not in cl[col_name]:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="The code '"+cd+"' is not present in the codes declared for dimension '"+col_name+"'. Please, check them.",
                                        location=IssueLocation(sheet_name=name, row=r + 1, column=c + 1)))
                else:
                    if col_name not in filter_:
                        lst2 = []
                        filter_[col_name] = lst2
                    else:
                        lst2 = filter_[col_name]
                    lst2.append(cd)
        elif we_have_time and col_name.lower() in ["startperiod", "starttime", "endperiod", "endtime"]:  # SPECIAL "WHERE" FOR TIME
            # TODO Instead, should use a single column, "Time", using the interval syntax of the Time column in the Data Input command
            # Interval of time periods
            lst = obtain_column(c, area[0] + 1, area[1])
            if len(lst) > 0:
                if col_name.lower() == "starttime":
                    col_name = "StartPeriod"
                elif col_name.lower() == "endtime":
                    col_name = "EndPeriod"
                filter_[col_name] = lst[0]  # In this case it is not a list, but a number or string !!!!
        elif col_name.lower() in ["outputdatasetname", "outputdataset", "result_name", "result name", "resultname"]:
            lst = obtain_column(c, area[0] + 1, area[1])
            if len(lst) > 0:
                result_name = lst[0]
                try:
                    parser_field_parsers.string_to_ast(simple_ident, result_name)
                except:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="Column '" + col_name + "' has an invalid dataset name '" + result_name + "'",
                                        location=IssueLocation(sheet_name=name, row=2, column=c + 1)))

    # If more than one agg function defined -> all must be defined
    # If no agg func defined -> assume AVG
    # If agg func defined only in first row -> extend to other columns
    agg_funcs = [v["agg_func"] for v in out_measures.values() if v["agg_func"]]
    if len(agg_funcs) > 1:
        first_agg_func = None
    elif len(agg_funcs) == 0:
        issues.append(Issue(itype=IType.WARNING,
                            description="No aggregation function specified. Assuming 'average'",
                            location=IssueLocation(sheet_name=name, row=1, column=aggregations_column)))
        first_agg_func = "avg"
    else:  # One aggregation function
        first_agg_func = out_measures[area[0]+1]["agg_func"]
        if not first_agg_func:
            issues.append(Issue(itype=IType.ERROR,
                                description="The aggregation function must be defined in the first row",
                                location=IssueLocation(sheet_name=name, row=1, column=aggregations_column)))

    if first_agg_func:
        for v in out_measures.values():
            if v.get("measure", None):
                v["agg_func"] = first_agg_func

    # Uniform rows, with the three values defined: measure, aggregation function and "measure as"
    for r, v in out_measures.items():
        measure = v.get("measure", None)
        agg_func = v.get("agg_func", None)
        measure_as = v.get("measure_as", None)
        if measure and not agg_func or not measure and agg_func:
            issues.append(Issue(itype=IType.ERROR,
                                description="Each measure must be associated with an aggregation function",
                                location=IssueLocation(sheet_name=name, row=r, column=measure_names_column)))
        elif measure and not measure_as:
            v["measure_as"] = measure + "_" + agg_func

    measures = [v["measure"] for v in out_measures.values() if v["measure"]]
    measures_as = [v["measure_as"] for v in out_measures.values() if v["measure_as"]]
    agg_funcs = [v["agg_func"] for v in out_measures.values() if v["agg_func"]]

    if len(measures) == 0:
        issues.append(Issue(itype=IType.ERROR,
                            description="At least one measure should be specified",
                            location=IssueLocation(sheet_name=name, row=1, column=measure_names_column)))

    # measures != agg_funcs && len(agg_funcs) == 1 --> OK
    if len(measures) != len(agg_funcs) and len(agg_funcs) != 1:
        issues.append(Issue(itype=IType.ERROR,
                            description="There must be one aggregation function (used for all measures) or one aggregation per measure",
                            location=IssueLocation(sheet_name=name, row=1, column=aggregations_column)))

    if not result_name:
        result_name = source + "_" + dataset_name
        issues.append(Issue(itype=IType.WARNING,
                            description="No result name specified. Assuming '"+result_name+"'",
                            location=IssueLocation(sheet_name=name, row=2, column=c + 1)))

    content = {"dataset_source": source,
               "dataset_name": dataset_name,
               "dataset_datetime": available_at_datetime,
               "where": filter_,
               "dimensions": [d for d in dims],
               "group_by": out_dims,
               "measures": measures,
               "agg_funcs": agg_funcs,
               "measures_as": measures_as,
               "result_name": result_name
               }
    return issues, None, content

