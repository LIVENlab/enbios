from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import IssuesLabelContentTripleType, AreaTupleType
from nexinfosys.command_generators import parser_field_parsers
from nexinfosys.common.helper import strcmp, obtain_dataset_metadata, \
    check_dataset_exists, create_dictionary


def parse_mapping_command(sh: Worksheet, area: AreaTupleType, origin, destination) -> IssuesLabelContentTripleType:
    """
    Map from a set of categories from an external dataset into a set of MuSIASEM categories
    If the categories do not exist, they are created flat. Later they can be turned into a hierarchy and the mapping
    will still hold

    The syntax of the mapping allows expressing MANY to ONE and also MANY to MANY correspondence.
    The mapping has to be complete (all elements from left side must be covered, if not "" is assumed on the right side)

    :param sh: Input worksheet
    :param area: Tuple (top, bottom, left, right) representing the rectangular area of the input worksheet where the
    command is present
    :param origin:
    :param destination:
    :return: list of issues (issue_type, message), command label, command content
    """
    some_error = False
    issues = []
    # Analyze Origin
    cell = sh.cell(row=area[0], column=area[2])
    col_name = cell.value
    if origin:
        if not strcmp(origin, col_name):
            some_error = True
            issues.append((3, "The Origin name is different in the sheet name and in the worksheet ("+origin+", "+col_name+")"))
    else:
        origin = col_name

    #   Obtain the source, the dataset and the dimension of "origin"
    spl = origin.split(".")
    if len(spl) == 3:  # Source.Dataset.Dimension
        s, ds, dim = spl
        s = s + "."
        origin_ok = True
    elif len(spl) == 2:  # Dataset.Dimension
        ds, dim = spl
        s = ""
        origin_ok = True
    else:
        origin_ok = False
        some_error = True
        issues.append((3, "Origin must specify a dataset and a dimension name separated by '.'"))

    if origin_ok:
        origin_dataset = s + ds
        origin_dim = dim

        if not check_dataset_exists(origin_dataset):
            some_error = True
            issues.append((3, "The Origin '" + origin_dataset + "' does not match any registered dataset"))
        else:
            dims, attrs, meas = obtain_dataset_metadata(ds)
            if origin_dim not in dims:
                some_error = True
                issues.append((3, "The Origin dataset '" + origin_dataset + "' does not have a dimension '" + origin_dim + "'"))

    # Analyze Destination
    cell = sh.cell(row=area[0], column=area[2] + 1)
    col_name = cell.value
    if destination:
        if not strcmp(destination, col_name):
            some_error = True
            issues.append((3, "The Destination name is different in the sheet name and in the worksheet (" + destination + ", " + col_name + ")"))
    else:
        destination = col_name

    #  Destination name must be a simple identity
    try:
        parser_field_parsers.simple_ident.parseString(destination, parseAll=True)
    except:
        some_error = True
        issues.append((3, "'" + destination + "' category name has to be a simple identifier"))

    if some_error:  # Issues at this point are errors, return if there are any
        return issues, None, None

    # Read mapping Origin to Destination
    o_dict = create_dictionary()
    for r in range(area[0] + 1, area[1]):
        o_value = sh.cell(row=r, column=area[2]).value  # First column -> Origin
        d_value = sh.cell(row=r, column=area[2] + 1).value  # Second column -> Destination
        try:
            exp_value = sh.cell(row=r, column=area[2] + 2).value  # Third column -> Weight (for Many to Many mappings)
            if exp_value:
                try:
                    exp_value = float(exp_value)
                except:  # If it is not possible, it maybe an expression, postpone conversion until usage
                    pass
            else:
                exp_value = 1.0  # If undefined -> Many to One
        except:
            exp_value = 1.0  # If undefined -> Many to One

        if not o_value and not d_value:
            # issues.append((2, "Row " + str(r) + ": Origin and Destination are not defined. Row skipped."))
            continue
        elif not o_value or not d_value:
            if not o_value and d_value:
                issues.append((2, "Row "+str(r)+": Origin not defined. Row skipped."))
            else:
                issues.append((2, "Row " + str(r) + ": Destination not defined. Row skipped."))
            continue

        o_value = str(o_value).lower()
        d_value = str(d_value).lower()
        if o_value in o_dict:
            lst = o_dict[o_value]
        else:
            lst = []
            o_dict[o_value] = lst
        # Check "d_value" is not being repeated for "o_value"
        if (len(lst) == 0) or (len(lst) >= 1 and d_value not in [d["d"] for d in lst]):
            lst.append({"d": d_value, "w": exp_value})
        else:
            issues.append((3, "Destination category '" + destination + "' has been repeated for origin category '" + o_value + "' at row '"+str(r)+"'"))

    # List of dictionaries, where each dictionary contains the specification of an origin "o"
    # For multiple entries (many to many map), the origin maps a list "to" of dictionaries "d", "e"
    content = {"origin_dataset": origin_dataset,  # Name of the origin dataset (may include the source name)
               "origin_dimension": origin_dim,  # Name of the origin dimension inside the dataset
               "destination": destination,  # Name of the destination hierarchy
               "map": [{"o": k, "to": v} for k, v in o_dict.items()]
               }
    label = ((content["origin_dataset"] + ".") if origin_dataset else "") + content["origin_dimension"] + " -> " + content["destination"]
    return issues, label, content
    # else:
    #     if not some_error:
    #         cmd = MappingCommand(label)
    #         cmd.json_deserialize(content)
    #     else:
    #         cmd = None
    #     return cmd, issues
