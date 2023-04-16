from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import IssuesLabelContentTripleType, AreaTupleType
from nexinfosys.command_generators import parser_field_parsers
from nexinfosys.command_generators.parser_spreadsheet_utils import obtain_rectangular_submatrices, \
    binary_mask_from_worksheet
from nexinfosys.common.helper import strcmp


def parse_scale_conversion_command(sh: Worksheet, area: AreaTupleType, name: str = None) -> IssuesLabelContentTripleType:
    """
    Analyze the input area
    Obtain the numerical part
    Read a row above and a column to the left, looking for source (left col) and target (row above) factor types

    FactorTypes do not need to exist previously, they can be created

    :param sh: Input worksheet
    :param area: Tuple (top, bottom, left, right) representing the rectangular area of the input worksheet where the
    command is present
    :return: list of issues (issue_type, message), command label, command content
    """

    def get_subrow(r, c1, c2):
        lst = []
        # To deal with combined cell ranges, store "previous" value, and if "" is found, assume it is a merged cell
        previous = None
        for c in range(c1, c2):
            v = sh.cell(row=r, column=c).value
            if not v:
                if previous:
                    lst.append(previous)
                else:
                    lst.append("")
            else:
                previous = v
                lst.append(v)

        return lst

    def get_subcolumn(c, r1, r2):
        lst = []
        # To deal with combined cell ranges, store "previous" value, and if "" is found, assume it is a merged cell
        # !!! This may not be correct at all times: when a cell is intentionally left blank
        # To solve this, use "sh.merged_cell_ranges" to check if the current cell (r, c) is inside a range
        previous = None
        for r in range(r1, r2):
            v = sh.cell(row=r, column=c).value
            if not v:
                if previous:
                    lst.append(previous)
                else:
                    lst.append("")
            else:
                previous = v
                lst.append(v)
        return lst

    # ---------------------------------------------

    some_error = False
    issues = []

    # Detect the matrix defining scales
    m = binary_mask_from_worksheet(sh, True)  # "True" is to focus on cells containing numbers
    # Locate the matrix with numbers. Assume this defines the labels to consider, they will be around the matrix
    t = obtain_rectangular_submatrices(m)[0]  # Take just the first tuple: U=t[0], D=t[1], L=t[2], R=t[3]
    t = (t[0]+1, t[1]+1, t[2]+1, t[3]+1)  # The previous calculation is done using Numpy, so it is Zero based. Correct this

    # Obtain the factor type names in the subrow on top of the matrix
    subrow = get_subrow(t[0]-1, t[2], t[3])
    # Obtain the factor type names in the subcolumn to the left of the matrix
    subcol = get_subcolumn(t[2]-1, t[0], t[1])

    # Check that we have valid factor type names
    for ft in subrow+subcol:
        try:
            parser_field_parsers.string_to_ast(parser_field_parsers.simple_h_name, ft)
        except:
            some_error = True
            issues.append((3, "'"+ft+"' is not a valid Factor Type name"))
    if some_error:
        return issues, None, None

    # Scan the matrix, creating scale records
    scales = []
    for i, r in enumerate(range(t[0], t[1])):
        for j, c in enumerate(range(t[2], t[3])):
            v = sh.cell(row=r, column=c).value
            if v:
                if not isinstance(v, str):
                    v = str(v)
                # Origin factor
                origin = subcol[i]
                # Destination factor
                destination = subrow[j]
                if strcmp(origin, destination):
                    issues.append((3, "A change of scale to the same factor type ("+origin+") is not allowed"))
                else:
                    try:
                        parser_field_parsers.string_to_ast(parser_field_parsers.expression_with_parameters, v)
                        # Add the scale
                        scales.append(dict(origin=origin, destination=destination, scale=v))
                    except:
                        issues.append((3, "The expression '"+v+"' at the intersection of factor types " + origin + " and " + destination + " is syntactically incorrect"))

    content = {"origin_factor_types": subcol,
               "destination_factor_types": subrow,
               "scales": scales
               }

    return issues, None, content
