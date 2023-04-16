from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import AreaTupleType, IssuesLabelContentTripleType
from nexinfosys.command_generators import parser_field_parsers


def parse_hierarchy_command(sh: Worksheet, area: AreaTupleType, name: str, n_type: str) -> IssuesLabelContentTripleType:
    """
    Analyze a "hierarchy" command expressed in a worksheet of a spreadsheet

    The resulting JSON will be:
    {
    "name": <hierarchy name>,
    "type": ("Category", "FactorType", "Processor"), (this determines if the hierarchy is "is-a" -categories or factor types- or "part-of" -processors-)
    "h": [{"name": ..., "description": ..., "expression": ..., children: []},
         ]
    }

    In a hierarchy only simple names (not hierarchic) are allowed. The full name is determined by its position in the tree
    At execution time, if the elements already exist, their location in the hierarchy is updated (and the description, if present, is added)

    :param sh: Input worksheet
    :param area: Tuple (top, bottom, left, right) representing the rectangular area of the input worksheet where the
    command is present
    :param n_type: Type of hierarchy node: "C" (Category), "I" (InterfaceType) or "P" (Processor)
    :return: list of issues [(issue_type, message)], command label, command content
    """
    some_error = False
    issues = []

    col_names = {("expression", "formula"): "expression",
                 ("code", "name"): "code",
                 ("description",): "description"
                 }

    # Scan columns to prepare:
    # * "expression_column". The column that can contain an expression (it is Optional)
    # * "levels". List of Levels, formed by pairs "code, description", where "description" is optional
    expression_column = None
    levels = []
    for c in range(area[2], area[3]):  # Scan all columns
        col_name = sh.cell(row=area[0], column=c).value
        if not col_name:
            continue

        for k in col_names:
            col_name = col_name.lower()
            if col_name in k:
                if col_name == "expression":
                    expression_column = c
                elif col_name == "code":
                    levels.append(tuple([c]))
                elif col_name == "description":
                    # Description if there is an active CODE. If the description for the active CODE was
                    # already satified, replace it...
                    if len(levels) > 0:
                        tmp = levels[-1]
                        levels[-1] = (tmp[0], c)  # Code, Description
                break

    # Now, scan rows.
    # Only one Level can be active at a time.
    # Current level starts in zero, and is updated in each row.
    # Level can increase by one with regard to the previous level, or freely decrease
    nodes = {}  # Store nodes to check expressions later. Row number is key of the dictionary, the node is the value
    nodes_stack = []
    current_level = -1
    for r in range(area[0]+1, area[1]):
        found = False
        for level, t in enumerate(levels):
            code_column = t[0]

            value = sh.cell(row=r, column=code_column).value
            if value:
                found = True
                break
        if found:
            # Value syntax. A simple identity name
            try:
                parser_field_parsers.string_to_ast(parser_field_parsers.simple_ident, value)
            except:
                issues.append((3, "The name of the category must be a simple name. Row "+str(r)))
            # Description
            if len(t) > 1:
                description_column = t[1]
                description = sh.cell(row=r, column=description_column).value
            else:
                description = None
            # Expression
            if expression_column:
                expression = sh.cell(row=r, column=expression_column).value
            else:
                expression = None
            # Create the hierarchy node
            n = dict(code=value, description=description, expression=expression, children=[])
            if not n["expression"]:
                del n["expression"]
            if not n["description"]:
                del n["description"]
            # Store the node
            nodes[r] = n

            # Process hierarchical information
            add_node = True
            if level == current_level + 1:
                # New (empty) list
                nodes_stack.append([])
                current_level = level
            elif level <= current_level:
                while current_level > level:
                    lst = nodes_stack.pop()  # Take and remove last element of the stack
                    current_level -= 1
                    if current_level >= 0:
                        # From the current level, children of the last node of the list are defined in "lst"
                        nodes_stack[current_level][-1]["children"] = lst
            else:
                issues.append((3, "Hierarchical level must increase by one, not more. Previous level was "+str(current_level)+", current is "+str(level)+". Row "+str(r)))
                add_node = False
            # Append the new node to the current level
            if add_node:
                nodes_stack[current_level].append(n)

    # Close
    while current_level > 0:
        lst = nodes_stack.pop()  # Take and remove last element of the stack
        current_level -= 1
        if current_level >= 0:
            # From the current level, children of the last node of the list are defined in "lst"
            nodes_stack[current_level][-1]["children"] = lst
    # Check that expressions are correct and that they refer to existing codes
    # TODO Check that expressions are not circular
    codes = set([n["code"].lower() for n in nodes.values()])
    for r, n in nodes.items():
        code = n["code"]
        if "expression" in n:
            expression = n["expression"]
            ast = parser_field_parsers.string_to_ast(parser_field_parsers.hierarchy_expression, expression)
            for p in ast["terms"]:
                if isinstance(p, str):
                    if p.lower() not in codes:
                        issues.append((3, "The code '"+p+"' in the expression '"+expression +
                                          "' (declaration of code '"+code+"') was not defined. Row: "+str(r)))

    content = {"name": name,
               "type": n_type,
               "h": nodes_stack[0]}

    return issues, None, content


