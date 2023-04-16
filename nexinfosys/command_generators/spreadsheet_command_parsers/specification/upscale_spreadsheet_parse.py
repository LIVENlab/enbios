from openpyxl.worksheet.worksheet import Worksheet

from nexinfosys import IssuesLabelContentTripleType, AreaTupleType
from nexinfosys.command_generators.parser_spreadsheet_utils import obtain_rectangular_submatrices, \
    binary_mask_from_worksheet


def parse_upscale_command(sh: Worksheet, area: AreaTupleType, name: str = None) -> IssuesLabelContentTripleType:
    """
    Analyze the input area
    Obtain the numerical part
    Obtain the tags identifying parent and child processors. Some tags may be used for both

    Read
    Most "parse" methods are mostly syntactic (as opposed to semantic). They do not check existence of names.
    But in this case, the valid field names are fixed beforehand, so they are checked at this time.
    Some of the fields will be controlled also, according to some

    :param sh: Input worksheet
    :param area: Tuple (top, bottom, left, right) representing the rectangular area of the input worksheet where the
    command is present
    :return: list of issues (issue_type, message), command label, command content
    """

    def get_subrow(r, c1, c2):
        lst = []
        # To deal with combined cell ranges, store "previous" value, and if "" is found, assume it is a merged cell
        previous = None
        for c in range(c1, c2):
            v = sh.cell(row=r, column=c).value
            if not v:
                if previous:
                    lst.append(previous)
                else:
                    lst.append("")
            else:
                previous = v
                lst.append(v)

        return lst

    def get_subcolumn(c, r1, r2):
        lst = []
        # To deal with combined cell ranges, store "previous" value, and if "" is found, assume it is a merged cell
        # !!! This may not be correct at all times: when a cell is intentionally left blank
        # To solve this, use "sh.merged_cell_ranges" to check if the current cell (r, c) is inside a range
        previous = None
        for r in range(r1, r2):
            v = sh.cell(row=r, column=c).value
            if not v:
                if previous:
                    lst.append(previous)
                else:
                    lst.append("")
            else:
                previous = v
                lst.append(v)
        return lst

    some_error = False
    issues = []

    # Detect the matrix defining scales
    m = binary_mask_from_worksheet(sh, True)  # "True" is to focus on cells containing numbers
    # Locate the matrix with numbers. Assume this defines the labels to consider, they will be around the matrix
    t = obtain_rectangular_submatrices(m)[0]  # Take just the first tuple: U=t[0], D=t[1], L=t[2], R=t[3]
    t = (t[0]+1, t[1]+1, t[2]+1, t[3]+1)  # The previous calculation is done using Numpy, so it is Zero based. Correct this

    # Obtain the parent, child and factor being scaled
    try:
        parent_child = sh.cell(row=t[0] - 1, column=t[2] - 1).value
        child, parent = parent_child.split("/")
        parent_processor_type = parent.strip()
        child_processor_type = child.strip()
    except:
        issues.append((3, "Could not obtain the parent and child processor types: '"+parent_child+"' in upscale command"))
        some_error = True

    scaled_factor = sh.cell(row=t[0]-2, column=t[2]-1).value
    if not scaled_factor:
        issues.append((3, "Factor not specified in upscale command"))
        some_error = True

    if some_error:
        return issues, None, None

    # # Merged cells
    # for ra in sh.merged_cell_ranges:
    #     t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
    #     mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
    #     v = m[mc[0], mc[2]]
    #     m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    # Analyze rows above the matrix, to see code lists
    subrows = []
    for r in range(t[0]-1, 0, -1):
        # Get the whole subrow into a list
        subrow = get_subrow(r, t[2], t[3])
        # If empty, break
        if not any(subrow):
            break
        subrows.append(subrow)

    # Analyze cols to the left of the matrix, to see code lists
    subcols = []
    for c in range(t[2]-1, 0, -1):
        # Get the whole subcolumn into a list
        subcol = get_subcolumn(c, t[0], t[1])
        # If empty, break
        if not any(subcol):
            break
        subcols.append(subcol)

    # "scales" list will contain a list of dictionaries with the keys "codes" which are the codes leading to a cell
    # in the table, and "weight" the number in the cell. To form "codes" a prefix, formed by codes in the columns to the
    # left, is concatenated to a suffix, formed by the codes in the rows on top of the matrix

    # Obtain suffix tuples from columns (as lists really)
    sufix_tuples = []
    for i, c in enumerate(range(t[2], t[3])):
        sufix_tuples.append([subrows[j][i] for j in range(len(subrows))])

    # Iterate every row, col in the factors matrix to obtain the SCALES list
    scales = []
    for i, r in enumerate(range(t[0], t[1])):
        prefix_tuple = [subcols[j][i] for j in range(len(subcols))]  # Subtuple from subcols
        for j, c in enumerate(range(t[2], t[3])):
            w = sh.cell(row=r, column=c).value
            scales.append(dict(codes=prefix_tuple+sufix_tuples[j], weight=w))

    # # Parent Processor type	Child Processor type	Scaled Factor	Source
    # parent_processor_type = None
    # child_processor_type = None
    # scaled_factor = None
    # source = None
    # for c in range(area[2], area[3]):  # Columns
    #     key = sh.cell(row=1, column=c).value
    #     value = sh.cell(row=2, column=c).value
    #     if key.lower() in ["parent"]:
    #         parent_processor_type = value
    #     elif key.lower() in ["child"]:
    #         child_processor_type = value
    #     elif key.lower() in ["scaled factor"]:
    #         scaled_factor = value
    #     elif key.lower() in ["source"]:  # "Observer"
    #         source = value

    content = {"parent_processor_type": parent_processor_type,
               "child_processor_type": child_processor_type,
               "scaled_factor": scaled_factor,
               "source": None,
               "scales": scales
               }
    label = "Upscale child '"+child_processor_type+"' into parent '"+parent_processor_type+"'"
    return issues, label, content


