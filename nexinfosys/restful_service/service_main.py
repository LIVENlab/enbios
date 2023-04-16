import binascii
import csv
import io
import json
import logging
import os
import traceback
import urllib
from pathlib import Path

import openpyxl
import pandas as pd
import pyximport
import redis
# from flask import (jsonify, abort, redirect, url_for,
#
#                    )
from NamedAtomicLock import NamedAtomicLock
from flask import (Response, request, session as flask_session, send_from_directory, redirect
                   )
from flask.helpers import get_root_path
from flask_cors import CORS
from flask_session import Session as FlaskSessionServerSide
from openpyxl.writer.excel import save_virtual_workbook
from werkzeug.exceptions import NotFound

from nexinfosys.ie_exports.jupyter_notebook import generate_jupyter_notebook_python, generate_jupyter_notebook_r
from nexinfosys.restful_service.rest_helper import gzipped

pyximport.install(reload_support=True, language_level=3)

# >>>>>>>>>> IMPORTANT <<<<<<<<<
# To debug in local mode, prepare an environment variable "MAGIC_NIS_SERVICE_CONFIG_FILE", with value "./nis_local.conf"
# >>>>>>>>>> IMPORTANT <<<<<<<<<
from nexinfosys.command_generators.parser_spreadsheet_utils import rewrite_xlsx_file
# from nexinfosys.command_generators.parser_spreadsheet_utils_accel import rewrite_xlsx_file

from nexinfosys.ie_exports.reference_of_commands import obtain_commands_help
from nexinfosys.command_definitions import commands
from nexinfosys.command_field_definitions import command_fields, _command_field_names
from nexinfosys.command_generators import Issue, IType
from nexinfosys.command_generators.parser_field_parsers import string_to_ast, simple_ident
from nexinfosys.common.helper import generate_json, str2bool, \
    download_file, any_error_issue, wv_upload_file
from nexinfosys.models.musiasem_methodology_support import *
from nexinfosys.restful_service import app, get_results_in_session
from nexinfosys.initialization import initialize_database_data, initialize_databases, get_parameters_in_state, \
    get_scenarios_in_state, register_external_datasources, get_graph_from_state, \
    get_dataset_from_state, get_model, get_geolayer, get_ontology, validate_command, command_field_help, comm_help
import nexinfosys
from nexinfosys.command_executors import create_command
from nexinfosys.command_executors.specification.metadata_command import generate_dublin_core_xml
from nexinfosys.model_services import State, get_case_study_registry_objects
from nexinfosys.model_services.workspace import InteractiveSession, CreateNew, ReproducibleSession, \
    execute_command_container, convert_generator_to_native, prepare_and_solve_model
from nexinfosys.restful_service import nis_api_base, nis_client_base, nis_external_client_base
from nexinfosys.models import log_level
from nexinfosys.serialization import serialize, deserialize, serialize_state, deserialize_state


# #####################################################################################################################
# >>>> BOOT TIME. FUNCTIONS AND CODE <<<<
# #####################################################################################################################

def printNProcessors(s, state):
    from nexinfosys.models.musiasem_concepts import Processor
    glb_idx, p_sets, hh, datasets, mappings = get_case_study_registry_objects(state)
    logging.debug("--------------------------------------------------------")
    logging.debug(f"--- {s} -----------------------------------------")
    logging.debug(f"Number of processors: {len(glb_idx.get(Processor.partial_key()))}")
    logging.debug("--------------------------------------------------------")
    logging.debug("--------------------------------------------------------")
    logging.debug("--------------------------------------------------------")


def construct_session_persistence_backend():
    # A REDIS instance needs to be available. Check it
    # A local REDIS could be as simple as:
    #
    # docker run --rm -p 6379:6379 redis:alpine
    #
    d = {}
    if 'REDIS_HOST' in app.config:
        r_host = app.config['REDIS_HOST']
        d["SESSION_KEY_PREFIX"] = "nis:"
        d["SESSION_PERMANENT"] = False
        rs2 = None
        if r_host == "redis_lite":
            try:
                import redislite
                rs2 = redislite.Redis("tmp_nis_backend_redislite.db")  # serverconfig={'port': '6379'}
                d["SESSION_TYPE"] = "redis"
                d["SESSION_REDIS"] = rs2
                # d["PERMANENT_SESSION_LIFETIME"] = 3600
            except ImportError as e:
                logging.error("Package 'redislite' not found. Please, either change REDIS_HOST configuration variable "
                              "to 'filesystem' or 'redis', or execute 'pip install redislite' and retry")
                sys.exit(1)
        elif r_host.startswith("filesystem:"):
            d["SESSION_TYPE"] = "filesystem"
            if app.config.get("REDIS_HOST_FILESYSTEM_DIR"):
                d["SESSION_FILE_DIR"] = app.config.get("REDIS_HOST_FILESYSTEM_DIR")
            d["SESSION_FILE_THRESHOLD"] = 100
            # d["SESSION_FILE_MODE"] = 666
        else:
            rs2 = redis.Redis(r_host)
            d["SESSION_TYPE"] = "redis"
            d["SESSION_REDIS"] = rs2
            # d["PERMANENT_SESSION_LIFETIME"] = 3600
        if rs2:
            try:
                logging.debug("Trying connection to REDIS '"+r_host+"'")
                rs2.ping()
                logging.debug("Connected to REDIS instance '"+r_host+"'")
            except:
                logging.debug("REDIS instance '"+r_host+"' not reachable, exiting now!")
                sys.exit(1)
        elif "SESSION_TYPE" not in d:
            logging.error("No session persistence backend configured, exiting now!")
            sys.exit(1)
    return d


# #####################################################################################################################
# >>>> THE INITIALIZATION CODE <<<<
#

logger = logging.getLogger(__name__)
logging.getLogger('flask_cors').level = logging.DEBUG
app.logger.setLevel(log_level)
logger.setLevel(log_level)
logging.basicConfig(level=logging.DEBUG)

lock = NamedAtomicLock("nis-backend-lock")
lock.acquire()
try:
    initialize_databases()
finally:
    lock.release()

nexinfosys.data_source_manager = register_external_datasources()

d = construct_session_persistence_backend()
if "SESSION_REDIS" in d:
    nexinfosys.redis = d["SESSION_REDIS"]
else:
    nexinfosys.redis = None

app.config.update(d)

FlaskSessionServerSide(app)  # Flask Session
CORS(app,                    # CORS
     resources={r"/nis_api/*": {"origins": "*"}},
     supports_credentials=True
     )

logging.debug(f"DB_CONNECTION_STRING: {app.config['DB_CONNECTION_STRING']}\n----------------------")
logging.debug(f'Assuming {os.environ[nexinfosys.cfg_file_env_var]} as configuration file')
logging.debug(f'command_field_names = {_command_field_names}')

# #####################################################################################################################
# >>>> UTILITY FUNCTIONS <<<<
# #####################################################################################################################


def reset_database():
    """
    Empty ALL data in the database !!!!

    Used in testing web services

    :return:
    """
    if is_testing_enabled():
        connection2 = nexinfosys.engine.connect()
        tables = ORMBase.metadata.tables
        table_existence = [nexinfosys.engine.dialect.has_table(connection2, tables[t].name) for t in tables]
        connection2.close()
        if False in table_existence:
            ORMBase.metadata.bind = nexinfosys.engine
            ORMBase.metadata.create_all()

    for tbl in reversed(ORMBase.metadata.sorted_tables):
        nexinfosys.engine.execute(tbl.delete())


def build_json_response(obj, status=200):
    return Response(generate_json(obj),
                    mimetype="text/json",
                    status=status)


def serialize_isession_and_close_db_session(sess: InteractiveSession):
    logging.debug("serialize_isession IN")
    # Serialize state
    if isinstance(sess._state, str):
        logging.debug("Str")
    sess._state = serialize_state(sess._state)  # TODO New

    # Serialize WorkSession apart, if it exists
    if sess._reproducible_session:
        csvs = sess._reproducible_session._session
        # csvs.version.state = st  # TODO New
        # sess._reproducible_session.state = st  # TODO New
        if csvs and csvs.version:  # FIX: "csvs" may be None in some situations
            o_list = [csvs.version.case_study, csvs.version, csvs]
            o_list.extend(csvs.commands)
            d_list = serialize(o_list)
            # JSON Pickle and save string
            s = jsonpickle.encode({"allow_saving": sess._reproducible_session._allow_saving, "pers": d_list})
            flask_session["rsession"] = s
            sess._reproducible_session = None
        else:
            # TODO New code. Test it
            logging.debug("Reproducible session corrupted. Closing Reproducible session")
            if "rsession" in flask_session:
                del flask_session["rsession"]
    else:
        if "rsession" in flask_session:
            del flask_session["rsession"]

    tmp = sess.get_sf()
    sess.set_sf(None)
    sess._reproducible_session = None
    # Serialize sess.state and sess._identity
    s = jsonpickle.encode(sess)
    flask_session["isession"] = s

    # # Save pickled state, for "in-vitro" analysis
    # with open("/home/rnebot/pickled_state", "w") as f:
    #     f.write(s)

    sess.set_sf(tmp)
    sess.close_db_session()

    logging.debug("serialize_isession OUT")


def deserialize_isession_and_prepare_db_session(return_error_response_if_none=True) -> InteractiveSession:
    logging.debug("deserialize_issesion IN")
    if "isession" in flask_session:
        s = flask_session["isession"]
        try:
            sess = jsonpickle.decode(s)
            if sess._state:
                sess._state = deserialize_state(sess._state)
            sess.set_sf(DBSession)
            if "rsession" in flask_session:
                rs = ReproducibleSession(sess)
                rs.set_sf(sess.get_sf())
                d = jsonpickle.decode(flask_session["rsession"])
                rs._allow_saving = d["allow_saving"]
                o_list = deserialize(d["pers"])
                rs._session = o_list[2]  # type: CaseStudyVersionSession
                sess._reproducible_session = rs
        except Exception as e:
            traceback.print_exc()
            sess = None
    else:
        sess = None

    logging.debug("deserialize_issesion OUT")

    if not sess and return_error_response_if_none:
        return NO_ISESS_RESPONSE
    else:
        return sess


def is_testing_enabled():
    if "TESTING" in app.config:
        if isinstance(app.config["TESTING"], bool):
            testing = app.config["TESTING"]
        else:
            testing = app.config["TESTING"].lower() in ["true", "1"]
    else:
        testing = False
    return testing


NO_ISESS_RESPONSE = build_json_response({"error": "No interactive session active. Please, open one first ('POST /isession')"}, 400)

# >>>> SPECIAL FUNCTIONS <<<<

# @app.before_request
# def print_headers():
#     print("HEADER Authorization")
#     found = False
#     for h in request.headers:
#         if h[0] in ["Authorization", "Autorizacion"]:
#             print(h[0] + ": " + str(h[1]))
#             found = True
#     if not found:
#         print("-- not sent --")


@app.after_request
def after_a_request(response):
    for i in request.cookies.items():
        response.set_cookie(i[0], i[1])

    if "__invalidate__" in flask_session:
        response.delete_cookie(app.session_cookie_name)

    return response

# #####################################################################################################################
# >>>> SERVE ANGULAR2 CLIENT FILES <<<<
# #####################################################################################################################


@app.route("/")
def index():
    return redirect(nis_client_base)


@app.route(nis_client_base + "/", methods=["GET"])
@app.route(nis_client_base + "/<path:path>", methods=["GET"])
@app.route(nis_external_client_base + "/<path:path>", methods=["GET"])
def send_web_client_file(path=None):
    """
    Serve files from the Angular2 client
    To generate these files (ON EACH UPDATE TO THE CLIENT:
    * CD to the Angular2 project directory
    * ng build --prod --aot --base-href /nis_client/
    * CP * <FRONTEND directory>

    :param path:
    :return:

    """
    def detect_mimetype(fn):
        if fn.lower().startswith("main.") and fn.lower().endswith(".js"):
            return "text/html"
        if fn.lower().endswith(".js"):
            return "application/javascript"
        elif fn.lower().endswith(".html"):
            return "text/html"
        elif fn.lower().endswith(".png"):
            return "image/png"
        elif fn.lower().endswith(".jpg") or fn.lower().endswith(".jpeg"):
            return "image/jpeg"
        elif fn.lower().endswith(".css"):
            return "text/css"
        elif fn.lower().endswith(".json"):
            return "application/json"
        elif fn.lower().endswith(".ico"):
            return "image/x-icon"
        elif fn.lower().endswith(".svg"):
            return "image/svg+xml"
        elif fn.lower().endswith(".eot"):
            return "application/vnd.ms-fontobject"
        elif fn.lower().endswith(".woff"):
            return "application/font-woff"
        elif fn.lower().endswith(".woff2"):
            return "application/font-woff2"
        elif fn.lower().endswith(".ttf"):
            return "application/x-font-ttf"
        else:
            return None

    base = Path(get_root_path("nexinfosys.restful_service"))
    base = str(base.parent)+os.sep+"frontend"
    logger.debug("BASE DIRECTORY: "+base)
    incoming_url = request.url_rule.rule

    if not path or path == "":
        path = "index.html"

    logging.debug(f"NIS (as Web Server), serving static file with path: {path}")

    if "config.json" in path:
        return build_json_response(dict(url=f"{request.host_url[:-1]}"), 200)

    if nis_external_client_base in incoming_url:
        # From outside
        if path == "index.html":
            # TODO Possibility of changing both the base and the file name
            # TODO The intention is to NOT show the "Login" possibilities, so
            # TODO users are always anonymous. To be discussed.
            base = get_root_path("clients/web")
            new_name = "index.html"
        else:
            new_name = path
    else:
        # From inside
        new_name = path

    mimetype = detect_mimetype(new_name)

    logger.debug(f"File: {new_name}; MIMETYPE: {mimetype}")

    try:
        return send_from_directory(base, new_name, mimetype=mimetype)
    except NotFound:
        return send_from_directory(base, "index.html", mimetype="text/html")


# #####################################################################################################################
# >>>> SERVE STATIC FILES <<<<
# #####################################################################################################################


@app.route(nis_api_base + "/static/<path:path>", methods=["GET"])
def send_static_file(path):
    """
    Serve files from the Angular2 client
    To generate these files (ON EACH UPDATE TO THE CLIENT:
    * CD to the Angular2 project directory
    * ng build --prod --aot --base-href /nis_client/
    * CP * <FRONTEND directory>

    :param path:
    :return:
    """
    base = Path(get_root_path("nexinfosys.restful_service"))
    base = str(base)+"/static"
    # logger.debug("BASE DIRECTORY: "+base)

    return send_from_directory(base, path)

# #####################################################################################################################
# >>>> RESTFUL INTERFACE <<<<
# #####################################################################################################################

# -- Special "give me state" for Case Study Management --


@app.route(nis_api_base + "/isession/rsession/state_summary", methods=["GET"])
def summary_status():  # Summary status
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    d = {}
    if isess:
        d["isession_open"] = True
        # Identity
        d["identified"] = isess.get_identity_id() is not None or isess.get_identity_id() != ""
        # Reproducible session
        if isess.reproducible_session_opened():
            d["rsession_open"] = True
            # Return a list with executed flags, a flag per command
            d["commands"] = [dict(executed=c.execution_end) for c in isess.reproducible_session.ws_commands]
        else:
            d["rsession_open"] = False
    else:
        d["isession_open"] = False

    return build_json_response(d, 200)


# -- Interactive session --


@app.route(nis_api_base + "/resetdb", methods=["POST"])
def reset_db():
    testing = is_testing_enabled()
    if testing:
        reset_database()
        initialize_database_data()
        interactive_session_close()  # Leave session if already in
        r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Illegal operation!!"}, 400)

    return r


@app.route(nis_api_base + "/isession", methods=["POST"])
def interactive_session_open():
    isess = deserialize_isession_and_prepare_db_session(False)
    if isess:
        r = build_json_response({"error": "Close existing interactive session ('DELETE /isession'"}, 400)
    else:
        isess = InteractiveSession(DBSession)
        serialize_isession_and_close_db_session(isess)
        r = build_json_response({}, 204)
    return r


@app.route(nis_api_base + "/isession", methods=["GET"])
def get_interactive_session():
    isess = deserialize_isession_and_prepare_db_session(False)
    if isess:
        st = "isession_open"
    else:
        st = "isession_closed"

    logging.debug("Get Isession: "+st)
    return build_json_response(st, 200)


@app.route(nis_api_base + "/isession/state", methods=["DELETE"])
def interactive_session_reset_state():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Clear state
    if isess.state:
        isess.reset_state()
        serialize_isession_and_close_db_session(isess)

    return build_json_response({}, 204)


# Set identity at this moment for the interactive session
@app.route(nis_api_base + "/isession/identity", methods=["PUT"])
def interactive_session_set_identity():
    # Recover InteractiveSession
    # if request.method=="OPTIONS":
    #     r = build_json_response({}, 200)
    #     h = r.headers
    #     h['Access-Control-Allow-Origin'] = "http://localhost:4200"
    #     h['Access-Control-Allow-Methods'] = "PUT,POST,DELETE,GET,OPTIONS"
    #     h['Access-Control-Max-Age'] = str(21600)
    #     h['Access-Control-Allow-Credentials'] = "true"
    #     h['Access-Control-Allow-Headers'] = "Content-Type, Authorization, Content-Length, X-Requested-With"
    #     return r

    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # If there is a current identity, issue an error. First "unidentify"
    if isess.get_identity_id():

        testing = is_testing_enabled()
        if testing and request.args.get("user") and isess.get_identity_id() == request.args.get("user"):
            result = True
        else:
            result = False
    else:
        # Two types of identification: external, using OAuth tokens, or application, using user+password
        application_identification = True
        if application_identification:
            if request.args.get("user"):
                testing = is_testing_enabled()
                result = isess.identify({"user": request.args.get("user"),
                                         "password": request.args.get("password", None)
                                         },
                                        testing=testing
                                        )
        else:
            # TODO Check the validity of the token using the right Authentication service
            result = isess.identify({"token": request.headers.get("token"),
                                     "service": request.headers.get("auth_service")
                                     }
                                    )
    serialize_isession_and_close_db_session(isess)

    r = build_json_response({"identity": isess.get_identity_id()} if result else {},
                            200 if result else 401)

    return r


@app.route(nis_api_base + "/isession/identity", methods=["GET"])
def interactive_session_get_identity():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    return build_json_response({"identity": isess.get_identity_id()})


# Set to anonymous user again (or "logout")
@app.route(nis_api_base + "/isession/identity", methods=["DELETE"])
def interactive_session_remove_identity():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Un-identify
    if isess.get_identity_id():
        if isess.reproducible_session_opened():
            # If reproducible session open, error!
            r = build_json_response({"error": "There is an open Reproducible Session. Close it first."}, 401)
        else:
            isess.unidentify()
            r = build_json_response({"identity": isess.get_identity_id()})

        serialize_isession_and_close_db_session(isess)
    else:
        r = build_json_response({"identity": isess.get_identity_id()})

    return r


# Close interactive session (has to log out if some identity is active)
@app.route(nis_api_base + "/isession", methods=["DELETE"])
def interactive_session_close():
    isess = deserialize_isession_and_prepare_db_session(False)

    if isess:
        isess.quit()

    flask_session.clear()
    flask_session["__invalidate__"] = True
    return build_json_response({})


@app.route(nis_api_base + '/isession/generator.json', methods=['POST'])
def convert_generator_to_json_generator():
    """
    Send the file to the service
    Convert to native
    Return it in JSON format

    :return:
    """
    # Check Interactive Session is Open. If not, open it
    isess = deserialize_isession_and_prepare_db_session(False)
    if not isess:
        isess = InteractiveSession(DBSession)

    testing = is_testing_enabled()
    if testing:
        result = isess.identify({"user": "test_user", "password": None}, testing=True)

    # Receive file
    generator_type, content_type, buffer, _, _ = receive_file_submission(request)

    if len(buffer) == 0:
        raise Exception("No content was received. Please check the original file exists.")

    output = convert_generator_to_native(generator_type, content_type, buffer)

    # Return the conversion
    r = build_json_response(output, 200)

    serialize_isession_and_close_db_session(isess)

    return r


@app.route(nis_api_base + '/isession/generator.to_dc.xml', methods=['POST'])
def convert_generator_to_dublin_core():
    """
    Send the file to the service
    Convert to native
    Return the Dublin Core XML record

    :return:
    """
    # Check Interactive Session is Open. If not, open it
    isess = deserialize_isession_and_prepare_db_session(False)
    if not isess:
        isess = InteractiveSession(DBSession)

    testing = is_testing_enabled()
    if testing:
        result = isess.identify({"user": "test_user", "password": None}, testing=True)

    # Receive file
    generator_type, content_type, buffer, _, _ = receive_file_submission(request)

    if len(buffer) == 0:
        raise Exception("No content was received. Please check the original file exists.")

    output = convert_generator_to_native(generator_type, content_type, buffer)
    xml = None
    for c in output:
        if "command" in c and c["command"] == "metadata" and "content" in c:
            xml = generate_dublin_core_xml(c["content"])
            break

    # Return the conversion
    if xml:
        r = Response(xml,
                     mimetype="text/xml",
                     status=200)
    else:
        r = build_json_response({"message": "Could not elaborate Dublin Core XML record from the input generator"}, 401)

    serialize_isession_and_close_db_session(isess)

    return r

# -- Reproducible Sessions --


@app.route(nis_api_base + "/isession/rsession", methods=["POST"])
def reproducible_session_open():
    def read_parameters(dd):
        nonlocal uuid2, read_uuid_state, create_new, allow_saving
        # Read query parameters
        uuid2 = dd.get("uuid")
        if "read_version_state" in dd:
            read_uuid_state = dd["read_version_state"]
            read_uuid_state = str2bool(read_uuid_state)
        if "create_new" in dd:
            create_new = str(dd["create_new"])
            if create_new.lower() in ["1", "case_study", "casestudy"]:
                create_new = CreateNew.CASE_STUDY
            elif create_new.lower() in ["2", "version", "case_study_version"]:
                create_new = CreateNew.VERSION
            else:
                create_new = CreateNew.NO
        if "allow_saving" in dd:
            allow_saving = dd["allow_saving"]
            allow_saving = allow_saving.lower() == "true"

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Check identity
    identity = isess.get_identity_id()

    # Retrieve parameters
    uuid2 = None
    read_uuid_state = None
    create_new = None
    allow_saving = None

    # First, read uploaded JSON
    if len(request.files) > 0:
        for k in request.files:
            buffer = bytes(request.files[k].stream.getbuffer())
            content_type = request.files[k].content_type
            break
    else:
        buffer = bytes(io.BytesIO(request.get_data()).getbuffer())
        if "Content-Type" in request.headers:
            content_type = request.headers["Content-Type"]

    if buffer:
        read_parameters(json.loads(buffer))
    if not uuid2 and not read_uuid_state and not create_new and not allow_saving:
        read_parameters(request.form)
        if not uuid2 and not read_uuid_state and not create_new and not allow_saving:
            read_parameters(request.args)

    if read_uuid_state is None:
        read_uuid_state = True
    if create_new is None:
        create_new = CreateNew.NO
    if allow_saving is None:
        allow_saving = True

    # Persistent object to open: None (new case study), UUID (case study version)
    if isess.reproducible_session_opened():
        r = build_json_response({"error": "There is an open Reproducible Session. Close it first."}, 401)
    else:
        if allow_saving and not identity:
            r = build_json_response({"error": "When 'allow_saving==true' an identity is required."}, 401)
        else:
            try:
                # TODO New, not checked
                isess.reset_state()

                isess.open_reproducible_session(case_study_version_uuid=uuid2,
                                                recover_previous_state=read_uuid_state,
                                                cr_new=create_new,
                                                allow_saving=allow_saving
                                                )
                r = build_json_response({}, 204)
            except Exception as e:
                s = "Exception trying to open reproducible session: "+str(e)
                logger.error(s)
                r = build_json_response({"error": s}, 401)

    #
    serialize_isession_and_close_db_session(isess)
    return r


@app.route(nis_api_base + "/isession/rsession", methods=["DELETE"])
def reproducible_session_save_close():  # Close the ReproducibleSession, with the option of saving it
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal it if not,
    save = request.args.get("save_before_close", "False")
    if save:
        save = str2bool(save)
    else:
        save = False
    # The reproducible session, if saved, could be assigned to an existing case study
    cs_uuid = request.args.get("cs_uuid", None)
    if cs_uuid:
        cs_uuid = str(cs_uuid)

    # If specified, it is the name for the case study Version
    cs_name = request.args.get("cs_name", None)
    if cs_name:
        cs_name = str(cs_name)

    # Close reproducible session
    if not isess.reproducible_session_opened():
        r = build_json_response({"error": "There is no open Reproducible Session. Cannot close"}, 401)
    else:
        try:
            uuid_, v_uuid, cs_uuid = isess.close_reproducible_session(issues=None,
                                                                      output=None,
                                                                      save=save,
                                                                      from_web_service=True,
                                                                      cs_uuid=cs_uuid,
                                                                      cs_name=cs_name)
            r = build_json_response({"session_uuid": str(uuid_),
                                     "version_uuid": str(v_uuid),
                                     "case_study_uuid": str(cs_uuid)
                                     },
                                    200)
        except Exception as e:
            s = "Exception trying to close reproducible session: " + str(e)
            logger.error(s)
            r = build_json_response({"error": s}, 401)

    serialize_isession_and_close_db_session(isess)
    return r


@app.route(nis_api_base + "/isession/rsession", methods=["GET"])
def reproducible_session_get_status():  # Return current status of ReproducibleSession
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        r = build_json_response("rsession_open", 200)
    else:
        r = build_json_response("rsession_closed", 200)

    return r


@app.route(nis_api_base + "/isession/rsession/command_generators/<order>", methods=["GET"])
def reproducible_session_get_command_generator(order):  # Return one of the command generators
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    order = int(order)
    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if order < len(isess.reproducible_session.ws_commands):
            c = isess.reproducible_session.ws_commands[order]
            r = Response(c.content, mimetype=c.content_type)
        else:
            r = build_json_response({"error":
                                     "Command number " + str(order) +
                                     " requested, only "+str(len(isess.reproducible_session.commands))+" available."})
    else:
        r = build_json_response("No open reproducible Session", 200)

    return r


# ----------------------------------------------------------------------------------------------------------------------
# State management: save, list, get, delete ("update"" is "save", overwrite always)
# ----------------------------------------------------------------------------------------------------------------------

@app.route(nis_api_base + "/isession/rsession/state", methods=["PUT"])
def reproducible_session_save_state():  # Save state
    """
    Save or overwrite state in-memory to a file at the backend side

    Receives a "code" Query parameter with the name for the saved state file (which must be unique, unless an overwrite
    is wanted)

    :return: Empty if everything is ok, Error if there is an issue
    """
    def ensure_dir(file_path):
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            code = request.args.get("code", None)
            try:
                string_to_ast(simple_ident, code)
            except:
                code = None
            if code is None:
                r = build_json_response({"error": "Query parameter 'code' is mandatory"}, 401)
            else:
                cs_path = nexinfosys.get_global_configuration_variable("CASE_STUDIES_DIR")
                ensure_dir(cs_path)
                # Save state
                s = serialize_state(isess.state)
                with open(cs_path+os.sep+code+".state_serialized", "wt") as f:
                    f.write(s)

                r = build_json_response({}, 204)
        else:
            r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Cannot save state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state", methods=["DELETE"])
def reproducible_session_delete_state():  # Delete state
    """
    Delete a saved state

    Receives a "code" Query parameter with the name for the saved state file to delete

    :return: Empty if everything is ok, Error if there is an issue
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            code = request.args.get("code", None)
            try:
                string_to_ast(simple_ident, code)
            except:
                code = None
            if code is None:
                r = build_json_response({"error": "Query parameter 'code' is mandatory"}, 401)
            else:
                cs_path = nexinfosys.get_global_configuration_variable("CASE_STUDIES_DIR")
                fname = cs_path+os.sep+code+".state_serialized"
                if os.path.exists(fname):
                    os.remove(fname)
                    r = build_json_response({}, 204)
                else:
                    r = build_json_response({"error": f"A state with code {code} did not exist"}, 401)
        else:
            r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Cannot delete state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state/", methods=["GET"])
def reproducible_session_list_states():  # List available states
    """
    List codes of all previously saved states

    :return: A JSON with a single entry "codes", with a list of the codes to address the saved states. Error if there is an issue
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        cs_path = nexinfosys.get_global_configuration_variable("CASE_STUDIES_DIR")
        lst = [f for f in os.listdir(cs_path) if os.path.isfile(f"{cs_path}{os.sep}{f}")]
        r = build_json_response({"codes": lst}, 204)
    else:
        r = build_json_response({"error": "Cannot return the list of states, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state", methods=["GET"])
def reproducible_session_load_state():
    """
    Loads a previously saved state in the reproducible session. After this call, output datasets can be retrieved or
    new parameters for the dynamic scenario submitted.

    A "code" Query parameter must be passed with a code for the saved state.

    :return: Empty if everything is ok (the state is on the backend side). Error if there is an issue
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        cs_path = nexinfosys.get_global_configuration_variable("CASE_STUDIES_DIR")
        code = request.args.get("code", None)
        try:
            string_to_ast(simple_ident, code)
        except:
            code = None
        if code is None:
            r = build_json_response({"error": "Query parameter 'code' is mandatory"}, 401)
        else:
            fname = cs_path + os.sep + code + ".state_serialized"
            with open(fname, "rt") as f:
                s = f.read()
                isess.state = deserialize_state(s)

            r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Cannot load state, no open reproducible session"}, 401)

    return r


# ----------------------------------------------------------------------------------------------------------------------

@app.route(nis_api_base + "/isession/rsession/state.pickled", methods=["GET"])
def reproducible_session_get_state():  # Return current status of ReproducibleSession
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            r = build_json_response(jsonpickle.encode(isess.state), 200)
        else:
            r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Cannot return state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query", methods=["GET"])
def reproducible_session_query_state():  # Query aspects of State
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            # TODO Parse query, execute it, return results
            # TODO By concept: Datasets, processors, factors, factor types, hierarchies, mappings, ISSUES (extra MuSIASEM, errors in some level: syntax, semantics, solving)
            # TODO Information: name, quantitites (attached to factors), relations, hierarchy (for hierarchies)
            # TODO By observer
            pass
        else:
            r = build_json_response({}, 204)
    else:
        r = build_json_response({"error": "Cannot return state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query/issues", methods=["GET"])
def reproducible_session_query_state_list_issues():  # Query list of issues IN the current state
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            issues = isess.state.get("_issues")
            if not issues:
                issues = []
            r = build_json_response({"issues": issues}, 200)
        else:
            r = build_json_response([], 204)
    else:
        r = build_json_response({"error": "Cannot return state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query/everything_executed", methods=["GET"])
def reproducible_session_query_state_everything_executed():  # Query if all commands have been executed
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        first_i = len(isess.reproducible_session.ws_commands)
        for i in range(len(isess.reproducible_session.ws_commands) - 1, -1, -1):
            c = isess.reproducible_session.ws_commands[i]
            if not c.execution_start:
                first_i = i
        r = build_json_response({"everything_executed": first_i == len(isess.reproducible_session.ws_commands)}, 200)
    else:
        r = build_json_response({"error": "Cannot return state, no open reproducible session"}, 401)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query/outputs", methods=["GET"])
@app.route(nis_api_base + "/isession/rsession/state_query/datasets", methods=["GET"])
def reproducible_session_query_state_list_results():  # Query list of outputs (not only datasets) IN the current state
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            r = build_json_response(get_results_in_session(isess), 200)
        else:
            r = build_json_response([], 204)
    else:
        r = build_json_response({"error": "Cannot return list of results, no reproducible session open"}, 401)

    printNProcessors("LIST OF OUTPUTS", isess.state)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query/webdav", methods=["PUT"])
def copy_resource_to_webdav():
    """
    Read a resource and put the result into WebDAV server
    PROBABLY REQUIRES MULTIPLE WORKERS because datasets are obtained via a recursive "RESTful call"

    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess
    tmp = request.get_json()
    source_url = tmp["sourceURL"]
    target_url = tmp["targetURL"]

    from urllib.parse import urlparse
    pr = urlparse(target_url)
    # Check host
    wv_host_name = nexinfosys.get_global_configuration_variable("FS_SERVER") \
        if nexinfosys.get_global_configuration_variable("FS_SERVER") else "nextcloud.data.magic-nexus.eu"
    if wv_host_name.lower() != pr.netloc:
        return build_json_response({"error": f"Cannot save the file in the requested server location, {pr.netloc}, which is different from the configured one, {wv_host_name}"}, 401)
    # Modify URL
    target_url = f"{pr.scheme}://{pr.netloc}{os.path.split(pr.path)[0]}"
    pr = urlparse(source_url)
    target_url += f"/{os.path.split(pr.path)[1]}"

    # READ (reentrant)
    self_schema = nexinfosys.get_global_configuration_variable("SELF_SCHEMA") \
        if nexinfosys.get_global_configuration_variable("SELF_SCHEMA") else request.host_url
    import requests
    requested_resource = f"{self_schema}{source_url[1:]}"
    logging.debug(f"REENTRANT REQUEST: {requested_resource}")
    r = requests.get(requested_resource, cookies=request.cookies, verify=False)
    # WRITE
    wv_upload_file(io.BytesIO(r.content), target_url)
    logging.debug(f"REQUESTED RESOURCE UPLOADED TO NEXTCLOUD at {target_url}")

    return build_json_response([], 204)


# -- DYNAMIC PARAMETERS --
@app.route(nis_api_base + "/isession/rsession/state_query/parameters", methods=["GET"])
def get_parameter_definitions():
    """
    Obtain a JSON enumerating the definition of all the parameters for the case study

    :param format:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    res = get_parameters_in_state(isess.state)

    return build_json_response(res, 200)


@app.route(nis_api_base + "/isession/rsession/state_query/parameters", methods=["PUT"])
def set_parameters_and_solve():
    """
    Create an "interactive" scenario, composed by a dictionary of parameter values,
    passed through a JSON in the request, and SOLVE this single scenario.

    As results, create a supermatrix containing only this scenario, and the MatrixIndicators

    :return:
    """
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    parameters = request.get_json()
    issues2 = prepare_and_solve_model(isess.state, parameters)

    # Return "issues2", issues found during the solving
    isess.state.set("_issues", issues2)

    # Return outputs (could be a list of binary files)
    r = build_json_response({"issues": convert_issues(issues2), "outputs": None}, 200)

    # Must serialize in order to later recover the datasets
    serialize_isession_and_close_db_session(isess)

    return r


@app.route(nis_api_base + "/isession/rsession/state_query/scenarios", methods=["GET"])
def get_scenarios():
    """
    Return a list scenarios and values for parameters in each of them

    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    scenarios = get_scenarios_in_state(isess.state)

    return build_json_response(scenarios, 200)


@app.route(nis_api_base + "/isession/rsession/state_query/geolayer.<format>", methods=["GET"])
def get_geolayer_service(format):
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    content, content_type, ok = get_geolayer(isess.state, format)
    return Response(content, mimetype=content_type, status=200 if ok else 401)


@app.route(nis_api_base + "/isession/rsession/state_query/ontology.<format>", methods=["GET"])
def get_ontology_service(format):
    # TODO OWLREADY2 installation on the Docker image issues a problem
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    content, content_type, ok = get_ontology(isess.state, format)
    return Response(content, mimetype=content_type, status=200 if ok else 401)


@app.route(nis_api_base + "/isession/rsession/state_query/python_script.<format>", methods=["GET"])
def get_python_script(format):
    """
script capaz de reproducir lo ejecutado
* login
* open
* load_workbook
* load_workbook desde Nextcloud, sin credenciales
* mostrar cómo obtener cada uno de los datasets, comentado (llamar a "query_state_list_results(isess)")
* mostrar cómo utilizar cada uno de los datasets, comentado también

* Jupyter sólo: script capaz de relanzar, selección de parámetros, reejecución, recogida de datasets (igual)

    :param format:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    output = None
    # Generate graph from State
    if isess.state:
        if format == "python":
            # TODO Prepare Python file
            output = io.StringIO()
            mimetype = "application/x-python-code"  # or text/x-python
        elif format == "jupyternotebook":
            output = generate_jupyter_notebook_python(isess.state)
            mimetype = "application/x-ipynb+json"  # TODO

    if output:
        return Response(output, mimetype=mimetype, status=200)
    else:
        return build_json_response({"error": F"Cannot return Python script, format '{format}' not recognized"}, 401)


@app.route(nis_api_base + "/isession/rsession/state_query/r_script.<format>", methods=["GET"])
def get_r_script(format):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    output = None
    # Generate graph from State
    if isess.state:
        if format == "r":
            # TODO Prepare R file
            output = io.StringIO()
            mimetype = "application/r-system"  # TODO
        elif format == "jupyternotebook":
            output = generate_jupyter_notebook_r(isess.state)
            mimetype = "application/x-ipynb+json"  # TODO

    if output:
        return Response(output, mimetype=mimetype, status=200)
    else:
        return build_json_response({"error": F"Cannot return R script, format '{format}' not recognized"}, 401)


@app.route(nis_api_base + "/isession/rsession/state_query/commands_reference_document.<format>", methods=["GET"])
def get_commands_reference_document(format):
    if format=="html":
        mimetype = "text/html"
    return Response(obtain_commands_help(format.lower()), mimetype)


@app.route(nis_api_base + "/isession/rsession/state_query/model.<format>", methods=["GET"])
def get_model_service(format):
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    content, content_type, ok = get_model(isess.state, format)
    return Response(content, mimetype=content_type, status=200 if ok else 401)


@app.route(nis_api_base + '/isession/rsession/state_query/flow_graph.<format>', methods=["GET"])
def obtain_flow_graph(format):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    output = None
    # Generate graph from State
    if isess.state:
        output, mimetype, ok = get_graph_from_state(isess.state, f"interfaces_graph.{format}")

    if output:
        r = Response(output, mimetype=mimetype, status=200)
    else:
        r = build_json_response({}, 200)

    return r


@app.route(nis_api_base + '/isession/rsession/state_query/processors_graph.<format>', methods=["GET"])
@app.route(nis_api_base + '/isession/rsession/query/processors_graph.<format>', methods=["GET"])
def obtain_processors_graph_visjs_format(format):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Generate graph from State
    output = None
    if isess.state:
        output, mimetype, ok = get_graph_from_state(isess.state, f"processors_graph.{format}")

    if output:
        r = Response(output, mimetype=mimetype, status=200)
    else:
        r = build_json_response({}, 200)

    return r


@app.route(nis_api_base + '/isession/rsession/state_query/sankey_graph.json', methods=["GET"])
@app.route(nis_api_base + '/isession/rsession/query/sankey_graph.json', methods=["GET"])
def obtain_sankey_graph():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Generate sanskey dictionary ready for plotly from State
    if isess.state:
        _, _, _, datasets, _ = get_case_study_registry_objects(isess.state)
        if datasets["flow_graph_matrix"]:
            df = datasets.get("flow_graph_matrix").data

            sankey = {}
            for p in list(set(df['Period'])):
                df_period = df[df['Period'] == p]
                tmp = {}
                for s in list(set(df_period['Scenario'])):
                    ds_scenario = df_period[df_period['Scenario'] == s]
                    processors = list(set(ds_scenario['source_processor'].append(ds_scenario['target_processor'])))
                    source = [processors.index(i) for i in list(ds_scenario['source_processor'])]
                    target = [processors.index(i) for i in list(ds_scenario['target_processor'])]
                    label = list(ds_scenario['source'] + ' to ' + ds_scenario['target'])
                    data = dict(
                        type='sankey',
                        node=dict(
                            pad=50,
                            thickness=100,
                            line=dict(
                                color="black",
                                width=0.5
                            ),
                            label=processors,

                        ),
                        link=dict(
                            source=source,
                            target=target,
                            value=list(ds_scenario['Value']),
                            label=label
                        ))

                    tmp[s] = data
                sankey[p] = tmp

            r = build_json_response(sankey, 200)

        else:
            r = build_json_response({}, 200)

    else:
        r = build_json_response({}, 200)

    return r


@gzipped
@app.route(nis_api_base + "/isession/rsession/state_query/datasets/<name>.<format>", methods=["GET"])
def reproducible_session_query_state_get_dataset(name, format):  # Query list of datasets IN the current state
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open, signal about it if not
    if isess.reproducible_session_opened():
        if isess.state:
            labels_enabled = request.args.get("labels", "True") == "True"
            content, content_type, ok = get_dataset_from_state(isess.state, name, format, labels_enabled)
            r = Response(content, mimetype=content_type, status=200 if ok else 401)
        else:
            r = build_json_response([], 204)
    else:
        r = build_json_response({"error": "Cannot return state, no open reproducible session"}, 401)

    return r


# ----------------------------------------------------------------------------------------------------------------------

@app.route(nis_api_base + "/isession/rsession/command", methods=["POST"])
def reproducible_session_append_single_command():  # Receive a JSON or CSV command from some externally executed generator

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open
    if isess.reproducible_session_opened():
        # Read content type header AND infer "generator_type"
        content_type = request.headers["Content-Type"]
        if content_type.lower() in ["application/json", "text/json", "text/csv"]:
            generator_type = "native"

        # Read binary content
        if len(request.files) > 0:
            for k in request.files:
                buffer = bytes(request.files[k].stream.getbuffer())
                break
        else:
            buffer = bytes(io.BytesIO(request.get_data()).getbuffer())

        # Read Execute and Register parameters
        execute = request.args.get("execute", "True")
        if execute:
            execute = str2bool(execute)
        register = request.args.get("register", "True")
        if register:
            register = str2bool(register)

        if isinstance(buffer, bytes):
            d = buffer.decode("utf-8")
        else:
            d = buffer
        d = json.loads(d)
        if isinstance(d, dict) and "command" in d and "content" in d:
            if "label" in d:
                n = d["label"]
            else:
                n = None
            cmd, syntax_issues = create_command(d["command"], n, d["content"])

        if register:
            isess.register_executable_command(cmd)
        if execute:
            issues, output = isess.execute_executable_command(cmd)
            # TODO Process "ret". Add issues to an issues list. Add output to an outputs list.

        r = build_json_response({}, 204)
        serialize_isession_and_close_db_session(isess)
    else:
        r = build_json_response({"error": "A reproducible session must be open in order to submit a command"}, 400)

    return r


def receive_file_submission(req):
    """
    Receive file submitted using multipart/form-data
    Return variables for the processing of the file as a command_executors generator

    :param req: The "request" object
    :return: A tuple (generator_type -str-, content_type -str-, buffer -bytes-, execute -bool-, register -bool-)
    """

    def parse_data_url(url):
        scheme, data = url.split(":", 1)
        assert scheme == "data", "unsupported scheme: " + scheme
        mediatype, data = data.split(",", 1)
        # base64 urls might have a padding which might (should) be quoted:
        data = urllib.parse.unquote_to_bytes(data)
        if mediatype.endswith(";base64"):
            return binascii.a2b_base64(data), mediatype[:-7] or None
        else:
            return data, mediatype or None

    # Read binary content
    if len(req.files) > 0:
        for k in req.files:
            buffer = bytes(req.files[k].stream.getbuffer())
            content_type = req.files[k].content_type
            it_is_url = False
            break
    else:
        buffer = bytes(io.BytesIO(req.get_data()).getbuffer())
        content_type = req.content_type
        it_is_url = buffer.startswith(b"data") or buffer.startswith(b"http")

    if it_is_url:
        url = buffer.decode("utf-8")
        if not url.startswith("data"):
            # Try a download from the URL
            # Check if it is a Google Drive file, a Nextcloud file or a freely downloadable file
            data = download_file(url)
            buffer = data.getvalue()
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            # It may be a DATA URL
            buffer, content_type = parse_data_url(url)
            # except:
            #     content_type = req.headers["Content-Type"]

    # Infer "generator_type" from content type
    if content_type.lower() in ["application/json", "text/csv"]:
        generator_type = "primitive"
    elif content_type.lower() in ["application/excel",
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
        generator_type = "spreadsheet"
    elif content_type.lower() in ["text/x-r-source"]:
        generator_type = "R-script"
    elif content_type.lower() in ["text/x-python", "text/x-python3", "application/x-python3"]:
        generator_type = "python-script"

    # # Write to file
    # with open("/home/rnebot/output_file.xlsx", "wb") as f:
    #     f.write(buffer)

    # Read Register and Execute parameters
    register = req.form.get("register")
    if not register:
        register = req.args.get("register")
        if not register:
            register = False
    execute = req.form.get("execute")
    if not execute:
        execute = req.args.get("execute")
        if not execute:
            execute = False
    execute = str2bool(execute)
    register = str2bool(register)

    return generator_type, content_type, buffer, execute, register


def reset_state_and_reproducible_session(isess: InteractiveSession):
    """
    Simple approach: reset state on every submission

    :param isess:
    :return:
    """
    isess.reset_state()
    isess.close_reproducible_session(issues=None, output=None, save=False, from_web_service=False, cs_uuid=None, cs_name=None)
    isess.open_reproducible_session(case_study_version_uuid=None,
                                    recover_previous_state=False,
                                    cr_new=True,
                                    allow_saving=True
                                    )


# #################################################################################################################### #
# MAIN POINT OF EXECUTION BY THE GENERIC CLIENT ("ANGULAR FRONTEND")                                                   #
# #################################################################################################################### #
def convert_issues(iss_lst):
    """
    Convert issues generated by the backend into a list of dictionaries as expected by the frontend
    :param iss_lst: Issues list
    :return: Issue list in frontend compatible format
    """
    out = []
    for i in iss_lst:
        location = dict(sheet_name="", row=None, col=None)
        i_type = "Error" if i.itype.value == 3 else ("Warning" if i.itype.value == 2 else "Info")
        if isinstance(i, Issue):
            if i.location is not None:
                location = dict(sheet_name=i.location.sheet_name, row=str(i.location.row), col=str(i.location.column))
            out.append(dict(**location, message=i_type + ": " +i.description, type=i.itype.value))
        else:
            out.append(dict(**location, message="Issue type unknown", type=3))
    return out


@app.route(nis_api_base + "/isession/rsession/generator", methods=["POST"])
def reproducible_session_append_command_generator():  # Receive a command_executors generator, like a Spreadsheet file, an R script, or a full JSON command_executors list (or other)

    import time
    logging.debug("### SUBMISSION STARTS ###")
    start = time.time()

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open
    if isess.reproducible_session_opened():
        # Reset!!
        # TODO Maybe do this only when some parameter is True
        reset_state_and_reproducible_session(isess)

        # Add system-level entities from JSON definition in "default_cmds"
        ret = isess.register_andor_execute_command_generator("json", "application/json", nexinfosys.default_cmds, False, True)

        # Check that objects have been properly registered
        # glb_idx, p_sets, hh, datasets, mappings = get_case_study_registry_objects(isess._state)
        # ps = glb_idx.get(Parameter.partial_key())
        # hs = glb_idx.get(Hierarchy.partial_key())

        # PARSE AND BUILD!!!
        generator_type, content_type, buffer, execute, register = receive_file_submission(request)

        try:
            ret = isess.register_andor_execute_command_generator(generator_type, content_type, buffer, register, execute)
            if isinstance(ret, tuple):
                issues = ret[0]
            else:
                issues = []

            # TODO CHECK SEMANTIC INCONSISTENCIES. Referred to values in Interfaces use either Parameters

            # SOLVE !!!!
            if not any_error_issue(issues):
                issues2 = prepare_and_solve_model(isess.state)
                issues.extend(issues2)
        except Exception as e:
            traceback.print_exc()  # Print the Exception to std output
            # Obtain trace as string; split lines in string; take the last three lines
            tmp = traceback.format_exc().splitlines()
            for i in range(len(tmp)-3, 0, -2):
                if tmp[i].find("nexinfosys") != -1:
                    tmp = [tmp[-1], tmp[i], tmp[i+1]]
                    break
            else:
                tmp = [tmp[-1], "Nexinfosys module not found", "Line not found"]
            exc_info = ' :: '.join([s.strip() for s in tmp])
            # Error Issue with the extracted Exception text
            issues = [Issue(itype=IType.ERROR,
                            description=f"UNCONTROLLED CONDITION: {exc_info}. Please, contact the development team.",
                            location=None)]

        # STORE the issues in the state
        # TODO If issues are produced by different generators, this will overwrite results from the previous generator
        isess.state.set("_issues", issues)

        # Return the issues if there were any.
        # TODO Return outputs (could be a list of binary files)

        r = build_json_response({"issues": convert_issues(issues), "outputs": None}, 200)

        # TODO Important!!! The R script generator can be executed remotely and locally. In the first case, it
        # TODO could be desired to store commands. But the library, when executed at the server, will be passed a flag
        # TODO to perform every call with the registering disabled.
        printNProcessors("SUBMISSION", isess.state)

        serialize_isession_and_close_db_session(isess)
    else:
        r = build_json_response({"error": "A reproducible session must be open in order to submit a generator"}, 400)

    endt = time.time()
    logging.debug(F"### SUBMISSION FINISHED: {endt-start} ###")

    return r


@app.route(nis_api_base + "/isession/rsession/ensure_executed", methods=["PUT"])
def reproducible_session_execute_not_executed_command_generators():  # Executes commands pending execution
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # A reproducible session must be open
    if isess.reproducible_session_opened():
        # TODO From last to first generator, find the first one NOT executed
        # TODO Execute them, one after the other
        # TODO If the case study is persisted, Store it again
        first_i = len(isess.reproducible_session.ws_commands)
        for i in range(len(isess.reproducible_session.ws_commands)-1, -1, -1):
            c = isess.reproducible_session.ws_commands[i]
            if not c.execution_start:
                first_i = i
        if first_i < len(isess.reproducible_session.ws_commands):
            # Execute!
            persist_version_state = None
            executed_cmds = []
            for i in range(first_i, len(isess.reproducible_session.ws_commands)):
                c = isess.reproducible_session.ws_commands[i]
                if persist_version_state is None:
                    persist_version_state = c.id is not None  # Persist if the command is already persisted
                # The state is modified
                try:
                    ret = isess.register_andor_execute_command_generator1(c, register=False, execute=True)
                    executed_cmds.append(c)
                except:
                    ret = ([("error", "Command execution did not end due to an Exception")])

                if isinstance(ret, tuple):
                    issues = ret[0]
                else:
                    issues = []

                # STORE the issues in the state
                # TODO If issues are produced by different generators, this will overwrite results from the previous generator
                isess.state.set("_issues", issues)
            if persist_version_state:  # TODO Does this work as expected?
                isess.reproducible_session.update_current_version_state(executed_cmds)

            # Return the issues if there were any.
            # TODO Return outputs (could be a list of binary files)
            r = build_json_response({"issues": issues, "outputs": None, "everything_executed": False}, 200)
        else:
            r = build_json_response({"everything_executed": True}, 200)


        # TODO Important!!! The R script generator can be executed remotely and locally. In the first case, it
        # TODO could be desired to store commands. But the library, when executed at the server, will be passed a flag
        # TODO to perform every call with the registering disabled.
        serialize_isession_and_close_db_session(isess)
    else:
        r = build_json_response({"error": "A reproducible session must be open in order to execute generators"}, 400)

    return r


# -- Reproducible Session Query --
# - INSTEAD OF COMMANDS, DIRECT EXECUTION (NOT REGISTERED)

# -----
@app.route(nis_api_base + '/nis_files.json', methods=['GET'])
def list_of_registered_nis_files():
    """
    Return a list of either importable or example NIS files

    :return:
    """
    # Elaborate a list of dictionaries
    files = []
    if "NIS_FILES_LIST" in app.config:
        example_files = app.config["NIS_FILES_LIST"].split(",")
        for url in example_files:
            try:
                tmp = download_file(url).getvalue().decode("UTF-8")
                df = pd.read_csv(io.StringIO(tmp), skipinitialspace=True, quoting=csv.QUOTE_ALL)
                df.columns = [c.strip().lower() for c in df.columns]
                for t in df.iterrows():
                    desc = ""
                    files.append(dict(name=t[1]["name"], url=t[1]["url"], example=t[1]["example"], description=desc))
            except Exception as e:
                traceback.print_exc()  # Print the Exception to std output

    if len(files) == 0:
        files = [
            dict(name="MuSIASEM hierarchies", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP4/D4.3%20global%20food%20supply%20and%20diets/README.md", example=False, description=""),
            dict(name="Water grammar", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP4/D4.3%20global%20food%20supply%20and%20diets/README.md", example=False, description=""),
            dict(name="Energy grammar", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP4/D4.3%20global%20food%20supply%20and%20diets/README.md", example=False, description=""),
            dict(name="Food grammar", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP4/D4.3%20global%20food%20supply%20and%20diets/README.md", example=False, description=""),
            dict(name="Biofuel", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP6/CS6_2_Biofuels/README.md", example=True, description=""),
            dict(name="Gran Canaria", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP6/CS6_6_Alternative_water_sources/README.md", example=True, description=""),
            dict(name="Tenerife", url="https://nextcloud.data.magic-nexus.eu/remote.php/webdav/NIS_internal/WP6/CS6_6_Alternative_water_sources/README.md", example=True, description=""),
        ]

    return build_json_response(files, 200)


# -- Case studies --


@app.route(nis_api_base + '/case_studies/', methods=['POST'])
def new_case_study_from_file():
    """
    Check that the user is authorized to submit a new case study
    Open a reproducible session
    Send the file to the service
    Close the reproducible session

    :return:
    """
    # Check Interactive Session is Open. If not, open it
    isess = deserialize_isession_and_prepare_db_session(False)
    if not isess:
        isess = InteractiveSession(DBSession)

    # TODO Check User Credentials (from Token)
    testing = is_testing_enabled()
    if testing:
        result = isess.identify({"user": "test_user", "password": None}, testing=True)

    # TODO Check User has Create New Case Study permissions

    # Receive file
    generator_type, content_type, buffer, execute, register = receive_file_submission(request)

    # Open Reproducible Session, NEW case study
    try:
        isess.open_reproducible_session(case_study_version_uuid=None,
                                        recover_previous_state=False,
                                        cr_new=CreateNew.CASE_STUDY,
                                        allow_saving=register
                                        )
    except Exception as e:
        s = "Exception trying to open reproducible session: "+str(e)
        logger.error(s)
        return build_json_response({"error": s}, 401)

    # Submit file to the Interactive Session (which has the open reproducible session)
    issues, output = isess.register_andor_execute_command_generator(generator_type, content_type, buffer, register, execute)

    # Close Reproducible Session
    isess.close_reproducible_session(issues=issues, output=output, save=register, from_web_service=False)

    # TODO Return the issues if there were any. Return outputs (could be a list of binary files)
    r = build_json_response({}, 204)

    serialize_isession_and_close_db_session(isess)

    return r


@app.route(nis_api_base + "/case_studies/", methods=["GET"])
def case_studies():  # List case studies available for current user
    """
Example:
[
{"resource": "/case_studies/<case study uuid>",
 "uuid": "<uuid>",
 "name": "Food in the EU",
 "oid": "zenodo.org/2098235",
 "internal_code": "CS1_F_E",
 "description": "...",
 "stats":
  {
   "n_versions": "<# of versions>",
   "n_commands": "<# of command_executors latest version>",
   "n_hierarchies": <# of hierarchies latest version>",
  }
 "versions": "/case_studies/<uuid>/short.json"
 "thumbnail": "/case_studies/<uuid>/thumbnail.svg|html|png"
},
...
]

    :return:
    """

    def get_avatar_path(cstudy):
        """
        From the areas of a case study, obtain the file name representing these areas

        :param cstudy: CaseStudy object
        :return: String with the URL subpath to the file name
        """
        areas = cstudy.areas
        if areas:
            name = ""
            if "W" in areas:
                name += "Water"
            if "E" in areas:
                name += "Energy"
            if "F" in areas:
                name += "Food"
            return "/static/images/" + name + "Nexus.png"
        else:
            return "/static/images/NoNexusAreas.png"  # TODO create this image

    def get_version_dict(vs):
        uuid3 = str(vs.uuid)
        authors = ", ".join([ss.who.name for ss in vs.sessions if ss.who])
        version = {"uuid": uuid3,
                   "cs_uuid": str(vs.case_study.uuid), # Redundant but helps in the user interface, to obtain the CS UUID
                   "authors": authors,
                   "creation_date": vs.creation_instant.isoformat(timespec="seconds") if vs.creation_instant else "<no date available>",
                   "cs_name": vs.name,
                   "name": vs.creation_instant.isoformat(sep=" ", timespec="seconds") + " [" +authors + "]",
                   "resource": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3,
                   "detail": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3,
                   "issues": None,  # [{"type": "error", "description": "syntax error in command ..."}],
                   }
        return version

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()

    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    if not isess:
        session = DBSession()
    else:
        session = isess.open_db_session()
    # TODO Obtain case studies FILTERED by current user permissions. Show case studies with READ access enabled
    # TODO Access Control
    # TODO CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # TODO CS in acl and group acl.detail and user in group
    lst = session.query(CaseStudy).all()
    lst2 = []
    for cs in lst:
        uuid2 = str(cs.uuid)
        vs_lst = sorted([get_version_dict(v) for v in cs.versions], key=lambda v: v["creation_date"], reverse=True)
        name = vs_lst[0]["cs_name"]
        d = {"resource": nis_api_base + "/case_studies/"+uuid2,
             "uuid": uuid2,
             "name": name if name else "<empty>",
             "oid": cs.oid if cs.oid else "<empty>",  # TODO
             "internal_code": cs.internal_code if cs.internal_code else "",  # TODO
             "description": cs.description if cs.description else "",  # TODO
             "stats": {
                 "n_versions": str(len(cs.versions)),
                 "n_commands": str(len([])),  # TODO
                 "n_hierarchies": str(len([])),  # TODO
             },
             "versions": vs_lst,
             "thumbnail": nis_api_base + "/case_studies/" + uuid2 + "/default_view.png",
             "thumbnail_png": nis_api_base + "/case_studies/" + uuid2 + "/default_view.png",
             "thumbnail_svg": nis_api_base + "/case_studies/" + uuid2 + "/default_view.svg",
             "avatar": nis_api_base + get_avatar_path(cs),  # Icon representing the type of Nexus study
             "case_study_permissions":
                 {
                     "read": True,
                     "annotate": True,
                     "contribute": True,
                     "share": False,
                     "delete": False
                 }
             }
        lst2.append(d)
        # print(json.dumps(lst2, default=json_serial, sort_keys=True, indent=JSON_INDENT, ensure_ascii=ENSURE_ASCII, separators=(',', ': '))
        #       )
    r = build_json_response(lst2)  # TODO Improve it, it must return the number of versions. See document !!!
    if isess:
        isess.close_db_session()
    else:
        DBSession.remove()

    return r


@app.route(nis_api_base + "/case_studies2/", methods=["GET"])
def case_studies2():  # List case studies
    """
Example:
[
{"resource": "/case_studies/<case study uuid>",
 "uuid": "<uuid>",
 "name": "Food in the EU",
 "oid": "zenodo.org/2098235",
 "internal_code": "CS1_F_E",
 "description": "...",
 "stats":
  {
   "n_versions": "<# of versions>",
   "n_commands": "<# of command_executors latest version>",
   "n_hierarchies": <# of hierarchies latest version>",
  }
 "versions": "/case_studies/<uuid>/short.json"
 "thumbnail": "/case_studies/<uuid>/thumbnail.svg|html|png"
},
...
]

    :return:
    """

    def get_avatar_path(cstudy):
        """
        From the areas of a case study, obtain the file name representing these areas

        :param cstudy: CaseStudy object
        :return: String with the URL subpath to the file name
        """
        areas = cstudy.areas
        if areas:
            name = ""
            if "W" in areas:
                name += "Water"
            if "E" in areas:
                name += "Energy"
            if "F" in areas:
                name += "Food"
            return "/static/images/" + name + "Nexus.png"
        else:
            return "/static/images/NoNexusAreas.png"  # TODO create this image

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()

    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    if not isess:
        session = DBSession()
    else:
        session = isess.open_db_session()
    # TODO Obtain case studies FILTERED by current user permissions. Show case studies with READ access enabled
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    base = app.config["APPLICATION_ROOT"]
    lst = session.query(CaseStudy).all()
    lst2 = []
    for cs in lst:
        uuid2 = str(cs.uuid)
        d = {"resource": nis_api_base + "/case_studies/"+uuid2,
             "uuid": uuid2,
             "name": cs.name if cs.name else "<empty>",
             "oid": cs.oid if cs.oid else "<empty>",  # TODO
             "internal_code": cs.internal_code if cs.internal_code else "",  # TODO
             "description": cs.description if cs.description else "",  # TODO
             "stats": {
                 "n_versions": str(len(cs.versions)),
                 "n_commands": str(len([])),  # TODO
                 "n_hierarchies": str(len([])),  # TODO
             },
             "versions": nis_api_base + "/case_studies/" + uuid2 + "/versions/",
             "thumbnail": nis_api_base + "/case_studies/" + uuid2 + "/default_view.png",
             "thumbnail_png": nis_api_base + "/case_studies/" + uuid2 + "/default_view.png",
             "thumbnail_svg": nis_api_base + "/case_studies/" + uuid2 + "/default_view.svg",
             "avatar": nis_api_base + get_avatar_path(cs),  # Icon representing the type of Nexus study
             "case_study_permissions":
                 {
                     "read": True,
                     "annotate": True,
                     "contribute": True,
                     "share": False,
                     "delete": False
                 }
             }
        lst2.append(d)
        # print(json.dumps(lst2, default=json_serial, sort_keys=True, indent=JSON_INDENT, ensure_ascii=ENSURE_ASCII, separators=(',', ': '))
        #       )
    r = build_json_response(lst2)  # TODO Improve it, it must return the number of versions. See document !!!
    if isess:
        isess.close_db_session()
    else:
        DBSession.remove()

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>", methods=["GET"])
@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/", methods=["GET"])
def case_study(cs_uuid):  # Information about case study
    """

{"case_study": "<uuid>",
 "name": "Food in the EU",
 "oid": "zenodo.org/2098235",
 "internal_code": "CS1_F_E",
 "resource": "/case_studies/<case study uuid>",
 "description": "...",
 "versions":
 [
  {"uuid": "<uuid>",
   "resource": "/case_studies/<case study uuid>/<version uuid>",
   "tag": "v0.1",
   "sessions":
   [
    {"uuid": "<uuid>",
     "open_date": "2017-09-20T10:00:00Z",
     "close_date": "2017-09-20T10:00:10Z",
     "client": "spreadsheet",
     "restart": True,
     "author": "<uuid>",
    },
    ...
   ]
   "detail": "/case_studies/<case study uuid>/<version uuid>/long.json"
   "generator": "/case_studies/<case study uuid>/<version uuid>/generator.xlsx",
   "state": "/case_studies/<case study uuid>/<version uuid>/state.xlsx",
   "issues": [{"type": "error", "description": "syntax error in command ..."}, ...],
  },
  ...
 ]
}

    :param cs_uuid:
    :return:
    """
    def get_version_dict(vs):
        # [
        #     {"uuid": "<uuid>",
        #      "resource": "/case_studies/<case study uuid>/<version uuid>",
        #      "tag": "v0.1",
        #      "sessions":
        #          [
        #              {"uuid": "<uuid>",
        #               "open_date": "2017-09-20T10:00:00Z",
        #               "close_date": "2017-09-20T10:00:10Z",
        #               "client": "spreadsheet",
        #               "restart": True,
        #               "author": "<uuid>",
        #               },
        #          ],
        #      "detail": "/case_studies/<case study uuid>/<version uuid>/long.json",
        #      "generator": "/case_studies/<case study uuid>/<version uuid>/generator.xlsx",
        #      },
        # ],
        def get_session_dict(ss):
            uuid4 = str(ss.uuid)
            v_session = {"uuid": uuid4,
                         "open_date": str(ss.open_instant),
                         "close_date": str(ss.close_instant),
                         "client": "spreadsheet",  # TODO Spreadsheet, R script, Python script, <Direct?>
                         "restart": ss.restarts,
                         "author": ss.who.name
                         }
            if mode == "tree":
                v_session = {"data": v_session}
            else:
                pass
            return v_session

        uuid3 = str(vs.uuid)
        version = {"uuid": uuid3,
                   "resource": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3,
                   "tag": "v0.1",
                   "detail": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3,
                   "state": nis_api_base + "/case_studies/" + uuid2 + "/versions/"+uuid3+"/state.xlsx",
                   "issues": None,  # [{"type": "error", "description": "syntax error in command ..."}],
                   "generator": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3+"/generator.xlsx",
                   }
        if mode == "tree":
            version = {"data": version, "children": [get_session_dict(s) for s in vs.sessions]}
        else:
            version["sessions"] = [get_session_dict(s) for s in vs.sessions]
        return version

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"

    mode = "tree"

    # Recover case studies READABLE by current user (or "anonymous")
    session = isess.open_db_session()
    # TODO Obtain case study, filtered by current user permissions
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    cs = session.query(CaseStudy).filter(CaseStudy.uuid == cs_uuid).first()
    if cs:
        uuid2 = str(cs.uuid)
        d = {"uuid": uuid2,
             "name": cs.name if cs.name else "<empty>",
             "oid": cs.oid if cs.oid else "<empty>",
             "internal_code": cs.internal_code if cs.internal_code else "",  # TODO
             "description": cs.description if cs.description else "",  # TODO
             "resource": nis_api_base + "/case_studies/"+uuid2,
             "versions": [get_version_dict(v) for v in cs.versions],
             "case_study_permissions":
                 {
                     "read": True,
                     "annotate": True,
                     "contribute": True,
                     "share": False,
                     "delete": False
                 },

             }
        # print(json.dumps(d, default=json_serial, sort_keys=True, indent=JSON_INDENT, ensure_ascii=ENSURE_ASCII, separators=(',', ': ')))
        r = build_json_response(d)
    else:
        r = build_json_response({"error": "The case study '"+cs_uuid+"' does not exist."}, 404)
    isess.close_db_session()

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>", methods=["POST"])
@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/", methods=["POST"])
def new_case_study_version_from_file(cs_uuid):
    """
    Check that the user is authorized to submit a new case study version
    Open a reproducible session
    Send the file to the service
    Close the reproducible session

    :param cs_uuid: UUID of case study
    :return:
    """
    # Check Interactive Session is Open. If not, open it
    isess = deserialize_isession_and_prepare_db_session(False)
    if not isess:
        isess = InteractiveSession(DBSession)

    # TODO Check User Credentials (from Token)
    testing = is_testing_enabled()
    if testing:
        result = isess.identify({"user": "test_user", "password": None}, testing=True)

    # TODO Check User has Write Case Study permissions

    # Receive file
    generator_type, content_type, buffer, execute, register = receive_file_submission(request)

    # Open Reproducible Session, NEW case study
    try:
        isess.open_reproducible_session(case_study_version_uuid=cs_uuid,
                                        recover_previous_state=False,
                                        cr_new=CreateNew.VERSION,
                                        allow_saving=register
                                        )
    except Exception as e:
        s = "Exception trying to open reproducible session: "+str(e)
        logger.error(s)
        return build_json_response({"error": s}, 401)


    # Submit file to the Interactive Session (which has the open reproducible session)
    issues, output = isess.register_andor_execute_command_generator(generator_type, content_type, buffer, register, execute)

    # Close Reproducible Session
    isess.close_reproducible_session(issues=issues, output=output, save=register, from_web_service=False)

    # TODO Return the issues if there were any. Return outputs (could be a list of binary files)
    r = build_json_response({}, 204)

    serialize_isession_and_close_db_session(isess)

    return r

# @app.route(nis_api_base + "/case_studies/<cs_uuid>", methods=["DELETE"])
# def case_study_delete(cs_uuid):  # DELETE a case study
#     # Recover InteractiveSession
#     isess = deserialize_isession_and_prepare_db_session()
#     if isess and isinstance(isess, Response):
#         return isess
#
#     # TODO Check permissions
#     # TODO If possible, deleet ALL the case study


@app.route(nis_api_base + "/case_studies/<cs_uuid>/default_view.png", methods=["GET"])
def case_study_default_view_png(cs_uuid):  # Return a view of the case study in PNG format, for preview purposes
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"

    return send_static_file("images/case_study_preview_placeholder.png")

    # # Recover case studies READABLE by current user (or "anonymous")
    # session = isess.open_db_session()
    # # TODO Obtain case study, filtered by current user permissions
    # # Access Control
    # # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # # CS in acl and group acl.detail and user in group
    # cs = session.query(CaseStudy).filter(CaseStudy.uuid == cs_uuid).first()
    # # TODO Scan variables. Look for the ones most interesting: grammar, data. Maybe cut processors.
    # # TODO Scan also for hints to the elaboration of this thumbnail
    # # TODO Elaborate View in PNG format
    # isess.close_db_session()
    # # TODO Return PNG image


@app.route(nis_api_base + "/case_studies/<cs_uuid>/default_view.svg", methods=["GET"])
def case_study_default_view_svg(cs_uuid):  # Return a view of the case study in SVG format, for preview purposes
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    session = isess.open_db_session()
    # TODO Obtain case study, filtered by current user permissions
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    cs = session.query(CaseStudy).filter(CaseStudy.uuid == cs_uuid).first()
    # TODO Scan variables. Look for the ones most interesting: grammar, data. Maybe cut processors.
    # TODO Scan also for hints to the elaboration of this thumbnail
    # TODO Elaborate View in SVG format
    isess.close_db_session()
    # TODO Return SVG image


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>", methods=["GET"])
def case_study_version(cs_uuid, v_uuid):  # Information about a case study version
    """

{"case_study": "<uuid>",
 "version": "<uuid>",
 "resource": "/case_studies/<case study uuid>/<version uuid>",
 "tag": "v0.1",
 "sessions":
   [
    {"uuid": "<uuid>",
     "open_date": "2017-09-20T10:00:00Z",
     "close_date": "2017-09-20T10:00:10Z",
     "client": "spreadsheet",
     "restart": True,
     "author": "<uuid>",
     "generator": "/case_studies/<case study uuid>/<version uuid>/<session uuid>/generator.xlsx",
     "state": "/case_studies/<case study uuid>/<version uuid>/<session uuid>/state.xlsx",
     "issues": [{"type": "error", "description": "syntax error in command ..."}, ...],
    },
    ...
   ]
 "command_executors":
 [
  {"type": "...",
   "label": "...",
   "definition": "/case_studies/<case study uuid>/<version uuid>/1.json"
  },
  ...
 ],
 "generator": "/case_studies/<case study uuid>/<version uuid>/generator.xlsx",
 "state": "/case_studies/<case study uuid>/<version uuid>/state.xlsx",
 "issues": [{"type": "error", "description": "syntax error in command ..."}, ...],
}

    :param cs_uuid:
    :param v_uuid:
    :return:
    """

    def get_version_dict(vs):
        # [
        #     {"uuid": "<uuid>",
        #      "resource": "/case_studies/<case study uuid>/<version uuid>",
        #      "tag": "v0.1",
        #      "sessions":
        #          [
        #              {"uuid": "<uuid>",
        #               "open_date": "2017-09-20T10:00:00Z",
        #               "close_date": "2017-09-20T10:00:10Z",
        #               "client": "spreadsheet",
        #               "restart": True,
        #               "author": "<uuid>",
        #               },
        #          ],
        #      "detail": "/case_studies/<case study uuid>/<version uuid>/long.json",
        #      "generator": "/case_studies/<case study uuid>/<version uuid>/generator.xlsx",
        #      },
        # ],
        def get_session_dict(ss):
            uuid4 = str(ss.uuid)
            v_session = {"uuid": uuid4,
                         "open_date": str(ss.open_instant),
                         "close_date": str(ss.close_instant),
                         "client": "spreadsheet",  # TODO Spreadsheet, R script, Python script, <Direct?>
                         "restart": ss.restarts,
                         "author": ss.who.name
                         }
            if mode == "tree":
                v_session = {"data": v_session}
            else:
                pass
            return v_session

        # Case Study UUID, Case Study Version UUID
        uuid2 = str(vs.case_study.uuid)
        uuid3 = str(vs.uuid)

        # Get active sessions
        act_sess = []
        for s in vs.sessions:
            if s.restarts:
                act_sess = []
            act_sess.append(s)

        # Load state (or EXECUTE IT!!! -CAN BE VERY SLOW!!-)
        if vs.state:
            # Deserialize
            st = deserialize_to_object(vs.state)
        else:
            st = State()  # Zero State, execute all commands in sequence
            for ws in act_sess:
                for c in ws.commands:
                    execute_command_container(st, c)

        # List command_executors lista -> diccionario ("data": {}, "children": [ ... ])
        lst_cmds = []
        for ws in act_sess:
            for c in ws.commands:
                d = {"type": c.generator_type,
                     "label": c.name if c.name else "<empty>",
                     "definition": nis_api_base + "/case_studies/" + uuid2 + "/versions/"+uuid3+"/sessions/"+str(ws.uuid)+"/command/"+str(c.order)
                     }
                if mode == "tree":
                    d = {"data": d}
                lst_cmds.append(d)

        # List of variables
        lst_vars = []
        for n in st.list_namespaces():
            for t in st.list_namespace_variables(n):
                d = {"name": t[0],
                     "type": str(type(t[1])),
                     "view": nis_api_base + "/case_studies/" + uuid2 + "/versions/"+uuid3+"/variables/"+str(t[0]),
                     "namespace": n
                     }
                if mode == "tree":
                    d = {"data": d}
                lst_vars.append(d)

        version = {"case_study": uuid2,
                   "version": uuid3,
                   "resource": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3,
                   "tag": "v0.1",
                   "generator": nis_api_base + "/case_studies/"+uuid2+"/versions/"+uuid3+"/generator.xlsx",
                   "state": nis_api_base + "/case_studies/" + uuid2 + "/versions/"+uuid3+"/state.xlsx",
                   "issues": [{"type": "error", "description": "syntax error in command ..."}
                              ],
                   "sessions": [get_session_dict(s) for s in vs.sessions],
                   "command_executors": lst_cmds,
                   "variables": lst_vars
                   }
        # if mode == "tree":
        #     version = {"data": version, "children": [get_session_dict(s) for s in vs.sessions]}
        # else:
        #     version["sessions"] = [get_session_dict(s) for s in vs.sessions]
        return version

    mode = "tree"

    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    session = isess.open_db_session()
    # TODO Obtain case study version, filtered by current user permissions
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    vs = session.query(CaseStudyVersion).filter(CaseStudyVersion.uuid == v_uuid).first()
    if not vs:
        r = build_json_response({"error": "The case study version '"+v_uuid+"' does not exist."}, 404)
    else:
        if str(vs.case_study.uuid) != cs_uuid:
            r = build_json_response({"error": "The case study '" + cs_uuid + "' does not exist."}, 404)
        else:
            r = build_json_response(get_version_dict(vs))
    isess.close_db_session()

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>", methods=["DELETE"])
def case_study_version_delete(cs_uuid, v_uuid):  # DELETE a case study version
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check user permissions
    # TODO If authorized, delete a case study version and all its sessions and commands


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/sessions/<s_uuid>", methods=["GET"])
def case_study_version_session(cs_uuid, v_uuid, s_uuid):  # Information about a session in a case study version
    """

{"case_study": "<uuid>",
 "version": "<uuid>",
 "session": "<uuid>",
 "resource": "/case_studies/<case study uuid>/<version uuid>/<session uuid>",
 "open_date": "2017-09-20T10:00:00Z",
 "close_date": "2017-09-20T10:00:10Z",
 "client": "spreadsheet",
 "restart": True,
 "author": "<uuid>",
 "generator": "/case_studies/<case study uuid>/<version uuid>/<session uuid>/generator.xlsx",
 "state": "/case_studies/<case study uuid>/<version uuid>/<session uuid>/state.xlsx",
 "issues": [{"type": "error", "description": "syntax error in command ..."}, ...],
 "command_executors":
 [
  {"type": "...",
   "label": "...",
   "definition": "/case_studies/<case study uuid>/<version uuid>/1.json"
  },
  ...
 ]
}

    :param cs_uuid:
    :param v_uuid:
    :param s_uuid:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    session = isess.open_db_session()
    # TODO Obtain case study version session, filtered by current user permissions
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    ss = session.query(CaseStudyVersionSession).filter(CaseStudyVersionSession.uuid == s_uuid).first()
    if not ss:
        r = build_json_response({"error": "The case study version session '"+s_uuid+"' does not exist."}, 404)
    else:
        if ss.version.uuid != v_uuid:
            r = build_json_response({"error": "The case study version '" + v_uuid + "' does not exist."}, 404)
        elif ss.version.case_study.uuid != cs_uuid:
            r = build_json_response({"error": "The case study '" + cs_uuid + "' does not exist."}, 404)
        else:
            # TODO Return the command OR generator
            # TODO The generator can be text or BINARY
            r = build_json_response(ss)  # TODO Improve it, it must return the number of versions. See document !!!
    isess.close_db_session()

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/sessions/<s_uuid>", methods=["DELETE"])
def case_study_version_session_delete(cs_uuid, v_uuid, s_uuid):  # DELETE a session in a case study version
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check user permissions
    # TODO If authorized, delete a case study version SESSION and all its commands


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/sessions/<s_uuid>/<command_order>", methods=["GET"])
def case_study_version_session_command(cs_uuid, v_uuid, s_uuid, command_order):
    """
        DOWNLOAD a command or generator, using the order, from 0 to number of command_executors - 1
        Commands are enumerated using "case_study_version_session()"
            (URL: "/case_studies/<cs_uuid>/versions/<v_uuid>/sessions/<s_uuid>")

    :param cs_uuid:
    :param v_uuid:
    :param s_uuid:
    :param command_order:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if not user:
        user = "_anonymous"
    # Recover case studies READABLE by current user (or "anonymous")
    session = isess.open_db_session()
    # TODO Obtain case study version session, filtered by current user permissions
    # Access Control
    # CS in acl and user in acl.detail and acl.detail is READ, WRITE,
    # CS in acl and group acl.detail and user in group
    if s_uuid == "-":
        vs = session.query(CaseStudyVersion).filter(CaseStudyVersion.uuid == v_uuid).first()
        if vs:
            ss = vs.sessions[0]
            s_uuid = ss.uuid
    else:
        ss = session.query(CaseStudyVersionSession).filter(CaseStudyVersionSession.uuid == s_uuid).first()
    if not ss:
        r = build_json_response({"error": "The case study version session '"+s_uuid+"' does not exist."}, 404)
    else:
        # if ss.version.uuid != v_uuid:
        #     r = build_json_response({"error": "The case study version '" + v_uuid + "' does not exist."}, 404)
        # elif ss.version.case_study.uuid != cs_uuid:
        #     r = build_json_response({"error": "The case study '" + cs_uuid + "' does not exist."}, 404)
        order = int(command_order)
        if order < len(ss.commands):
            c = ss.commands[order]
            r = Response(c.content, mimetype=c.content_type)
        else:
            r = build_json_response({"error":
                                     "Command number " + str(order) +
                                     " requested. The session '"+s_uuid+"' only has "+str(len(ss.commands))+"."})
        # r.headers['Access-Control-Allow-Origin'] = "*"

    isess.close_db_session()

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/variables/", methods=["GET"])
def case_study_version_variables(cs_uuid, v_uuid):  # List of variables defined in a case study version
    """
    Return the list of ALL variables defined in the case study version

    :param cs_uuid:
    :param v_uuid:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check READ permission of the user to the case study
    # Open temporary reproducible session
    try:
        isess.open_reproducible_session(case_study_version_uuid=v_uuid,
                                        recover_previous_state=True,
                                        cr_new=CreateNew.NO,
                                        allow_saving=False
                                        )

        # A reproducible session must be open, signal about it if not
        if isess.state:
            # List all available variables, from state. A list of dictionaries "name", "type" and "namespace"
            lst = []
            for n in isess.state.list_namespaces():
                # Add ALL variables EXCEPT the internal ones (which start with "_")
                lst.extend([{"name": t[0],
                             "type": str(type(t[1])),
                             "namespace": n} for t in isess.state.list_namespace_variables(n) if not t[0].startswith("_")
                            ]
                           )

            r = build_json_response(lst, 200)
        else:
            r = build_json_response({"error": "No state available for Case Study Version '"+v_uuid+"'"}, 404)

        # Close temporary reproducible session
        isess.close_reproducible_session(issues=None, output=None, save=False, from_web_service=True)
    except Exception as e:
        r = build_json_response({"error": str(e)}, 404)

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/variables/<name>", methods=["GET"])
def case_study_version_variable(cs_uuid, v_uuid, name):  # Information about a case study version variable
    """
    Return the value of the requested variable

    :param cs_uuid: Case Study UUID
    :param v_uuid: Version UUID
    :param name: Variable name
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check READ permission of the user to the case study
    # Open temporary reproducible session
    try:
        isess.open_reproducible_session(case_study_version_uuid=v_uuid,
                                        recover_previous_state=True,
                                        cr_new=CreateNew.NO,
                                        allow_saving=False
                                        )

        # A reproducible session must be open, signal about it if not
        if isess.state:
            # TODO Parse Variable name can be "namespace'::'name"
            # TODO For now, just the variable name
            v = isess.state.get(name)
            if v:
                r = build_json_response({name: v}, 200)
            else:
                r = build_json_response(
                    {"error": "The requested variable name ('"+name+"') has not "
                              "been found in the Case Study Version '" + v_uuid + "'"}, 404)
        else:
            r = build_json_response({"error": "No state available for Case Study Version '" + v_uuid + "'"}, 404)

        # Close temporary reproducible session
        isess.close_reproducible_session(issues=None, output=None, save=False, from_web_service=True)
    except Exception as e:
        r = build_json_response({"error": str(e)}, 404)

    return r


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/variables/<name>/views/", methods=["GET"])
def case_study_version_variable_views(cs_uuid, v_uuid, name):  # Information about a case study version variable views
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check READ permission of the user to the case study
    # TODO Return the different views on a variable


@app.route(nis_api_base + "/case_studies/<cs_uuid>/versions/<v_uuid>/variables/<name>/views/<view_type>", methods=["GET"])
def case_study_version_variable_view(cs_uuid, v_uuid, name, view_type):  # A view of case study version variable
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Check READ permission of the user to the case study
    # TODO Return a view of the requested variable

# -- Users --


@app.route(nis_api_base + "/users/", methods=["GET"])
def list_users():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin":
        session = isess.open_db_session()
        lst = session.query(User).all()
        r = build_json_response(lst)
        isess.close_db_session()
    else:
        r = build_json_response({"error": "Users list can be obtained only by 'admin' user"}, 401)

    return r


@app.route(nis_api_base + "/users/<id>", methods=["GET"])
def get_user(id):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin" or user == id:
        session = isess.open_db_session()
        u = session.query(User).filter(User.name == id).first()
        r = build_json_response(u)  # TODO Improve it !!!
        isess.close_db_session()
    else:
        r = build_json_response({"error": "User '"+id+"' can be obtained only by 'admin' or '"+id+"' user"}, 401)

    return r


@app.route(nis_api_base + "/users/<id>", methods=["PUT"])
def put_user(id):  # Used also to deactivate user
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin" or user == id:
        session = isess.open_db_session()
        u = session.query(User).filter(User.name == id).first()
        if not u:
            r = build_json_response({"error": "User '"+id+"' does not exist"}, 404)
        else:
            # TODO Update "u" fields
            session.commit()
        r = build_json_response(u)  # TODO Improve it !!!
        isess.close_db_session()
    else:
        r = build_json_response({"error": "User '"+id+"' can be modified only by 'admin' or '"+id+"' user"}, 401)

    return r


@app.route(nis_api_base + "/users/", methods=["POST"])
def post_user():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin":
        session = isess.open_db_session()
        # Read JSON
        if request.content_type != "application/json":
            raise Exception("Only application/json data is allowed")
        if not request.data:
            raise Exception("No data received")
        j = request.data.decode()
        j = json.loads(j)
        u = session.query(User).filter(User.name == j["name"]).first()
        if not u:
            # Create User
            u = User()
            u.name = j["name"]
            session.add(u)
            session.commit()
            r = build_json_response(u)
        else:
            r = build_json_response({"error": "User '"+j["name"]+"' already exists"}, 422)
        isess.close_db_session()
    else:
        r = build_json_response({"error": "A user can be created only by 'admin'"}, 401)

    return r

# -- Groups --


@app.route(nis_api_base + "/groups/", methods=["GET"])
def list_groups():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess


@app.route(nis_api_base + "/groups/<id>", methods=["GET"])
def get_group(id):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin" or user == id:
        session = isess.open_db_session()
        u = session.query(Group).filter(Group.name == id).first()
        r = build_json_response(u)  # TODO Improve it !!!
        isess.close_db_session()
    else:
        r = build_json_response({"error": "Group '" + id + "' can be obtained only by 'admin' or '" + id + "' user"},
                                401)

    return r


@app.route(nis_api_base + "/groups/<id>", methods=["PUT"])
def put_group(id):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin" or user == id:
        session = isess.open_db_session()
        u = session.query(Group).filter(Group.name == id).first()
        if not u:
            r = build_json_response({"error": "Group '" + id + "' does not exist"}, 404)
        else:
            # TODO Update "u" fields
            session.commit()
        r = build_json_response(u)  # TODO Improve it !!!
        isess.close_db_session()
    else:
        r = build_json_response({"error": "Group '" + id + "' can be modified only by 'admin' or '" + id + "' user"},
                                401)

    return r


@app.route(nis_api_base + "/users/", methods=["POST"])
def post_group():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    user = isess.get_identity_id()
    if user and user == "admin":
        session = isess.open_db_session()
        # Read JSON
        if request.content_type != "application/json":
            raise Exception("Only application/json data is allowed")
        if not request.data:
            raise Exception("No data received")
        j = request.data.decode()
        j = json.loads(j)
        u = session.query(Group).filter(Group.name == j["name"]).first()
        if not u:
            # TODO Create Group
            u = Group()
            u.name = j["name"]
            session.add(u)
            session.commit()
            r = build_json_response(u)
        else:
            r = build_json_response({"error": "Group '" + j["name"] + "' already exists"}, 422)
        isess.close_db_session()
    else:
        r = build_json_response({"error": "A group can be created only by 'admin'"}, 401)

    return r


# -- Permissions --


def acl():
    pass

# -- Reusable objects --


@app.route(nis_api_base + "/sources/", methods=["GET"])
def data_sources():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # TODO Authentication, authorization

    # Enumerate sources
    dsm = nexinfosys.data_source_manager
    lst = dsm.get_supported_sources()

    return build_json_response(dict(sources=lst))


@app.route(nis_api_base + "/sources/<source_id>", methods=["GET"])
@app.route(nis_api_base + "/sources/<source_id>/databases/", methods=["GET"])
def data_source_databases(source_id):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Enumerate source databases
    dsm = nexinfosys.data_source_manager
    if source_id == "-":
        source_id = None
    lst = dsm.get_databases(source_id)

    ret_lst = []
    for i in lst:
        ret_lst.append(dict(source=i[0], databases=[dict(code=c.code, description=c.description) for c in i[1]]))

    return build_json_response(ret_lst)


@app.route(nis_api_base + "/sources/<source_id>/databases/<database_id>", methods=["GET"])
@app.route(nis_api_base + "/sources/<source_id>/databases/<database_id>/datasets/", methods=["GET"])
def data_source_database_datasets(source_id, database_id):
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Enumerate source+database datasets
    dsm = nexinfosys.data_source_manager
    if source_id == "-":
        source_id = None
        database_id = None
    if database_id in ("-", "None"):
        database_id = None
    lst = dsm.get_datasets(source_id, database_id)

    base = request.base_url+"/datasets/"
    return build_json_response([dict(source=i[0],
                                     datasets=[dict(code=j[0], description=j[1], info_url=base+j[0]) for j in i[1]]) for i in lst]
                               )


@app.route(nis_api_base + "/sources/<source_id>/databases/<database_id>/datasets/<dataset_id>", methods=["GET"])
def data_source_database_dataset_detail(source_id, database_id, dataset_id):
    """
    Return a JSON with the method "GET" and the possible values for the dimensions
    Also parameters to return a table of tuples or a precomputed pivot table
    Also return the address of the endpoint to query the dataset using SDMX. This be

    :param id:
    :param database_id:
    :param dataset_id:
    :return:
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    dsm = nexinfosys.data_source_manager
    if source_id == "-":
        source_id = None
        database_id = None
    if database_id == "-":
        database_id = None
    if not dataset_id:
        raise Exception("It is mandatory to define the dataset name when requesting the dataset parameters")

    if not source_id:
        _, _, _, datasets, _ = get_case_study_registry_objects(isess.state)
        from nexinfosys.ie_imports.data_source_manager import DataSourceManager
        source_id = DataSourceManager.obtain_dataset_source(dataset_id, datasets)

    ds = dsm.get_dataset_structure(source_id, dataset_id)
    dims = []
    for d in ds.dimensions:
        cl = []
        if d.get_hierarchy():
            # CodeList has one or more levels ".levels" property
            # CodeListLevel has zero or more Codes ".codes" property
            if isinstance(d.get_hierarchy, list):
                for v in d.get_hierarchy().codes:
                    cl.append(dict(code=v.name, description=v.description, level=v.level.name if v.level else None))
            else:  # Fix: codes can be in a dictionary
                for v in d.get_hierarchy().codes.values():
                    cl.append(dict(code=v.name, description=v.description, level=v.level.name if v.level else None))

        dims.append(dict(code=d.code, description=d.description, is_time=d.is_time, is_measure=d.is_measure, attributes=d.attributes, code_list=cl))

    d = dict(id=ds.id, code=ds.code, description=ds.description, data_dictionary=ds.data_dictionary, attributes=ds.attributes,
             dimensions=dims)

    return build_json_response(d)


@app.route(nis_api_base + "/isession/external_xslx", methods=["PUT"])
def download_external_xlsx():  # From the URL of an external XLSX, obtain it and return it
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    buffer = bytes(io.BytesIO(request.get_data()).getbuffer())
    url = buffer.decode("utf-8")
    data = download_file(url)
    try:
        xl = openpyxl.load_workbook(data, data_only=True)
        rewrite_xlsx_file(xl, copy_style=False)
    except Exception as e:
        logging.error("Exception rewriting XLSX. Is openpyxl==2.4.8 installed?. Check with 'pip freeze | grep "
                      "openpyxl'. If that is the case, fix with 'pip install openpyxl==2.4.8'")
        raise e

    data = save_virtual_workbook(xl)

    r = Response(data,
                 mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 status=200)

    return r


@app.route(nis_api_base + "/isession/regenerate_xlsx", methods=["POST"])
def regenerate_xlsx_file():  # Receive an XLSX workbook, regenerate it
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    generator_type, content_type, buffer, execute, register = receive_file_submission(request)
    try:
        xl = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
        rewrite_xlsx_file(xl)
        # rewrite_xlsx_file(xl, copy_style=False)
    except Exception as e:
        print("Exception rewriting XLSX. Is openpyxl==2.4.8 installed?. Check with 'pip freeze | grep openpyxl'. "
              "If that is the case, fix with 'pip install openpyxl==2.4.8'")
        raise e

    tmp = save_virtual_workbook(xl)
    r = Response(tmp,
                 mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 status=200)

    return r


@app.route(nis_api_base + "/commands_and_fields", methods=["GET"])
def obtain_commands_and_their_fields():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    j = {}
    for k, v in command_fields.items():
        j["name"] = k
        flds = []
        for f in v:
            flds.append(f.allowed_names)
        j["fields"] = flds
    return j


@app.route(nis_api_base + "/validate_command_record", methods=["POST"])
def validate_command_record():
    """
    A function for on-line, field by field or row by row validation of syntax
    (the client can send what the user just entered, the server, this function, will respond
     None the field is ok, and an error message if not)

    The input comes in a JSON field "content":
    {"command": "<command name",
     "fields": {"<field name>": "<value", ...}
    }

    :return: A dictionary with the same fields of the input dictionary, whose values are the diagnosis, None being
            everything-ok, and a string being a message describing the problem.
    """
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    # Read request
    command_content_to_validate = request.get_json()
    result, status = validate_command(command_content_to_validate)
    return build_json_response(result, 200 if status else 400)


def get_misc_cmd_help(cmd_name):
    if cmd_name == "metadata":
        return {"type": "Metadata", "by_rows": False, "name": "Metadata", "template":
"Case study code\n\
Case study name\n\
Title\n\
Subject, topic and/or keywords\n\
Description\n\
Geographical level\n\
Dimensions\n\
Reference documentation\n\
Authors\n\
Date of elaboration\n\
Temporal situation\n\
Geographical location\n\
DOI\n\
Language\n\
Restriction level\n\
Version", "examples": [
"Case study code\tCS3_R_WEF_P-0.1\n\
Case study name\n\
Title\tSoslaires\n\
Subject, topic and/or keywords\n\
Description\tA small scale system combining Energy, Water and Food\n\
Geographical level\tLocal\n\
Dimensions\tEnergy\tWater\tFood\n\
Reference documentation\n\
Authors\tAna Musicki\tBaltasar Peñate\tTarik Serrrano\n\
Date of elaboration\t2016\n\
Temporal situation\t2016\n\
Geographical location\tGran Canaria\n\
DOI\n\
Language\tEnglish\n\
Restriction level\tPublic\n\
Version\tV0.1"]}
    elif cmd_name == "pedigree_matrix":
        return {"type": "Metadata", "name": "Pedigree", "template":
"Code\t<Phase name #1>\t<Phase name #2>\t<Phase name #3>\t...",
         "examples": [
"Code\tTheoreticalStructures\tDataInput\tPeerAcceptance\tColleagueConsensus\n\
4\tEstablishedTheory\tExperimentalData\tTotal\tAllButCranks\n\
3\tTheoreticallyBasedModel\tHistoricFieldData\tHigh\tAllButRebels\n\
2\tComputationalModel\tCalculatedData\tMedium\tCompetingSchools\n\
1\tStatisticalProcessing\tEducatedGuess\tLow\tEmbryonicField\n\
0\tDefinitions\tUneducatedGuess\tNone\tNoOpinion",

"Code\tModelStructure\tDataInput\tTesting\n\
4\tComprehensive\tReview\tCorroboration\n\
3\tFiniteElementApproximation\tHistoricField\tComparison\n\
2\tTransferFunction\tExperimental\tUncertaintyAnalysis\n\
1\tStatisticalProcessing\tCalculated\tSensitivityAnalysis\n\
0\tDefinitions\tExpertGuess\tNone",

"Code\tDefinitionsAndStandards\tDataCollectionAndAnalysis\tInstitutionalCulture\tReview\n\
5\tNegotiation\tTaskForce\tDialogue\tExternal\n\
4\tScience\tDirectSurvey\tAccomodation\tIndependent\n\
3\tConvenience\tIndirectEstimate\tObedience\tRegular\n\
2\tSymbolism\tEducatedGuess\tEvasion\tOccasional\n\
1\tInertia\tFiat\tNoContact\tNone\n\
0\tUnknown\tUnknown\tUnknown\tUnknown"
         ]
         }
    elif cmd_name == "datasetdata":
        return {"type": "Input", "name": "DatasetData", "template":
            "<Dataset concept #1>\t<Dataset concept #2>\t<Dataset concept #3>\t...",
                "examples": [
"Country\tYear\tWaterConsumption\n\
ES\t2015\t102\n\
ES\t2016\t110\n\
IT\t2015\t130\n\
IT\t2016\t140\n",

"Tech\tScale\tUnitEnergyConsumption\n\
Coal\tMiddle\t1.4\n\
Coal\tLarge\t1.3\n\
Coal\tVeryLarge\t1.2\n\
Nuclear\tLarge\t1.3\n\
Nuclear\tVeryLarge\t1.15\n"
                ]
                }
    else:
        return None


def get_regular_cmd_help(cmd: nexinfosys.Command):
    ctype = str(cmd.cmd_type)
    cmdflds = command_fields.get(cmd.name, None)
    examples = cmd.direct_examples
    files = cmd.files
    return dict(type=ctype,
                name=cmd.allowed_names[0],
                template="\t".join([f.allowed_names[0] for f in cmdflds if "@" not in f.allowed_names[0] and not f.deprecated]),
                examples=[]
                )


@app.route(nis_api_base + "/commands_reference.json", methods=["GET"])
def obtain_commands_reference():
    # Recover InteractiveSession
    isess = deserialize_isession_and_prepare_db_session()
    if isess and isinstance(isess, Response):
        return isess

    d = []
    sequence = [nexinfosys.CommandType.core,
                nexinfosys.CommandType.input,
                nexinfosys.CommandType.analysis,
                nexinfosys.CommandType.metadata,
                nexinfosys.CommandType.convenience,
                nexinfosys.CommandType.misc]
    for ctype in sequence:
        for cmd in commands:
            if cmd.is_v2 and cmd.cmd_type == ctype:
                tmp = get_misc_cmd_help(cmd.name)
                if tmp:
                    d.append(tmp)
                elif command_fields.get(cmd.name, None):
                    d.append(get_regular_cmd_help(cmd))

    return build_json_response([e for e in d if e])

"""
    d = [
        {"type": "External dataset", "name": "Mapping", "template":
            "<Source dimension from external dataset>\t<Target internal taxonomy>\t<Weight (optional, default 1 (many-to-one), <1 for many-to-many mappings)>",
         "examples": [
             "nrg_110a.PRODUCT\tMuSIASEM_EC\n\
             2100\tHeat\n\
             2200\tHeat\n\
             2410\tHeat\n\
             3214\tHeat\n\
             3215\tHeat\n\
             3215\tFeedstock\n\
             3220\tHeat\n\
             3234\tFuel\n\
             3235\tFuel\n\
             3244\tFuel\n\
             3246\tFuel\n\
             3247\tFuel\n\“n-2”“n-2”
             3250\tFeedstock\n\
             3260\tFuel\n\
             3270A\tHeat\n\
             3280\tFeedstock\n\
             3285\tHeat\n\
             4000\tHeat\n\
             5532\tHeat\n\
             5541\tHeat\n\
             5542\tHeat\n\
             55431\tHeat\n\
             55432\tHeat\n\
             5544\tHeat\n\
             5545\tFuel\n\
             5550\tHeat\n\
             6000\tElectricity\n\
             7100\tHeat\n\
             7200\tHeat\n\
             ",
             "nrg_110a.INDIC_NRG\tMuSIASEM_Sector\n\
             B_101300\tES\n\
             B_101825\tMQ\n\
             B_102030\tAFO\n\
             B_102020\tFI\n\
             B_101805\tIS\n\
             B_101810\tNF\n\
             B_101815\tCP\n\
             B_101820\tNM\n\
             B_101830\tFT\n\
             B_101835\tTL\n\
             B_101840\tPPP\n\
             B_101846\tTE\n\
             B_101847\tMA\n\
             B_101851\tWWP\n\
             B_101852\tCO\n\
             B_101853\tNS\n\
             B_102035\tSG"
         ]
         },
        {"type": "External dataset", "name": "Parameters", "template":
            "Name\tValue\tType\tGroup\tDescription",
         "examples": [
             "Name\tValue\tType\tGroup\tDescription\n\
             p1\t3\tnumber\t\tParameter  # 1\n\
             p2\t3.5\tnumber\t\tParameter two"
         ]
         },
        {"type": "Specification", "by_rows": False, "name": "Metadata", "template":
"Case study code\n\
Case study name\n\
Title\n\
Subject, topic and/or keywords\n\
Description\n\
Geographical level\n\
Dimensions\n\
Reference documentation\n\
Authors\n\
Date of elaboration\n\
Temporal situation\n\
Geographical location\n\
DOI\n\
Language\n\
Restriction level\n\
Version", "examples": [
"Case study code\tCS3_R_WEF_P-0.1\n\
Case study name\n\
Title\tSoslaires\n\
Subject, topic and/or keywords\n\
Description\tA small scale system combining Energy, Water and Food\n\
Geographical level\tLocal\n\
Dimensions\tEnergy\tWater\tFood\n\
Reference documentation\n\
Authors\tAna Musicki\tBaltasar Peñate\tTarik Serrrano\n\
Date of elaboration\t2016\n\
Temporal situation\t2016\n\
Geographical location\tGran Canaria\n\
DOI\n\
Language\tEnglish\n\
Restriction level\tPublic\n\
Version\tV0.1"]},
        {"type": "Specification", "name": "Processors", "template":
"Name\tLevel\tFF_TYPE\tVAR\tVALUE\tUNIT\tRELATIVE TO\tUNCERTAINTY\tASSESSMENT\tPEDIGREE\\nMATRIX\tPEDIGREE\tTIME\tGEO\tSCALE\tSOURCE\tCOMMENTS",
         "examples": [
"Name\tLevel\tFF_TYPE\tVAR\tVALUE\tUNIT\tRELATIVE TO\tUNCERTAINTY\tASSESSMENT\tPEDIGREE\\nMATRIX\tPEDIGREE\tTIME\tGEO\tSCALE\tSOURCE\tCOMMENTS\n\
WindFarm\tN-1\tInt_In_Fund\tHA\t660\thours\t\t\t\t\t\tYear\t\t\t\t\n\
WindFarm\tN-1\tInt_In_Fund\tHA_cost\t1800\t€\t\t\t\t\t\t2016\t\t\t\t\n\
WindFarm\tN-1\tInt_Out_Flow\tWindElectricity\t9.28\tGWh\t\t\t\t\t\tYear\t\t\t\t11,8% Energy transformation efficiency from wind to electricity\n\
ElectricGrid\tN\tExt_In_Flow\tGridElectricity\t6.6\tGWh\t\t\t\t\t\tYear\t\t\t\t0.429 M€ income from energy sale"]},
        {"type": "Specification", "name": "Upscale", "template":
"<factor name>\t\n\
<child processor type> / <parent processor type>\t<one or more codes from predefined categories. One or more rows allowed, from this row upwards>\n\
<one or more codes from predefined categories. One or more columns allowed, from this column to the left>\
",
         "examples": [
"LU\tGH\tGH\tOF\n\
Farm / AgrarianRegion\tMCR1\tMCR2\tMCR1\n\
AR1\t0.00\t0.06\t0.94\n\
AR2\t0.15\t0.85\t0.00\n\
AR3\t0.19\t0.77\t0.04\n\
AR4\t0.03\t0.05\t0.92\n\
AR5\t0.00\t0.00\t1.00\n\
AR6\t0.00\t0.87\t0.13"
         ]
         },
        {"type": "Specification", "name": "Structure", "template":
"Origin\tRelation\tDestination\tDestination\tDestination",
         "examples": [
"Origin\tRelation\tDestination\tDestination\tDestination\tDestination\tDestination\tDestination\tDestination\tDestination\n\
WindFarm:WindElectricity\t>\t1/(0.5*p1)>DesalinationPlant:WindElectricity\tElectricGrid\t\t\t\t\t\t\n\
ElectricGrid\t>\tDesalinationPlant:GridElectricity\t\t\t\t\t\t\t\n\
DesalinationPlant:DesalinatedWater\t>\tFarm:BlueWater\t\t\t\t\t\t\t\n\
Farm\t|\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:LU\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:HA\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:IrrigationCapacity\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:BlueWater\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:Agrochemicals\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:Fuel\t>\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:GreenWater\t<\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:MaterialWaste\t<\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:DiffusivePollution\t<\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm:CO2\t<\tCantaloupe\tWatermelon\tTomato\tZucchini\tBeans\tPumpkin\tBanana\tMoringa\n\
Farm\t<\tCantaloupe:Cantaloupe\tWatermelon:Watermelon\tTomato:Tomato\tZucchini:Zucchini\tBeans:Beans\tPumpkin:Pumpkin\tBanana:Banana\tMoringa:Moringa"
         ]
         },
        {"type": "Specification", "name": "Taxonomy_F", "template":
"Code\tDescription\tCode\tDescription\tExpression",
         "examples": [
"Code\tDescription\tCode\tDescription\tExpression\n\
Vegetables\tAll kinds of vegetables\n\
\t\tCantaloupe\n\
\t\tWatermelon\t\n\
\t\tTomato\n\
\t\tZucchini\n\
\t\tBeans\n\
\t\tPumpkin\n\
\t\tBanana\n\
\t\tMoringa"
         ]
         },
        {"type": "Specification", "name": "Pedigree", "template":
"Code\t<Phase name #1>\t<Phase name #2>\t<Phase name #3>\t...",
         "examples": [
"Code\tTheoreticalStructures\tDataInput\tPeerAcceptance\tColleagueConsensus\n\
4\tEstablishedTheory\tExperimentalData\tTotal\tAllButCranks\n\
3\tTheoreticallyBasedModel\tHistoricFieldData\tHigh\tAllButRebels\n\
2\tComputationalModel\tCalculatedData\tMedium\tCompetingSchools\n\
1\tStatisticalProcessing\tEducatedGuess\tLow\tEmbryonicField\n\
0\tDefinitions\tUneducatedGuess\tNone\tNoOpinion",
"Code\tModelStructure\tDataInput\tTesting\n\
4\tComprehensive\tReview\tCorroboration\n\
3\tFiniteElementApproximation\tHistoricField\tComparison\n\
2\tTransferFunction\tExperimental\tUncertaintyAnalysis\n\
1\tStatisticalProcessing\tCalculated\tSensitivityAnalysis\n\
0\tDefinitions\tExpertGuess\tNone",
"Code\tDefinitionsAndStandards\tDataCollectionAndAnalysis\tInstitutionalCulture\tReview\n\
5\tNegotiation\tTaskForce\tDialogue\tExternal\n\
4\tScience\tDirectSurvey\tAccomodation\tIndependent\n\
3\tConvenience\tIndirectEstimate\tObedience\tRegular\n\
2\tSymbolism\tEducatedGuess\tEvasion\tOccasional\n\
1\tInertia\tFiat\tNoContact\tNone\n\
0\tUnknown\tUnknown\tUnknown\tUnknown"
         ]
         },
        {"type": "Specification", "name": "Composition_P", "template":
"Code\tDescription\tCode\tDescription",
         "examples": [
"Code\tDescription\tCode\tDescription\tCode\tDescription\tCode\tDescription\tCode\tDescription\n\
Society\tEncompassess the human realm\n\
\t\tHH\tHousehold Sector\n\
\t\tPW\tPaid Work Sector\n\
\t\t\t\tSG\tService & Government\n\
\t\t\t\tPS\tPrimary & Secondary\n\
\t\t\t\t\t\tBM\tBuilding & Manufacturing\n\
\t\t\t\t\t\tPF\tPrimary flows\n\
\t\t\t\t\t\t\t\tAG\tAgriculture\n\
\t\t\t\t\t\t\t\tEM\tEnergy & Mining"
         ]
         },
        {"type": "Specification", "name": "Taxonomy_C", "template":
"Code\tDescription\tCode\tDescription\tExpression",
         "examples": [
         ]
         },
        {"type": "Specification", "name": "References", "template":
"ref_id\t<list of columns depending on the type reference (bibliographic, geographic, provenance, see examples)>",
         "examples": [
"ref_id\tTitle\tDate\tBoundingBox\tTopicCategory\tDescription\tMetadataPointOfContact\tAnnote\tDataLocation",
"ref_id\tEntry_Type\tAddress\tAnnote\tBookTitle\tChapter\tCrossRef\tEdition\tEditor\tHowPublished\tInstitution\tJournal\tKey\tMonth\tNote\tNumber\tOrganization\tPages\tPublisher\tSchool\tSeries\tTitle\tType\tURL\tVolume\tYear",
"ref_id\tAgentType\tAgent\tActivities\tEntities"
         ]
         },
        {"type": "Specification", "name": "Scale", "template":
"<A matrix having as row starts the origin factor type names, as column headers the target factor type names",
         "examples": []
         },
        {"type": "Analysis", "name": "Indicators", "template":
            "Name\tFormula\tDescription\tBenchmark\tBenchmark\tBenchmark\tBenchmark",
         "examples": []
         }
    ]
"""


@app.route(nis_api_base + "/command_reference.json", methods=["POST"])
def command_help():
    """
    A function for on-line help for a command

    The input comes in a JSON field "content":
    {"command": "<command name"
    }

    :return: A dictionary with the same fields passed in the input dictionary, whose values are the help divided in
        sections: explanation, allowed_values, formal syntax and examples
    """

    # Read request
    command_content_to_validate = request.get_json()
    result, status = comm_help(command_content_to_validate)

    return build_json_response(result, status)


@app.route(nis_api_base + "/command_fields_reference.json", methods=["POST"])
def command_fields_help():
    """
    A function for on-line, field by field help

    The input comes in a JSON field "content":
    {"command": "<command name",
     "fields": ["<field name>", "<field_name>"]
    }

    :return: A dictionary with the same fields passed in the input dictionary, whose values are the help divided in
        sections: explanation, allowed_values, formal syntax and examples
    """

    # Read request
    command_content_to_validate = request.get_json()
    result, status = command_field_help(command_content_to_validate)
    return build_json_response(result, 200 if status else 400)


# @app.route(nis_api_base + "/sources/<id>/databases/<database_id>/datasets/<dataset_id>", methods=["GET"])
# def data_source_database_dataset_query(id, database_id, dataset_id):
#     """
#     This is the most powerful data method, allowing to
#
#     :param id:
#     :param database_id:
#     :param dataset_id:
#     :return:
#     """
#     # Recover InteractiveSession
#     isess = deserialize_isession_and_prepare_db_session()
#     if isess and isinstance(isess, Response):
#         return isess


def data_processes():
    pass


def nusap_data_pedigree():
    pass


def grammars():
    pass


def mappings():
    """
    From an external dataset to internal categories
    :return: 
    """
    pass


def hierarchies():
    a= 6
    pass

# -- Test --


@app.route('/test', methods=['GET'])
@app.route(nis_api_base + '/test', methods=['GET'])
def hello():
    logger.debug("LOG!!!")
    return build_json_response({"hello": "world"})


if __name__ == '__main__':
    # xl = openpyxl.load_workbook("/home/rnebot/Dropbox/nis-internal-tests/issue_report.xlsx", data_only=True)
    # rewrite_xlsx_file(xl)
    # xl.save("/home/rnebot/Downloads/borrame.xlsx")
    # sys.exit(0)

    # from tasks import add
    # from celery.task.control import inspect
    # import time
    # def f():
    #     t = []
    #     for i in range(10):
    #         t.append(add.delay(i, i + 1))
    #     i = inspect()
    #     st = [ti.ready() for ti in t]
    #     while not all(st):
    #         print(f"Completos: {sum(st)}; quedan {len(st)-sum(st)}")
    #         print(i.active())
    #         time.sleep(1)
    #         st = [ti.ready() for ti in t]
    # f()

    # 1) GUNICORN
    # (start REDIS first at localhost:6379. E.g.: docker run --rm --name redis-local -p 6379:6379 redis:alpine)
    #
    # cd ~/AA_MAGIC/nis-nexinfosys
    # export MAGIC_NIS_SERVICE_CONFIG_FILE=/home/rnebot/Dropbox/nis-nexinfosys-config/nis_local.conf
    # gunicorn --bind 0.0.0.0:8080 --workers 3 nexinfosys.restful_service.service_main:app

    # 2) DOCKER. BASIC DEPLOYMENT
    #
    # PREVIOUSLY, COMPILE FRONTEND
    # cd ~/GoogleDrive/AA_MAGIC/nis-frontend
    # npm install
    # rm dist -fr
    # node --max_old_space_size=8192 node_modules/@angular/cli/bin/ng build --prod -c production_local --aot --base-href /nis_client/
    # rm ~/GoogleDrive/AA_MAGIC/nis-nexinfosys/frontend/* -fr
    # cp -r ~/GoogleDrive/AA_MAGIC/nis-frontend/dist/* ~/GoogleDrive/AA_MAGIC/nis-nexinfosys/frontend
    #
    # 2) (continuation) DOCKER COMMANDS (example)
    # docker network create nis-net
    # docker run --rm --name redis-local --net nis-net -p 6379:6379 redis:alpine
    # docker create --name nis-local --net nis-net -p 5000:80 -v /home/rnebot/DATOS/docker_magic_nis:/srv -e MAGIC_NIS_SERVICE_CONFIG_FILE="nis_local_redis_docker.conf" magic-nis

    # cs = CaseStudy()
    # vs1 = CaseStudyVersion()
    # vs1.case_study = cs
    # vs2 = CaseStudyVersion()
    # vs2.case_study = cs
    #
    # lst = [cs, vs1, vs2]
    # d_list = serialize(lst)
    # lst2 = deserialize(d_list)
    # sys.exit(1)
    # >>>>>>>>>> IMPORTANT <<<<<<<<<
    # For debugging in local mode, prepare an environment variable "MAGIC_NIS_SERVICE_CONFIG_FILE", with value "./nis_local.conf"
    # >>>>>>>>>> IMPORTANT <<<<<<<<<

    # >>>>>>>>>> IMPORTANT <<<<<<<<<
    # "cannot connect to X server" error when remote debugging?
    # Execute "Xvfb :99 -ac -noreset" in the remote server and uncomment the following line
    # os.environ["DISPLAY"] = ":99"
    app.run(host='0.0.0.0',
            debug=True,
            use_reloader=False,  # Avoid loading twice the application
            processes=1,
            threaded=False)  # Default port, 5000
