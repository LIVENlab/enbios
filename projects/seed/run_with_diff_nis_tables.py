import json
from pathlib import Path
import pandas as pd

import yaml
from openpyxl.reader.excel import load_workbook

from enbios.processing.main import Enviro
from enbios2.const import BASE_DATA_PATH

cfg_file_path = (BASE_DATA_PATH / "AlexEnbios1/base.json")
base_output_dir = (BASE_DATA_PATH / "AlexEnbios1/output")

config_data: dict = {}


def base2json():
    """
    convert the yml to json. just need to do that once in the beginning.
    :return:
    """
    d = yaml.load(cfg_file_path.read_text(encoding="utf-8"), Loader=yaml.FullLoader)
    json.dump(d, open(cfg_file_path.parent / "base.json", "w", encoding="utf-8"), indent=2, ensure_ascii=False)


def update_nis_table(nis_file_path: Path) -> None:
    # TODO: implement
    # 1. read the nis file
    # 2. update the nis table (in base.json:nis_table)
    # open excel file on specific sheet
    workbook = load_workbook(filename=config_data["nis_table"])
    sheet = workbook["Interfaces"]
    assert sheet["L1"].value == "Value"
    assert sheet["C1"].value == "Interface"

    # row 2 cuz, 1 are the header and 2. is the reference interface, which is always 1
    for row in sheet.iter_rows(values_only=True, min_row=3):
        interface = row[2]
        value = row[11]
        # make a lookup in the nis file
    workbook.save(filename=config_data["nis_table"])


def get_sorted_nis_update_files() -> list[Path]:
    # todo: implement
    pass
    return [Path("/")]


def update_config_field(field, value):
    """
    change a field in the config. used to change the output directory
    :param field:
    :param value:
    """
    if field not in config_data:
        raise Exception(f"Field {field} not found in config file")
    config_data[field] = value
    json.dump(config_data, cfg_file_path.open("w", encoding="utf-8"), indent=2, ensure_ascii=False)


def update_output_path(scenario_index: int) -> Path:
    """
    create an output path for a scenario
    :param scenario_index:
    :return:
    """
    output_path = base_output_dir / f"scenario_{scenario_index}"
    output_path.mkdir(exist_ok=True, parents=True)
    return output_path


def run_with_diff_nis_tables():
    global config_data
    t = Enviro()
    t.set_cfg_file_path(cfg_file_path.as_posix())

    config_data = t._cfg

    nis_files = get_sorted_nis_update_files()

    for index, nis_file in enumerate(nis_files):
        print("Running scenario", index)
        update_nis_table(nis_file)
        update_config_field("output_directory", update_output_path(index).as_posix())
        # t.compute_indicators_from_base_and_simulation(keep_fragment_file=False)


run_with_diff_nis_tables()
