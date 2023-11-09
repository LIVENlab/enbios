import json
import os
from csv import DictReader
from pathlib import Path
from typing import Generator, Union, Optional

try:
    from yaml import CLoader as Loader
except ImportError:
    from yaml import Loader
import openpyxl
import xmltodict as xmltodict

import yaml

from enbios.const import BASE_DATA_PATH
from enbios.generic.enbios2_logging import get_logger

logger = get_logger(__name__)


def read_data(path: Path, config: Optional[dict] = None):
    """
    Read data from file. Formats supported: json, csv, excel
    - json is read straight into a dict
    - csv is read into a DictReader object
    - excel is read into a dict of sheet names and lists of rows

    :return:
    """
    if path.suffix == ".json":
        return json.loads(path.read_text(encoding="utf-8"))
    elif path.suffix == ".csv":
        if not config:
            config = {}
        return list(DictReader(path.open(encoding="utf-8"), **config))
    # excel
    elif path.suffix == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return {sheet.title: list(sheet.values) for sheet in workbook.worksheets}
    # xml
    elif path.suffix == ".xml":
        try:
            import xmltodict
        except ImportError:
            raise ImportError("xmltodict not installed")
        return xmltodict.parse(path.read_text(encoding="utf-8"))
    else:
        raise NotImplementedError(f"File format {path.suffix} not supported")


class DataPath(Path):
    _flavour = Path(".")._flavour  # type: ignore

    def __new__(cls, *args, **kwargs):
        return super().__new__(cls, BASE_DATA_PATH, *args, **kwargs)

    def read_data(self, config: Optional[dict] = None):
        """
        Read data from file. Formats supported: json, csv, excel
        - json is read straight into a dict
        - csv is read into a DictReader object
        - excel is read into a dict of sheet names and lists of rows

        :return:
        """
        return read_data(self, config=config)


class ReadPath(Path):
    """
    Checks on instantiation that the file exists
    """

    _flavour = Path(".")._flavour  # type: ignore

    def __new__(cls, *args, **kwargs):
        instance = super().__new__(cls, *args, **kwargs)
        if not instance.exists():
            raise FileNotFoundError(f"File {instance} does not exist")
        return instance

    def read_data(self, config: Optional[dict] = None):
        """
        Read data from file. Formats supported: json, csv, excel
        - json is read straight into a dict
        - csv is read into a DictReader object
        - excel is read into a dict of sheet names and lists of rows

        :return:
        """
        return read_data(self, config=config)

    def iter_data(self) -> Union[dict, Generator[dict, None, Optional[dict]]]:
        if self.suffix == ".json":
            logger.warning("Reading json completely not as iterator")
            return json.loads(self.read_text(encoding="utf-8"))
        elif self.suffix in [".yaml", ".yml"]:
            return yaml.load(self.read_text(encoding="utf-8"), Loader=Loader)
        elif self.suffix == ".csv":
            reader = DictReader(self.open(encoding="utf-8"))
            for row in reader:
                yield row
        elif self.suffix == ".xlsx":
            workbook = openpyxl.load_workbook(self, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    yield {sheet.title: row}
        elif self.suffix == ".xml":
            logger.info("Reading xml completely not as iterator")
            return dict(xmltodict.parse(self.read_text(encoding="utf-8")))
        return None


class ReadDataPath(ReadPath):
    def __new__(cls, *args, **kwargs):
        instance = super().__new__(cls, BASE_DATA_PATH, *args, **kwargs)
        if not instance.exists():
            raise FileNotFoundError(f"File {instance} does not exist")
        return instance


PathLike = Union[str, os.PathLike]
