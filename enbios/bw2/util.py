from pathlib import Path
from typing import Generator, Iterator, Literal, Optional

import bw2data
import bw2io
import openpyxl
from bw2calc import LCA
from bw2data import databases as bw_databases
from bw2data.backends import Activity, ExchangeDataset, ActivityDataset
from bw2data.project import projects as bw_projects
from bw2io import SingleOutputEcospold2Importer
from scipy.sparse import csr_matrix
from tqdm import tqdm

from enbios.generic.files import PathLike
from enbios.base.models import ExperimentHierarchyNodeData


def info_exchanges(act: Activity) -> dict:
    """
    Show exchanges of an activity
    :param act:
    :return:
    """
    exchanges = act.exchanges()
    for exc in exchanges:
        print(exc)
    for exc in act.upstream():
        print(exc)
    return {exc.input: exc for exc in act.exchanges()}


def iter_exchange_by_ids(
    ids: Iterator[int], batch_size: int = 1000
) -> Generator[ExchangeDataset, None, None]:
    """
    Iterate over exchanges by ids
    :param ids:
    :param batch_size:
    :return:
    """
    last_batch: bool = False
    while True:
        batch_ids: list[int] = []
        for i in range(batch_size):
            try:
                batch_ids.append(next(ids))
            except StopIteration:
                last_batch = True
                break
        for exc in ExchangeDataset.select().where(ExchangeDataset.id.in_(batch_ids)):
            yield exc
        if last_batch:
            break


def iter_activities_by_codes(
    codes: Iterator[str], batch_size: int = 1000
) -> Generator[ActivityDataset, None, None]:
    """
    Iterate over activities by codes
    :param codes:
    :param batch_size:
    :return:
    """
    last_batch: bool = False
    while True:
        batch_codes: list[str] = []
        for i in range(batch_size):
            try:
                batch_codes.append(next(codes))
            except StopIteration:
                last_batch = True
                break
        for act in ActivityDataset.select().where(ActivityDataset.code.in_(batch_codes)):
            yield act
        if last_batch:
            break


def get_activity(code: str) -> Activity:
    """
    Get activity by code
    :param code:
    :return:
    """
    activity_ds = ActivityDataset.get_or_none(ActivityDataset.code == code)
    if not activity_ds:
        raise ValueError(f"Activity with code '{code}' does not exist")
    return Activity(activity_ds)


def full_duplicate(activity: Activity, code=None, **kwargs) -> Activity:
    """
    Make a copy of an activity with its upstream exchanges
    (Otherwise, you cannot calculate the lca of the copy)
    :param activity: the activity to copy
    :param code: code of the new activity
    :param kwargs: other data for the copy
    :return: new activity
    """
    activity_copy = activity.copy(code, **kwargs)
    for upstream in activity.upstream():
        upstream.output.new_exchange(
            input=activity_copy, type=upstream["type"], magnitude=upstream.magnitude
        ).save()
    activity_copy.save()
    return activity_copy


def clean_delete(activity: Activity):
    """
    Delete an activity and its upstream exchanges.
    Otherwise, the system will be corrupted
    :param activity: The activity to delete
    """
    for link in activity.upstream():
        link.delete()
    activity.delete()


def report():
    current_ = bw2data.projects.current
    projects = list(bw2data.projects)
    for project in projects:
        print(project)
        bw_projects.set_current(project.name)
        databases = list(bw_databases)
        print(databases)
    bw2data.projects.set_current(current_)


def bw_unit_fix(unit_str: str):
    if unit_str == "kilowatt hour":
        return "kilowatt_hour"
    if unit_str == "unit":
        return "unspecificEcoinventUnit"
    unit_str = unit_str.replace("-","_")
    return unit_str


def delete_db(project_name: str, db_name: str):
    if project_name not in bw2data.projects:
        raise ValueError(f"Project '{project_name}'does not exist")
    bw2data.projects.set_current(project_name)
    if db_name not in bw2data.databases:
        raise ValueError(f"Database '{db_name}'does not exist")
    print("proceeding to delete database ...")
    bw2data.Database(db_name).delete_instance()


def delete_project(project_name: str):
    if project_name == "default":
        print("Not gonna delete project 'default'")
        return
    if project_name not in bw2data.projects:
        print(f"Project '{project_name}'does not exist")
        return
    bw2data.projects.delete_project(project_name, True)


def delete_all_projects():
    resp = input("Are you sure you want to delete all projects? [y]")
    if resp != "y":
        print("cancelled")
        return
    for project in bw2data.projects:
        if project.name != "default":
            bw2data.projects.delete_project(project.name, True)
            print(f"Deleted {project.name}")
    bw2data.projects.purge_deleted_directories()


"""
Following two functions are from bw_utils...
"""


def _check_lca(
    lca: LCA,
    make_calculations: bool = True,
    inventory_name: Literal["inventory", "characterized_inventory"] = "inventory",
):
    if not hasattr(lca, "inventory"):
        if make_calculations:
            print("calculating inventory")
            lca.lci()
        else:
            raise ValueError("Must do lci first")
    if inventory_name == "inventory":
        return

    if not hasattr(lca, "characterization_matrix"):
        print("loading lcia data")
        lca.load_lcia_data()

    if not hasattr(lca, "characterized_inventory"):
        if make_calculations:
            print("calculating lcia")
            lca.lcia()
        else:
            raise ValueError("Must do lcia first")


def split_inventory(
    lca: LCA,
    technosphere_activities: list[int],
    inventory_name: Literal["inventory", "characterized_inventory"] = "inventory",
    make_calculations: bool = True,
) -> csr_matrix:
    """
    Split the results of a lcia calculation into groups. Each group is a list of activities, specified by their ids.
    Calculations of lci and lcia are performed when they are missing and `make_calculations` is set to `True`
    :param lca: bw LCA object
    :param technosphere_activities: list of (technosphere) activity-groups, activities are specified by their 'id'
    :param make_calculations: make lci and lcia calculations if the corresponding matrices are missing
    :return: a list of sparse matrices, which are characterized inventories split by the activity groups.
    score can be calculated by calling `sum()` for any matrix.
    """
    _check_lca(lca, make_calculations, inventory_name)
    inventory = getattr(lca, inventory_name)
    # do matrix multiplication for each final location
    return inventory[:, [lca.dicts.activity[c] for c in technosphere_activities]]


"""
This package has just one function for setting up a brightway project with a evoincent database
"""


def safe_setup_ecoinvent(
    project_name: str,
    ecoinvent_db_path: str,
    db_name: str,
    delete_project: bool = False,
    delete_if_unlinked: bool = True,
):
    """
    Initiate a project with a ecoinvent database
    :param project_name: new project name
    :param ecoinvent_db_path: dataset path of the ecoinvent database
    :param db_name: name of the new database containing the ecoinvent database
    :param delete_project: Delete existing project
    :param delete_if_unlinked: Delete project if linking didn't work
    :return:
    """
    if project_name in bw2data.projects:
        if delete_project:
            bw2data.projects.delete_project(project_name, True)
        else:
            raise KeyError(f"Project '{project_name}' already exists")
    db_path = Path(ecoinvent_db_path)
    spold_files_glob = db_path.glob("*.spold")
    if not db_path.exists:
        raise FileNotFoundError(f"{ecoinvent_db_path} does not exist")
    if not next(spold_files_glob):
        raise KeyError(f"There are no spold files in {ecoinvent_db_path}")
    bw2data.projects.create_project(project_name)
    bw2data.projects.set_current(project_name)
    bw2io.bw2setup()

    imported = SingleOutputEcospold2Importer(ecoinvent_db_path, db_name)
    imported.apply_strategies()
    # print(type(imported))
    if imported.all_linked:
        imported.write_database()
    else:
        if delete_if_unlinked:
            bw2data.projects.delete_project(project_name, True)
        else:
            imported.write_unlinked(f"{db_name}_unlinked")


def update_ecoinvent_activity_code(
    experiment_hierarchy: dict, activity_code_update_file: PathLike, worksheet: str
) -> dict:
    """
     update the codes of an hierarchy, between the 2 ecoinvent versions.
    :return: a new, updated
    """
    hierarchy: ExperimentHierarchyNodeData = ExperimentHierarchyNodeData.model_validate(
        experiment_hierarchy
    )
    activity_code_update_file_ = Path(activity_code_update_file)
    if not activity_code_update_file_.exists():
        raise FileNotFoundError(f"{activity_code_update_file} does not exist")

    activity_codes: dict[str, ExperimentHierarchyNodeData] = {}

    # go through the hierarchy and find the nodes, which have a brightway adapter and a code
    def rec_change_code(node: ExperimentHierarchyNodeData):
        from enbios.bw2.brightway_experiment_adapter import BrightwayAdapter

        if getattr(node, "adapter", None) in [
            BrightwayAdapter.name(),
            BrightwayAdapter.node_indicator(),
        ]:
            code = node.config.get("code", None)
            if not code:
                print(f"node : {node.name} has no code. Nothing to do")
            else:
                activity_codes[code] = node
        if hasattr(node, "children"):
            for child in node.children:
                rec_change_code(child)

    rec_change_code(hierarchy)

    if not activity_code_update_file_.suffix == ".xlsx":
        raise ValueError(f"File {activity_code_update_file} is not supported")

    # load worksheet
    wb = openpyxl.load_workbook(activity_code_update_file_, read_only=True)
    if worksheet not in wb.sheetnames:
        raise ValueError(
            f"Worksheet {worksheet} not found in {activity_code_update_file}"
        )
    ws = wb[worksheet]

    # check the versions, or logging
    first_row = [cell.value for cell in ws[1]]
    versions = list(filter(lambda value: value, first_row))
    if len(versions) != 2:
        raise ValueError(
            f"First row of {activity_code_update_file} must contain 2 versions"
        )
    print(f"Version update: {versions[0]} -> {versions[1]}")

    # find the columns with the codes
    prev_index: Optional[int] = None
    new_index: Optional[int] = None
    activity_name_index: Optional[int] = None
    for idx, cell in enumerate(ws[2]):
        if cell.value == "Activity Name" and not activity_name_index:
            activity_name_index = idx
        if cell.value == "Activity UUID":
            if prev_index is None:
                prev_index = idx
            else:
                new_index = idx

    if prev_index is None or new_index is None:
        raise ValueError("Activity UUID columns not found")

    activities_todo = list(activity_codes.keys())
    for row in tqdm(ws.iter_rows(min_row=3, values_only=True)):
        prev_code = row[prev_index]
        # check if we already replaced it (codes appear twice with different products)
        if prev_code in activity_codes and prev_code in activities_todo:
            activity_codes[prev_code].config["code"] = row[new_index]
            print(
                f"change code of '{row[activity_name_index]}': {prev_code} -> {row[new_index]}"
            )
            activities_todo.remove(prev_code)
            if len(activities_todo) == 0:
                break
    if activities_todo:
        print(f"Following codes have not been found: {activities_todo}")

    return hierarchy.model_dump(exclude_unset=True, exclude_defaults=True)


if __name__ == "__main__":
    print(
        update_ecoinvent_activity_code(
            {
                "name": "cool",
                "aggregator": "sum",
                "children": [
                    {
                        "name": "a1",
                        "adapter": "bw",
                        "config": {"code": "4fe91148-1e26-59dd-91c4-52a70be24882"},
                    }
                ],
            },
            "Correspondence File v3.9.1 - v.3.10-1.xlsx",
            "Cut-off",
        )
    )
