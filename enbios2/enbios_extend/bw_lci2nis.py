import ast
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import bw2data
import bw2data as bd
import openpyxl
import pandas as pd
import pint
from bw2data.backends import Activity, ActivityDataset
from pint import DimensionalityError, UndefinedUnitError

from enbios2.bw2.project_index import set_bw_current_project
from enbios2.const import BASE_DATA_PATH

output_path = BASE_DATA_PATH / "/enbios/bw2nis"

ureg = pint.UnitRegistry()


def long_to_short_unit(long_unit):
    try:
        unit = ureg(long_unit)
        # print(unit)
        short_unit = '{:~}'.format(unit).split()[1:][0]
    except (KeyError, ValueError, Exception) as err:
        return None

    return short_unit
#
#
flow_unit_lookup_data = {}
unit_lookup_data = {}


def easy_unit_lookup(name) -> str:
    global flow_unit_lookup_data
    if not flow_unit_lookup_data:
        flow_unit_lookup_data = json.load((BASE_DATA_PATH / "enbios/bioflow_unit_lookup.json").open())
    unit = flow_unit_lookup_data.get(name)
    if unit:
        return unit
    else:
        return get_unit_lookup(name)
        # print("No unit found for biosphere activity", name, "using 'kg' as default")


def from_manual_lookup(name, unit):
    data = json.load((BASE_DATA_PATH / "enbios/manual_unit_lookup.json").open())
    res_unit = data.get(unit)
    if res_unit:
        return res_unit
    else:
        print(f"No unit found for biosphere activity {name}, {unit} using 'kg' as default. put something for '{unit}' into manual lookup")
        return "kg"


def get_unit_lookup(name, folder: Optional[Path] = None) -> str:
    global unit_lookup_data
    if not folder:
        folder = BASE_DATA_PATH / "temp"
    if not unit_lookup_data:
        if (folder / "units.json").exists():
            unit_lookup_data = json.load((folder / "units.json").open())
        else:
            unit_lookup_data = {}
    unit = unit_lookup_data.get(name)
    if unit:
        # print(name, unit)
        res_unit = long_to_short_unit(unit)
        if not res_unit:
            res_unit = from_manual_lookup(name, unit)
        return res_unit
    print("searching for biosphere activity", name)
    res = bw2data.Database("biosphere3").search(name)
    if res:
        unit = res[0]["unit"]
        flow_unit_lookup_data[name] = unit
        json.dump(flow_unit_lookup_data, (folder / "units.json").open("w"))
        # print(name, unit)
        res_unit =  long_to_short_unit(unit)
        if not res_unit:
            res_unit = from_manual_lookup(name, unit)
        return res_unit

    else:
        print("No unit found for biosphere activity", name, "using 'kg' as default")


# def long_to_short_unit(long_unit):
#     try:
#         unit = ureg(long_unit)
#         short_unit = '{:~}'.format(unit).split()[1:]
#         # print(unit, short_unit)
#     except (KeyError, ValueError, AttributeError, DimensionalityError, UndefinedUnitError) as err:
#         print(long_unit)
#         print(err)
#         short_unit = long_unit  # If the long_unit is not recognized, return it as-is.
#     return short_unit


def export_solved_inventory(activity: Activity, method: tuple[str, ...],
                            out_path: Optional[str] = None,
                            sphere: Optional[str] = "biosphere") -> pd.DataFrame:
    """
    All credits to Ben Portner.
    :param activity:
    :param method:
    :param out_path:
    :return:
    """
    # print("calculating lci")
    lca = activity.lca(method, 1)
    lca.lci()
    array = lca.inventory.sum(axis=1)
    if hasattr(lca, 'dicts'):
        mapping = lca.dicts._dicts[sphere]
    else:
        raise ValueError("No dicts found")
        # mapping = lca.biosphere_dict
    data = []
    for key, row in mapping.items():
        amount = array[row, 0]
        data.append((bd.get_activity(key), row, amount))
    data.sort(key=lambda x: abs(x[2]))

    df = pd.DataFrame([{
        'row_index': row,
        'amount': amount,
        'name': flow.get('name'),
        'unit': easy_unit_lookup(flow.get('name')),
        'categories': str(flow.get('categories'))
    } for flow, row, amount in data])
    if out_path:
        df.to_excel(out_path)
    return df


def interface_type_template():
    return {
        "InterfaceTypeHierarchy": "LCI",
        "InterfaceType": None,
        "Sphere": None,
        "RoegenType": "Flow",
        "ParentInterfaceType": "",
        "Formula": "",
        "Description": "",
        "Unit": None,
        "OppositeSubsystemType": None,
        "Attributes": "",
        "@EcoinventName": None
    }


def base_processors_template():
    return {
        "ProcessorGroup": "NewStructurals",
        "Processor": None,
        "ParentProcessor": "",
        "SubsystemType": "",
        "System": "",
        "FunctionalOrStructural": "Structural",
        "Accounted": "No",
        "Stock": "",
        "Description": "",
        "GeolocationRef": "",
        "GeolocationCode": "",
        "GeolocationLatLong": "",
        "Attributes": "",
        "@EcoinventName": None,
        "@EcoinventFilename": None,
        "@EcoinventCarrierName": None,
        "@region": ""
    }


def interfaces_template():
    return {
        "Processor": None,
        "InterfaceType": None,
        "Interface": None,
        "Sphere": "",
        "RoegenType": "",
        "Orientation": None,
        "OppositeSubsystemType": "",
        "GeolocationRef": "",
        "GeolocationCode": "",
        "I@compartment": None,
        "I@subcompartment": None,
        "Value": None,
        "Unit": "",
        "RelativeTo": None,
        "Uncertainty": "",
        "Assessment": "",
        "PedigreeMatrix": "",
        "Pedigree": "",
        "Time": "Year",
        "Source": None,
        "NumberAttributes": "",
        "Comments": "",
    }


def get_nis_name(original_name):
    """
    Convert the original_name to a name valid for NIS
    from nexinfosys

    :param original_name:
    :return:
    """
    if not isinstance(original_name, str):
        original_name = str(original_name)
    if original_name.strip() == "":
        return ""
    else:
        prefix = original_name[0] if original_name[0].isalpha() else "_"
        remainder = original_name[1:] if original_name[0].isalpha() else original_name

        return prefix + re.sub("[^0-9a-zA-Z_]", "_", remainder)


@dataclass
class NisSheetDfs:
    interface_types_df: pd.DataFrame
    bare_processors_df: pd.DataFrame
    interfaces_df: pd.DataFrame


def read_exising_nis_file(file_path) -> NisSheetDfs:
    """
    read sheets and convert to dataframes
    :param file_path:
    :return:
    """
    with pd.ExcelFile(file_path) as xls:
        interface_types_df = pd.read_excel(xls, "InterfaceTypes", header=0)
        bare_processors_df = pd.read_excel(xls, "BareProcessors", header=0)
        interfaces_df = pd.read_excel(xls, "Interfaces", header=0)
    return NisSheetDfs(interface_types_df, bare_processors_df, interfaces_df)


def insert_activity_processor(activity: Activity, nis_dataframes: NisSheetDfs, lci_result: pd.DataFrame,
                              new_file_loc: Path):
    # insert new interface_types
    unique_interface_types = set(nis_dataframes.interface_types_df["@EcoinventName"].unique())
    # print(len(unique_interface_types))
    lci_result = lci_result[lci_result["amount"] > 0]
    lci_interface_types = set(lci_result["name"].unique())
    # difference between the two sets
    missing_interface_types = lci_interface_types - unique_interface_types
    # get the rows, where the interface type is missing
    rows_to_add = lci_result[lci_result["name"].isin(missing_interface_types)]

    new_workbook = openpyxl.Workbook()
    new_workbook.remove(new_workbook["Sheet"])
    new_workbook.create_sheet("InterfaceTypes")

    interface_type_new_rows = []
    for new_type in rows_to_add.iterrows():
        # print(new_type)
        row = new_type[1]

        new_row = {**interface_type_template(), **{
            "InterfaceType": get_nis_name(row["name"]),
            "Sphere": "Biosphere",
            "Unit": easy_unit_lookup(row["name"]),
            # opposite = "Environment" if props["sphere"].lower() == "biosphere" else ""
            "OppositeSubsystemType": "Environment",
            "@EcoinventName": row["name"]
        }}
        interface_type_new_rows.append(new_row)

    # insert the new rows into the new sheet
    new_sheet = new_workbook["InterfaceTypes"]
    fieldnames = list(interface_type_new_rows[0].keys())
    new_sheet.append(fieldnames)
    for row in interface_type_new_rows:
        new_sheet.append((row[k] for k in fieldnames))

    # save the workbook
    # new_workbook.save(new_file_loc.as_posix())

    unique_bareprocessors = set(nis_dataframes.bare_processors_df["Processor"].unique())
    # # difference between the two sets
    # missing_interface_types = lci_interface_types - unique_interface_types
    # # get the rows, where the interface type is missing
    # rows_to_add = lci_result[lci_result["name"].isin(missing_interface_types)]

    bareprocessor_new_rows = []

    if not (carrier_name := activity.get("carrierName")):
        print(f"activity data of {activity['name']} does not contain 'carrierName'. Setting it to electricity")

    if activity["name"] not in unique_bareprocessors:
        new_row = {**base_processors_template(), **{
            "Processor": get_nis_name(activity["name"]),
            "@EcoinventName": activity["name"],
            "@EcoinventFilename": "no-file",
            "@EcoinventCarrierName": carrier_name
        }}
        bareprocessor_new_rows.append(new_row)

    new_interface_rows = []
    # print(lci_result)

    new_sheet = new_workbook.create_sheet("BareProcessors")
    fieldnames = list(bareprocessor_new_rows[0].keys())
    new_sheet.append(fieldnames)
    for row in bareprocessor_new_rows:
        new_sheet.append((row[k] for k in fieldnames))

    new_workbook.save(new_file_loc.as_posix())

    # and add the first row for it
    main_name = get_nis_name("_".join(activity._data["synonyms"]))
    main_row = {
        "Orientation": "Output",
    }

    # new_interface_rows.append(main_row)

    for row in lci_result.iterrows():
        # print(row)
        data = row[1]
        categories = ast.literal_eval(data["categories"])
        compartment = categories[0]
        subcompartment = categories[1] if len(categories) > 1 else "unspecified"
        new_interface_rows.append({**interfaces_template(),
                                   "Processor": get_nis_name(activity["name"]),
                                   "InterfaceType": get_nis_name(data["name"]),
                                   "Interface": get_nis_name(data["name"] + "_" + compartment + "_" + subcompartment),
                                   "Orientation": "Input",
                                   "I@compartment": compartment,
                                   "I@subcompartment": subcompartment,
                                   "Value": data["amount"],
                                   "Unit": None,
                                   "RelativeTo": main_name,  #
                                   "Source": "BW",
                                   })

    new_sheet = new_workbook.create_sheet("Interfaces")
    fieldnames = list(new_interface_rows[0].keys())
    new_sheet.append(fieldnames)
    for row in new_interface_rows:
        new_sheet.append((row[k] for k in fieldnames))

    new_workbook.save(new_file_loc.as_posix())


# if __name__ == "__main__":
#     projects = bd.projects
#
#     set_bw_current_project(version="3.9.1", system_model="cutoff")
#
#     # generate = False
#     # if generate:
#     #     bi.bw2setup()
#     #     if "ei39" not in bd.databases:
#     #         im = SingleOutputEcospold2Importer(
#     #             (BASE_DATA_PATH / f"ecoinvent/ecoinvent 3.9_cutoff_ecoSpold02/datasets").as_posix(),
#     #             "ei39")
#     #     im.apply_strategies()
#     #     if im.statistics()[2] == 0:
#     #         im.write_database()
#     #     else:
#     #         print("unlinked exchanges")
#     # #
#     # activities = bd.Database("cutoff391",).search("heat and power co-generation, oil", filter={"location": "PT"})
#     # activity = activities[0]
#     # print(activity._document.code)
#     activity_ds = ActivityDataset.select().where(ActivityDataset.code == "a8fe0b37705fe611fac8004ca6cb1afd")
#     activity = Activity(activity_ds[0])
#     print(activity._document.name)
#     #
#     base_folder = BASE_DATA_PATH / "nis_tests"
#     nis_file = base_folder / "output.xlsx"
#     #
#     dataframes = read_exising_nis_file(nis_file)
#     lci_result = export_solved_inventory(activity,
#                                          ('CML v4.8 2016', 'acidification',
#                                           'acidification (incl. fate, average Europe total, A&B)'),
#                                          "test.xlsx")
#
#     insert_activity_processor(activity, dataframes, lci_result, base_folder / "output2.xlsx")
