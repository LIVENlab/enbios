import json
from csv import DictReader
from pathlib import Path
from typing import Generator, Optional

import openpyxl
import xmltodict as xmltodict

import yaml

from enbios2.const import BASE_DATA_PATH
from enbios2.generic.enbios2_logging import get_logger

logger = get_logger(__file__)


class DataPath(Path):
    _flavour = Path('.')._flavour

    def __new__(cls, *args, **kwargs):
        return super().__new__(cls, BASE_DATA_PATH, *args, **kwargs)


class ReadPath(Path):
    """
    Checks on instantiation that the file exists
    """
    _flavour = Path('.')._flavour

    def __new__(cls, *args, **kwargs):
        instance = super().__new__(cls, *args, **kwargs)
        if not instance.exists():
            raise FileNotFoundError(f"File {instance} does not exist")
        return instance

    def read_data(self, config: Optional[dict] = None):
        """
        Read data from file. Formats supported: json, csv, excel
        - json is read straight into a dict
        - csv is read into a DictReader object
        - excel is read into a dict of sheet names and lists of rows

        :return:
        """
        if self.suffix == ".json":
            return json.loads(self.read_text(encoding="utf-8"))
        elif self.suffix == ".csv":
            return list(DictReader(self.open(encoding="utf-8"), **config))
        # excel
        elif self.suffix == ".xlsx":
            workbook = openpyxl.load_workbook(self, read_only=True, data_only=True)
            return {sheet.title: list(sheet.values) for sheet in workbook.worksheets}
        # xml
        elif self.suffix == ".xml":
            try:
                import xmltodict
            except ImportError:
                raise ImportError("xmltodict not installed")
            return xmltodict.parse(self.read_text(encoding="utf-8"))

    def iter_data(self) -> Generator[dict, None, dict]:
        if self.suffix == ".json":
            logger.warning("Reading json completely not as iterator")
            return json.loads(self.read_text(encoding="utf-8"))
        elif self.suffix in [".yaml", ".yml"]:
            yaml.load(self.read_text(encoding="utf-8"))
        elif self.suffix == ".csv":
            reader = DictReader(self.open(encoding="utf-8"))
            for row in reader:
                yield row
        elif self.suffix == ".xlsx":
            workbook = openpyxl.load_workbook(self, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    yield {sheet.title: row}
        elif self.suffix == ".xml":
            logger.info("Reading xml completely not as iterator")
            return dict(xmltodict.parse(self.read_text(encoding="utf-8")))


class ReadDataPath(ReadPath):
    def __new__(cls, *args, **kwargs):
        instance = super().__new__(cls, BASE_DATA_PATH, *args, **kwargs)
        if not instance.exists():
            raise FileNotFoundError(f"File {instance} does not exist")
        return instance
